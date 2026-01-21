import os
import csv
import zipfile
import logging
import traceback
import sys
import queue
import shutil
from datetime import datetime, date, time, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading

import calc
from openpyxl import load_workbook
from tkcalendar import DateEntry

import db
import report

try:
    import winsound
except ImportError:
    winsound = None
try:
    import requests
except ImportError:
    requests = None

DATE_FMT = "YYYY-MM-DD"
TIME_FMT = "HH:MM"
KEEPALIVE_SECONDS = 300
REGIONS = ["Ankara", "Izmir", "Bursa", "Istanbul"]
VIEW_REGIONS = ["Tum Bolgeler"] + REGIONS
DEFAULT_OIL_INTERVAL_KM = 14000
DEFAULT_OIL_SOON_KM = 2000
LOG_DIR = os.path.join(os.path.dirname(db.DB_DIR), "logs")
LOG_PATH = os.path.join(LOG_DIR, "rainstaff.log")

VEHICLE_CHECKLIST = [
    ("body_dent", "Govde ezik/cizik"),
    ("paint_damage", "Boya hasari"),
    ("interior_clean", "Ic temizligi"),
    ("smoke_smell", "Sigara kokusu"),
    ("tire_condition", "Lastik durumu"),
    ("lights", "Far/stop/sinyal"),
    ("glass", "Camlar"),
    ("warning_lamps", "Ikaz lambalari"),
    ("water_level", "Su seviyesi"),
]

EMP_HEADER_ALIASES = {
    "full_name": ["ad soyad", "adsoyad", "calisan", "calisan adi", "name", "full_name"],
    "identity_no": ["tckn", "tc", "tc kimlik", "identity", "identity_no"],
    "department": ["departman", "department"],
    "title": ["unvan", "title"],
}

TS_HEADER_ALIASES = {
    "employee": ["calisan", "ad soyad", "employee", "full_name", "name"],
    "work_date": ["tarih", "date", "work_date"],
    "start_time": ["giris", "start", "start_time"],
    "end_time": ["cikis", "end", "end_time"],
    "break_minutes": ["mola", "mola dk", "mola dakika", "break", "break_minutes"],
    "is_special": ["ozel gun", "ozel", "resmi tatil", "special", "is_special"],
    "notes": ["not", "notes", "aciklama"],
}


def normalize_header(value):
    """Lowercase and strip spaces/underscores for loose header matching."""
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def build_header_aliases(alias_config):
    """Expand alias lists into normalized lookup sets per target column."""
    result = {}
    for target, aliases in alias_config.items():
        normalized = {normalize_header(target)}
        for alias in aliases:
            normalized.add(normalize_header(alias))
        result[target] = normalized
    return result


def days_until(date_str):
    """Return days from today to the given ISO date (positive = future, negative = past)."""
    if not date_str:
        return None
    try:
        target = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    return (target - datetime.now().date()).days


def normalize_date_value(value):
    """Import icin flexible tarih normalizasyonu; Excel float veya string kabul eder."""
    if value is None or value == "":
        raise ValueError("Tarih bos olamaz.")
    if isinstance(value, str):
        return normalize_date(value)
    if isinstance(value, (int, float)):
        try:
            dt = datetime.fromordinal(int(value) + 693594)
            return dt.strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            raise ValueError(f"Tarih formati gecersiz: {value}")
    raise ValueError(f"Tarih formati gecersiz: {value}")


def normalize_time_value(value):
    """Import icin flexible saat normalizasyonu."""
    if value is None or value == "":
        raise ValueError("Saat bos olamaz.")
    if isinstance(value, str):
        return normalize_time(value)
    if isinstance(value, (int, float)):
        try:
            total_minutes = int(round(value * 24 * 60))
            hours = (total_minutes // 60) % 24
            minutes = total_minutes % 60
            return f"{hours:02d}:{minutes:02d}"
        except (ValueError, OverflowError):
            raise ValueError(f"Saat formati gecersiz: {value}")
    raise ValueError(f"Saat formati gecersiz: {value}")


def parse_bool(value):
    """String veya int boolean'a cevir."""
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    if isinstance(value, str):
        return value.lower().strip() in {"1", "true", "yes", "evet", "e"}
    return bool(value)


def split_display_name(display, regions):
    """`Adi (Region)` formatindan baz adi ve bolgeyi ayir."""
    text = str(display or "").strip()
    if not text:
        return "", None
    if text.endswith(")") and "(" in text:
        base, suffix = text.rsplit("(", 1)
        base = base.strip()
        region = suffix[:-1].strip()  # sondaki ) kaldir
        if region in regions:
            return base, region
        if region == "-":  # belirsiz/boş
            return base, None
    return text, None


def week_start_from_date(value):
    """Verilen tarihi ait haftanin pazartesi baslangicina cek."""
    iso = normalize_date(str(value))
    d = datetime.strptime(iso, "%Y-%m-%d").date()
    start = d - timedelta(days=d.weekday())  # Pazartesi
    return start.strftime("%Y-%m-%d")


def week_end_from_start(week_start):
    """Hafta baslangicindan pazar gununu uret."""
    d = datetime.strptime(week_start, "%Y-%m-%d").date()
    return (d + timedelta(days=6)).strftime("%Y-%m-%d")


def normalize_time_in_var(var):
    """StringVar icindeki saati normalize eder; hatada eski degeri korur."""
    try:
        normalized = normalize_time(var.get())
        var.set(normalized)
    except Exception:
        pass


def ensure_app_dirs():
    data_dir = os.path.join(os.path.dirname(__file__), "data")
    if not os.path.isdir(data_dir):
        os.makedirs(data_dir, exist_ok=True)


def setup_logging():
    os.makedirs(LOG_DIR, exist_ok=True)
    logger = logging.getLogger("rainstaff")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # Konsola da yaz (komut penceresinde anlık görmek için)
    try:
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
    except Exception:
        # Konsol eklenemese de dosyaya yazmaya devam etsin
        pass

    def handle_uncaught(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            return
        logger.error("Unhandled exception", exc_info=(exc_type, exc_value, exc_traceback))

    sys.excepthook = handle_uncaught
    return logger


class LogQueueHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue
        self.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

    def emit(self, record):
        try:
            msg = self.format(record)
            self.log_queue.put(msg)
        except Exception:
            pass


def parse_int(value, default=0):
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def parse_float(value, default=0.0):
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def normalize_date(value):
    if value is None or value == "":
        raise ValueError("Tarih bos olamaz.")
    value = str(value).strip()
    if not value:
        raise ValueError("Tarih bos olamaz.")
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError("Tarih formati gecersiz. Ornek: 2026-01-05 veya 05.01.2026")


def normalize_time(value):
    if value is None:
        raise ValueError("Saat bos olamaz.")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)) and 0 <= value < 1:
        total_minutes = int(round(value * 24 * 60))
        hours = (total_minutes // 60) % 24
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    text = str(value).strip()
    if not text:
        raise ValueError("Saat bos olamaz.")
    text = text.replace(".", ":")
    if text.isdigit() and len(text) in (3, 4):
        if len(text) == 3:
            text = "0" + text
        return f"{text[:2]}:{text[2:]}"
    if ":" in text:
        parts = text.split(":")
        if len(parts) >= 2:
            return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    raise ValueError("Saat formati gecersiz. Ornek: 09:30")


def normalize_vehicle_status(status):
    """Araç muayene durumunu normalize et (Olumsuz/Olumlu/Belirsiz)"""
    if status is None:
        return "Belirsiz"
    text = str(status).strip().lower()
    if "olumsuz" in text or "bad" in text or "0" in text or "no" in text:
        return "Olumsuz"
    if "olumlu" in text or "good" in text or "ok" in text or "1" in text or "yes" in text:
        return "Olumlu"
    return "Belirsiz"


EMP_HEADER_MAP = build_header_aliases(EMP_HEADER_ALIASES)
TS_HEADER_MAP = build_header_aliases(TS_HEADER_ALIASES)


def map_headers(header_row, header_map):
    mapping = {}
    for idx, value in enumerate(header_row):
        key = normalize_header(value)
        for target, aliases in header_map.items():
            if key in aliases:
                mapping[target] = idx
    return mapping


def load_tabular_file(path):
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.get_dialect("excel")
            reader = csv.reader(handle, dialect)
            for row in reader:
                rows.append(row)
    else:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    return [row for row in rows if any(cell is not None and str(cell).strip() for cell in row)]


def create_labeled_entry(parent, label, textvariable, width=24):
    frame = ttk.Frame(parent)
    ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=(0, 8))
    ttk.Entry(frame, textvariable=textvariable, width=width).pack(side=tk.LEFT)
    return frame


def create_labeled_date(parent, label, textvariable, width=12):
    frame = ttk.Frame(parent)
    ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=(0, 8))
    date_entry = DateEntry(frame, textvariable=textvariable, width=width, date_pattern="yyyy-mm-dd")
    date_entry.pack(side=tk.LEFT)
    return frame, date_entry


def create_time_entry(parent, label, textvariable, width=8):
    frame = ttk.Frame(parent)
    ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=(0, 8))
    entry = ttk.Entry(frame, textvariable=textvariable, width=width)
    entry.pack(side=tk.LEFT)
    return frame


def set_time_vars(time_value, textvariable):
    try:
        normalized = normalize_time(time_value)
    except ValueError:
        return
    textvariable.set(normalized)


def clear_date_entry(entry):
    try:
        entry.delete(0, tk.END)
    except Exception:
        pass


def ensure_logo_asset(path):
    if os.path.isfile(path):
        return
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return
    width, height = 320, 80
    img = Image.new("RGB", (width, height), "#e7eef6")
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle([8, 8, width - 8, height - 8], radius=18, fill="#cfe1f2")
    # Minimalist rain drop icon
    drop_x, drop_y = 28, 18
    draw.polygon(
        [(drop_x + 16, drop_y), (drop_x, drop_y + 24), (drop_x + 16, drop_y + 46), (drop_x + 32, drop_y + 24)],
        fill="#2f6fed",
        outline="#2a5fd1",
    )
    draw.ellipse([drop_x + 6, drop_y + 24, drop_x + 26, drop_y + 44], fill="#2f6fed", outline="#2a5fd1")
    draw.ellipse([drop_x + 14, drop_y + 10, drop_x + 20, drop_y + 16], fill="#bfe0ff")
    text = "RAINSTAFF"
    try:
        font = ImageFont.truetype("Consola.ttf", 22)
    except Exception:
        font = ImageFont.load_default()
    draw.text((84, 22), text, fill="#0f2438", font=font)
    try:
        sub_font = ImageFont.truetype("Consola.ttf", 14)
    except Exception:
        sub_font = ImageFont.load_default()
    draw.text((86, 48), "PUANTAJ", fill="#425466", font=sub_font)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    img.save(path)


# Rehber kaldırıldı - Modern ERP tasarımıyla değiştirildi


def load_logo_image(path, target_height=48):
    try:
        from PIL import Image, ImageTk
    except Exception:
        Image = None
        ImageTk = None

    if Image and ImageTk:
        try:
            img = Image.open(path)
            ratio = target_height / float(img.height)
            target_width = int(img.width * ratio)
            img = img.resize((target_width, target_height), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None
    try:
        return tk.PhotoImage(file=path)
    except Exception:
        return None


class PuantajApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Rainstaff ERP - Puantaj Yönetimi")
        self.geometry("1280x800")
        self.minsize(1024, 720)

        self.logger = setup_logging()
        self.report_callback_exception = self._handle_tk_exception
        self.log_queue = queue.Queue()
        self.log_handler = LogQueueHandler(self.log_queue)
        if self.logger:
            self.logger.addHandler(self.log_handler)

        self.current_user = None
        self.current_region = None
        self.is_admin = False

        self.settings = db.get_all_settings()
        entry_region = self.settings.get("admin_entry_region", "Ankara")
        view_region = self.settings.get("admin_view_region", "Tum Bolgeler")
        if view_region == "ALL":
            view_region = "Tum Bolgeler"
        self.admin_entry_region_var = tk.StringVar(value=entry_region)
        self.admin_view_region_var = tk.StringVar(value=view_region)
        self.employee_map = {}
        self.employee_display_names = []
        self.employee_details = {}
        self.vehicle_map = {}
        self.driver_map = {}
        self.driver_display_names = []
        self.fault_map = {}
        self.service_visit_map = {}
        self.shift_template_map = {}
        self.status_var = tk.StringVar()
        self.ts_original = None
        self.ts_editing_id = None
        self.vehicle_original_plate = None
        self._tab_loaded = {}

        if not self._login_prompt():
            self.destroy()
            return

        self._show_loading("Yukleniyor...")
        self.after(10, self._finish_startup)

    def _handle_tk_exception(self, exc, val, tb):
        if self.logger:
            self.logger.error("Tkinter callback error", exc_info=(exc, val, tb))
        messagebox.showerror("Hata", "Beklenmeyen bir hata olustu. Log kaydi alindi.")

    def _log_action(self, action, detail=""):
        if not self.logger:
            return
        user = self.current_user or "unknown"
        region = self.current_region or "ALL"
        suffix = f" | {detail}" if detail else ""
        self.logger.info("ACTION: %s | user=%s | region=%s%s", action, user, region, suffix)

    def _finish_startup(self):
        self.after(10, self._startup_step_prepare)
        if self.logger:
            self.logger.info("Uygulama basladi")

    def _startup_step_prepare(self):
        self.after(10, self._startup_step_style)

    def _startup_step_style(self):
        self._configure_style()
        self.after(10, self._startup_step_ui)

    def _startup_step_ui(self):
        self._build_ui()
        self.after(10, self._startup_step_data)

    def _startup_step_data(self):
        self._load_tab_data(self.tab_employees)
        self._start_keepalive()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._hide_loading()

    def _show_loading(self, text):
        overlay = tk.Toplevel(self)
        overlay.title("Yükleniyor")
        overlay.geometry("360x200")
        overlay.resizable(False, False)
        overlay.configure(bg="#1E1E1E")  # Koyu arka plan
        overlay.transient(self)
        overlay.grab_set()
        overlay.protocol("WM_DELETE_WINDOW", lambda: None)

        # Koyu card
        card = tk.Frame(overlay, bg="#2A2A2A", highlightbackground="#3A3A3A", highlightthickness=1)
        card.place(x=16, y=16, relwidth=1, relheight=1, width=-32, height=-32)

        # Logo
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "rainstaff_logo_1.png")
        logo_img = load_logo_image(logo_path, target_height=40)
        if logo_img:
            self._loading_logo = logo_img
            tk.Label(card, image=logo_img, bg="#2A2A2A").pack(pady=(32, 16))
        else:
            # Silik altın logo
            tk.Label(card, text="RAINSTAFF", bg="#2A2A2A", fg="#C9A961", 
                    font=("Segoe UI", 14, "bold")).pack(pady=(32, 16))

        # Yükleniyor metni - silik gri
        tk.Label(card, text=text, bg="#2A2A2A", fg="#B0B0B0", 
                font=("Segoe UI", 10)).pack(pady=(0, 20))

        # Spinner - silik mavi
        spinner = tk.Canvas(card, width=40, height=40, bg="#2A2A2A", highlightthickness=0)
        spinner.pack(pady=(0, 24))

        # Silik mavi ring
        arc = spinner.create_arc(2, 2, 38, 38, start=0, extent=280, style="arc", 
                                width=3, outline="#5B9BD5")

        def step(angle=0):
            if getattr(self, "_loading_overlay", None) is None:
                return
            spinner.itemconfigure(arc, start=angle)
            overlay.after(20, step, (angle + 12) % 360)

        step()
        self._loading_overlay = overlay
        self._loading_spinner = spinner

    def _hide_loading(self):
        if hasattr(self, "_loading_spinner"):
            self._loading_spinner = None
        if hasattr(self, "_loading_overlay"):
            try:
                self._loading_overlay.grab_release()
                self._loading_overlay.destroy()
            except Exception:
                pass
            self._loading_overlay = None

    def _start_keepalive(self):
        self._keepalive_stop = threading.Event()
        thread = threading.Thread(target=self._keepalive_worker, daemon=True)
        thread.start()

    def _keepalive_worker(self):
        while not self._keepalive_stop.wait(KEEPALIVE_SECONDS):
            if requests is None:
                continue
            settings = db.get_all_settings()
            enabled = settings.get("sync_enabled") == "1"
            sync_url = settings.get("sync_url", "").strip()
            token = settings.get("sync_token", "").strip()
            if not enabled or not sync_url:
                continue
            try:
                headers = {"X-API-KEY": token} if token else {}
                url = sync_url.rstrip("/") + "/health"
                requests.get(url, headers=headers, timeout=6)
            except Exception:
                pass

    def _on_close(self):
        if hasattr(self, "_keepalive_stop"):
            self._keepalive_stop.set()
        self.destroy()

    def _login_prompt(self):
        dialog = tk.Toplevel(self)
        dialog.title("Giris")
        dialog.resizable(False, False)
        dialog.configure(bg="#1E1E1E")  # Koyu arka plan
        dialog.transient(self)
        dialog.grab_set()

        # Card görünümü için sabit boyut ve merkezleme
        dialog.geometry("420x460")
        dialog.update_idletasks()
        if self.winfo_ismapped():
            px = self.winfo_rootx()
            py = self.winfo_rooty()
            pw = self.winfo_width()
            ph = self.winfo_height()
            dx = px + (pw - 420) // 2
            dy = py + (ph - 460) // 2
            dialog.geometry(f"420x460+{dx}+{dy}")

        card = tk.Frame(dialog, bg="#2A2A2A", highlightbackground="#3A3A3A", highlightthickness=1)
        card.place(x=16, y=16, relwidth=1, relheight=1, width=-32, height=-32)

        logo_path = os.path.join(os.path.dirname(__file__), "assets", "rainstaff_logo_1.png")
        logo_img = load_logo_image(logo_path, target_height=48)
        if logo_img:
            self._login_logo = logo_img
            tk.Label(card, image=logo_img, bg="#2A2A2A").pack(pady=(48, 12))
        else:
            tk.Label(card, text="RAINSTAFF", bg="#2A2A2A", fg="#C9A961",
                     font=("Segoe UI", 16, "bold")).pack(pady=(48, 12))

        tk.Label(card, text="Giris Yapin", bg="#2A2A2A", fg="#E0E0E0",
                 font=("Segoe UI", 16, "bold")).pack(pady=(0, 24))

        form_frame = tk.Frame(card, bg="#2A2A2A")
        form_frame.pack(padx=56, fill=tk.X)

        username_var = tk.StringVar()
        password_var = tk.StringVar()

        tk.Label(form_frame, text="Kullanici Adi", bg="#2A2A2A", fg="#5B9BD5",
                 font=("Segoe UI", 10, "bold"), anchor="w").pack(fill=tk.X, pady=(0, 6))
        username_entry = tk.Entry(form_frame, textvariable=username_var,
                                  font=("Segoe UI", 11), relief="solid", bd=1,
                                  bg="#363636", fg="#E0E0E0",
                                  insertbackground="#5B9BD5",
                                  highlightthickness=1, highlightbackground="#454545")
        username_entry.pack(fill=tk.X, ipady=10)

        tk.Label(form_frame, text="Sifre", bg="#2A2A2A", fg="#5B9BD5",
                 font=("Segoe UI", 10, "bold"), anchor="w").pack(fill=tk.X, pady=(24, 6))
        password_entry = tk.Entry(form_frame, textvariable=password_var, show="●",
                                  font=("Segoe UI", 11), relief="solid", bd=1,
                                  bg="#363636", fg="#E0E0E0",
                                  insertbackground="#5B9BD5",
                                  highlightthickness=1, highlightbackground="#454545")
        password_entry.pack(fill=tk.X, ipady=10)

        status_var = tk.StringVar()
        error_label = tk.Label(card, textvariable=status_var, bg="#2A2A2A", fg="#E57373",
                               font=("Segoe UI", 9))
        error_label.pack(pady=(20, 0))

        success = {"ok": False}

        def attempt_login():
            user = db.verify_user(username_var.get().strip(), password_var.get().strip())
            if not user:
                status_var.set("❌ Kullanici adi veya sifre hatali")
                if self.logger:
                    self.logger.warning("Giris basarisiz: %s", username_var.get().strip())
                return
            self.current_user = user["username"]
            self.is_admin = user["role"] == "admin"
            region = user.get("region") or "Ankara"
            self.current_region = region
            if self.is_admin:
                self.admin_entry_region_var.set(region)
                self.admin_view_region_var.set("Tum Bolgeler")
            else:
                self.admin_entry_region_var.set(region)
                self.admin_view_region_var.set(region)
            success["ok"] = True
            dialog.destroy()

        btn_frame = tk.Frame(card, bg="#2A2A2A")
        btn_frame.pack(pady=(28, 40), padx=56, fill=tk.X)

        login_btn = tk.Button(btn_frame, text="Giris Yap", command=attempt_login,
                              bg="#5B9BD5", fg="#1E1E1E", font=("Segoe UI", 11, "bold"),
                              relief="flat", cursor="hand2", bd=0)
        login_btn.pack(fill=tk.X, ipady=12)

        def on_enter(_event):
            login_btn.config(bg="#7BB3E0")

        def on_leave(_event):
            login_btn.config(bg="#5B9BD5")

        login_btn.bind("<Enter>", on_enter)
        login_btn.bind("<Leave>", on_leave)

        dialog.bind("<Return>", lambda _e: attempt_login())
        username_entry.focus_set()

        self.wait_window(dialog)
        return success["ok"]

    def _view_region(self):
        if self.is_admin:
            value = self.admin_view_region_var.get().strip()
            if value in {"Tum Bolgeler", "ALL"}:
                return None
            return value or None
        return self.current_region

    def _refresh_region_views(self):
        self.refresh_employees()
        self.refresh_timesheets()
        self.refresh_admin_summary()
        self.refresh_vehicles()
        self.refresh_drivers()
        self.refresh_faults()
        self.refresh_service_visits()
        self.refresh_vehicle_dashboard()

    def _entry_region(self):
        if self.is_admin:
            value = self.admin_entry_region_var.get().strip()
            return value or "Ankara"
        return self.current_region

    def _configure_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        # GECE TEMASI - Silik mavi, gri ve altın tonları
        primary = "#5B9BD5"
        primary_hover = "#7BB3E0"
        accent_gold = "#C9A961"

        bg_app = "#1E1E1E"
        bg_content = "#2A2A2A"
        bg_elevated = "#323232"
        bg_input = "#363636"
        bg_hover = "#3A3A3A"

        text_primary = "#E0E0E0"
        text_secondary = "#B0B0B0"
        text_disabled = "#707070"

        self.configure(bg=bg_app)

        style.configure("Header.TLabel",
            font=("Segoe UI", 18, "bold"),
            foreground=accent_gold,
            background=bg_app,
            padding=(0, 8, 0, 12))
        style.configure("SubHeader.TLabel",
            font=("Segoe UI", 12),
            foreground=text_secondary,
            background=bg_app,
            padding=(0, 4, 0, 8))

        style.configure("Section.TLabelframe",
            padding=(24, 20),
            relief="flat",
            borderwidth=1,
            background=bg_content)
        style.configure("Section.TLabelframe.Label",
            font=("Segoe UI", 11, "bold"),
            foreground=primary,
            background=bg_content,
            padding=(0, 0, 0, 8))

        style.configure("Accent.TButton",
            padding=(20, 10),
            background=primary,
            foreground=bg_app,
            borderwidth=0,
            relief="flat",
            font=("Segoe UI", 10, "bold"))
        style.map("Accent.TButton",
            background=[("active", primary_hover), ("pressed", primary_hover)],
            foreground=[("active", bg_app), ("pressed", bg_app)])

        style.configure("TButton",
            padding=(16, 8),
            background=bg_elevated,
            foreground=text_primary,
            borderwidth=1,
            relief="solid",
            font=("Segoe UI", 10))
        style.map("TButton",
            background=[("active", bg_hover), ("pressed", bg_hover)],
            foreground=[("active", primary), ("pressed", primary)])

        style.configure("Treeview",
            rowheight=36,
            fieldbackground=bg_content,
            background=bg_content,
            foreground=text_primary,
            borderwidth=1,
            font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background=bg_elevated,
            foreground=primary,
            borderwidth=1,
            relief="flat",
            padding=(8, 8))
        style.map("Treeview.Heading",
            background=[("active", bg_hover)],
            foreground=[("active", accent_gold)])
        style.map("Treeview",
            background=[("selected", "#3A4A5A")],
            foreground=[("selected", text_primary)])

        style.configure("TFrame", background=bg_app)
        style.configure("Card.TFrame", background=bg_content, relief="flat")
        style.configure("TLabel",
            background=bg_app,
            font=("Segoe UI", 10),
            foreground=text_primary)
        style.configure("CardLabel.TLabel",
            background=bg_content,
            font=("Segoe UI", 10),
            foreground=text_primary)

        style.configure("TNotebook",
            background=bg_app,
            borderwidth=0,
            tabmargins=(0, 0, 0, 0))
        style.configure("TNotebook.Tab",
            padding=(24, 12),
            borderwidth=0,
            font=("Segoe UI", 10))
        style.map("TNotebook.Tab",
            background=[("selected", bg_content), ("!selected", bg_app)],
            foreground=[("selected", primary), ("!selected", text_secondary)],
            expand=[("selected", [1, 1, 1, 0])])

        style.configure("TEntry",
            padding=(12, 10),
            borderwidth=1,
            relief="solid",
            fieldbackground=bg_input,
            foreground=text_primary,
            selectbackground=bg_hover,
            selectforeground=text_primary,
            insertwidth=2,
            insertcolor=primary,
            font=("Segoe UI", 10))
        style.map("TEntry",
            fieldbackground=[("focus", bg_elevated)],
            bordercolor=[("focus", primary)])

        style.configure("TCombobox",
            padding=(12, 10),
            borderwidth=1,
            fieldbackground=bg_input,
            readonlybackground=bg_input,
            foreground=text_primary,
            selectbackground=bg_hover,
            selectforeground=text_primary,
            arrowsize=14,
            font=("Segoe UI", 10))
        style.map("TCombobox",
            fieldbackground=[("readonly", bg_input), ("!active", bg_input)],
            foreground=[("readonly", text_primary)],
            selectbackground=[("readonly", bg_hover)],
            selectforeground=[("readonly", text_primary)])

        # Dropdown list ve genel selection renkleri
        self.option_add("*TCombobox*Listbox.background", bg_content)
        self.option_add("*TCombobox*Listbox.foreground", text_primary)
        self.option_add("*TCombobox*Listbox.selectBackground", bg_hover)
        self.option_add("*TCombobox*Listbox.selectForeground", text_primary)
        self.option_add("*Entry.selectBackground", bg_hover)
        self.option_add("*Entry.selectForeground", text_primary)

        style.configure("TScrollbar",
            troughcolor=bg_content,
            background=bg_hover,
            bordercolor=bg_hover,
            arrowcolor=text_secondary)

        style.configure("TLabelframe", background=bg_content, borderwidth=0)
        style.configure("TLabelframe.Label", background=bg_content, foreground=text_secondary)
        style.configure("TMenubutton", background=bg_content, foreground=text_primary, borderwidth=0)
        style.configure("Toolbutton", background=bg_content, foreground=text_primary)
        style.configure("TCheckbutton", background=bg_app, foreground=text_primary, indicatorcolor=bg_input)

    def _build_ui(self):
        self.title("Rainstaff Puantaj")
        self.geometry("1280x800")
        self.configure(bg="#1E1E1E")

        header = tk.Frame(self, bg="#2A2A2A", height=64)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        logo_container = tk.Frame(header, bg="#2A2A2A")
        logo_container.place(x=32, y=16)

        logo_path = os.path.join(os.path.dirname(__file__), "assets", "rainstaff_logo_1.png")
        self._logo_image = load_logo_image(logo_path, target_height=32)
        if self._logo_image:
            tk.Label(logo_container, image=self._logo_image, bg="#2A2A2A").pack(side=tk.LEFT)
        else:
            tk.Label(logo_container, text="RAINSTAFF", bg="#2A2A2A",
                fg="#C9A961", font=("Segoe UI", 12, "bold")).pack(side=tk.LEFT)

        user_info = tk.Frame(header, bg="#2A2A2A")
        user_info.place(relx=1.0, x=-32, y=20, anchor="ne")

        user_text = f"{self.current_user}"
        if self.is_admin:
            user_text += " (Admin)"
        tk.Label(user_info, text=user_text, bg="#2A2A2A",
            fg="#B0B0B0", font=("Segoe UI", 10)).pack(side=tk.RIGHT)

        divider = tk.Frame(self, bg="#3A3A3A", height=1)
        divider.pack(fill=tk.X)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        self.tab_employees = ttk.Frame(self.notebook)
        self.tab_timesheets = ttk.Frame(self.notebook)
        self.tab_reports = ttk.Frame(self.notebook)
        self.tab_settings = ttk.Frame(self.notebook)
        self.tab_admin = ttk.Frame(self.notebook)
        self.tab_vehicles = ttk.Frame(self.notebook)
        self.tab_dashboard = ttk.Frame(self.notebook)
        self.tab_service = ttk.Frame(self.notebook)
        self.tab_logs = ttk.Frame(self.notebook)
        self.tab_stock = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_dashboard, text="Dashboard")
        self.notebook.add(self.tab_timesheets, text="Puantaj")
        self.notebook.add(self.tab_employees, text="Çalışanlar")
        self.notebook.add(self.tab_vehicles, text="Araçlar")
        self.notebook.add(self.tab_service, text="Servis")
        self.notebook.add(self.tab_stock, text="Stok Yönetimi")
        self.notebook.add(self.tab_reports, text="Raporlar")
        self.notebook.add(self.tab_admin, text="Yönetim")
        self.notebook.add(self.tab_settings, text="Ayarlar")
        self.notebook.add(self.tab_logs, text="Loglar")

        self.tab_employees_body = self._make_tab_scrollable(self.tab_employees)
        self.tab_timesheets_body = self._make_tab_scrollable(self.tab_timesheets)
        self.tab_reports_body = self._make_tab_scrollable(self.tab_reports)
        self.tab_settings_body = self._make_tab_scrollable(self.tab_settings)
        self.tab_admin_body = self._make_tab_scrollable(self.tab_admin)
        self.tab_vehicles_body = self._make_tab_scrollable(self.tab_vehicles)
        self.tab_dashboard_body = self._make_tab_scrollable(self.tab_dashboard)
        self.tab_service_body = self._make_tab_scrollable(self.tab_service)
        self.tab_logs_body = self._make_tab_scrollable(self.tab_logs)
        self.tab_stock_body = self._make_tab_scrollable(self.tab_stock)

        self._build_employees_tab()
        self._build_timesheets_tab()
        self._build_reports_tab()
        self._build_settings_tab()
        self._build_admin_tab()
        self._build_vehicles_tab()
        self._build_dashboard_tab()
        self._build_service_tab()
        self._build_stock_tab()
        self._build_logs_tab()

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        status_bar = ttk.Label(self, textvariable=self.status_var, anchor=tk.W, foreground="#5B9BD5")
        status_bar.pack(fill=tk.X, padx=10, pady=(0, 8))
        self.status_var.set("Hazir")

    def _on_tab_changed(self, _event):
        current = self.notebook.nametowidget(self.notebook.select())
        self._load_tab_data(current)

    def _load_tab_data(self, tab):
        if self._tab_loaded.get(tab):
            return
        if tab is self.tab_employees:
            self.refresh_employees()
        elif tab is self.tab_timesheets:
            self.refresh_employees()
            self.refresh_timesheets()
        elif tab is self.tab_reports:
            self.refresh_report_archive()
        elif tab is self.tab_settings:
            self.refresh_shift_templates()
        elif tab is self.tab_admin:
            self.refresh_admin_summary()
        elif tab is self.tab_vehicles:
            self.refresh_vehicles()
            self.refresh_drivers()
            self.refresh_faults()
            self.refresh_service_visits()
        elif tab is self.tab_dashboard:
            self.refresh_vehicle_dashboard()
        elif tab is self.tab_service:
            self.refresh_service_visits()
        self._tab_loaded[tab] = True

    def _make_tab_scrollable(self, tab):
        canvas = tk.Canvas(tab, highlightthickness=0)
        vscroll = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=canvas.yview)
        hscroll = ttk.Scrollbar(tab, orient=tk.HORIZONTAL, command=canvas.xview)
        canvas.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)

        tab.rowconfigure(0, weight=1)
        tab.columnconfigure(0, weight=1)
        canvas.grid(row=0, column=0, sticky="nsew")
        vscroll.grid(row=0, column=1, sticky="ns")
        hscroll.grid(row=1, column=0, sticky="ew")

        content = ttk.Frame(canvas)
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfigure(window_id, width=e.width),
        )
        content.bind("<Enter>", lambda _e: self._bind_canvas_mousewheel(canvas))
        content.bind("<Leave>", lambda _e: self._unbind_canvas_mousewheel(canvas))
        return content

    def _make_window_scrollable(self, window):
        canvas = tk.Canvas(window, highlightthickness=0)
        vscroll = ttk.Scrollbar(window, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)

        window.rowconfigure(0, weight=1)
        window.columnconfigure(0, weight=1)
        canvas.grid(row=0, column=0, sticky="nsew")
        vscroll.grid(row=0, column=1, sticky="ns")

        content = ttk.Frame(canvas)
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(window_id, width=e.width))
        content.bind("<Enter>", lambda _e: self._bind_canvas_mousewheel(canvas))
        content.bind("<Leave>", lambda _e: self._unbind_canvas_mousewheel(canvas))
        return content

    def _bind_canvas_mousewheel(self, canvas):
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

    def _unbind_canvas_mousewheel(self, canvas):
        canvas.unbind_all("<MouseWheel>")

    def _drain_log_queue(self):
        if not hasattr(self, "log_text"):
            return
        drained = False
        while True:
            try:
                msg = self.log_queue.get_nowait()
            except queue.Empty:
                break
            drained = True
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.insert(tk.END, msg + "\n")
            self.log_text.configure(state=tk.DISABLED)
        if drained:
            total_lines = int(self.log_text.index("end-1c").split(".")[0])
            if total_lines > 2000:
                self.log_text.configure(state=tk.NORMAL)
                self.log_text.delete("1.0", f"{total_lines - 2000}.0")
                self.log_text.configure(state=tk.DISABLED)
            self.log_text.see(tk.END)
        self.after(500, self._drain_log_queue)

    def clear_log_view(self):
        if hasattr(self, "log_text"):
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.delete("1.0", tk.END)
            self.log_text.insert(tk.END, "Log temizlendi.\n")
            self.log_text.configure(state=tk.DISABLED)

    def trigger_sync(self, reason="manual", force=False):
        if force and hasattr(self, "sync_enabled_var"):
            enabled = self.sync_enabled_var.get()
            sync_url = self.sync_url_var.get().strip()
            token = self.sync_token_var.get().strip()
        else:
            self.settings = db.get_all_settings()
            enabled = self.settings.get("sync_enabled") == "1"
            sync_url = self.settings.get("sync_url", "").strip()
            token = self.settings.get("sync_token", "").strip()
        if not enabled:
            self.status_var.set("Senkron kapali")
            return
        if requests is None:
            self.status_var.set("Senkron icin requests kurulu degil")
            return
        if not sync_url:
            self.status_var.set("Senkron URL bos")
            return
        self.status_var.set("Senkron basladi...")
        thread = threading.Thread(target=self._sync_worker, args=(sync_url, token, reason), daemon=True)
        thread.start()

    def _sync_worker(self, sync_url, token, reason):
        """Senkronizasyon worker; upload + download + merge logic (19 Ocak)."""
        msg = None
        try:
            # Step 1: Upload local DB to server
            # Get region from settings or use default
            settings = db.get_all_settings()
            user_region = settings.get("user_region", "Ankara")
            current_region = user_region or "ALL"
            
            with open(db.DB_PATH, "rb") as handle:
                files = {"db": ("puantaj.db", handle, "application/octet-stream")}
                headers = {
                    "X-API-KEY": token,
                    "X-Region": current_region,
                    "X-Reason": reason
                }
                url = sync_url.rstrip("/") + "/sync"
                resp = requests.post(url, headers=headers, files=files, timeout=10)

            # Basarili upload doğrulama
            if resp.status_code != 200:
                msg = f"Senkron hatasi: Upload HTTP {resp.status_code}"
                if self.logger:
                    self.logger.warning("Cloud sync upload error: %s", msg)
                self.after(0, lambda: self._notify_sync_result(msg, reason))
                return

            # Step 2: Download merged DB from server
            headers = {"X-API-KEY": token}
            download_url = sync_url.rstrip("/") + "/sync/download"
            
            resp = requests.get(download_url, headers=headers, timeout=10)
            
            if resp.status_code != 200:
                msg = f"Senkron hatasi: Download HTTP {resp.status_code}"
                if self.logger:
                    self.logger.warning("Cloud sync download error: %s", msg)
                self.after(0, lambda: self._notify_sync_result(msg, reason))
                return

            # Step 3: Backup current local database
            import shutil
            backup_path = db.DB_PATH + ".sync_backup"
            if os.path.isfile(db.DB_PATH):
                shutil.copy2(db.DB_PATH, backup_path)

            # Step 4: Write downloaded database as new local DB
            with open(db.DB_PATH, "wb") as f:
                f.write(resp.content)

            msg = "Senkron basarili"
            if self.logger:
                self.logger.info("Cloud sync completed (upload+download+merge): %s", reason)
            
        except requests.Timeout:
            msg = "Senkron hatasi: Baglanti timeout"
            if self.logger:
                self.logger.warning("Cloud sync timeout")
        except requests.RequestException as e:
            msg = f"Senkron hatasi: {str(e)[:80]}"
            if self.logger:
                self.logger.warning("Cloud sync request error: %s", str(e))
        except Exception as e:
            msg = f"Senkron hatasi: {str(e)[:80]}"
            if self.logger:
                self.logger.error("Cloud sync unexpected error: %s", str(e))

        if msg:
            self.after(0, lambda: self._notify_sync_result(msg, reason))

    def manual_sync(self):
        if not self.sync_enabled_var.get():
            messagebox.showwarning("Uyari", "Senkron kapali. Ayarlardan acin.")
            return
        self.trigger_sync("manual", force=True)

    def _notify_sync_result(self, message, reason):
        self.status_var.set(message)
        if reason == "manual":
            messagebox.showinfo("Senkron", message)

    # Employees tab
    def _build_employees_tab(self):
        form = ttk.LabelFrame(self.tab_employees_body, text="Calisan Bilgisi", style="Section.TLabelframe")
        form.pack(fill=tk.X, padx=6, pady=6)

        self.emp_id_var = tk.StringVar()
        self.emp_name_var = tk.StringVar()
        self.emp_identity_var = tk.StringVar()
        self.emp_department_var = tk.StringVar()
        self.emp_title_var = tk.StringVar()

        row1 = ttk.Frame(form)
        row1.pack(fill=tk.X, pady=4)
        create_labeled_entry(row1, "Ad Soyad", self.emp_name_var, 30).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row1, "TCKN", self.emp_identity_var, 18).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row1, "Departman", self.emp_department_var, 18).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row1, "Unvan", self.emp_title_var, 18).pack(side=tk.LEFT, padx=6)

        btn_row = ttk.Frame(form)
        btn_row.pack(fill=tk.X, pady=6)
        ttk.Button(btn_row, text="Kaydet", style="Accent.TButton", command=self.add_or_update_employee).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(btn_row, text="Sil", command=self.delete_employee).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="Temizle", command=self.clear_employee_form).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Excel/CSV Iceri Aktar", command=self.import_employees).pack(side=tk.LEFT, padx=6)

        list_frame = ttk.Frame(self.tab_employees_body)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        columns = ("id", "name", "identity", "department", "title", "region")
        self.employee_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        self.employee_tree.heading("id", text="ID")
        self.employee_tree.heading("name", text="Ad Soyad")
        self.employee_tree.heading("identity", text="TCKN")
        self.employee_tree.heading("department", text="Departman")
        self.employee_tree.heading("title", text="Unvan")
        self.employee_tree.heading("region", text="Bolge")
        self.employee_tree.column("id", width=60, anchor=tk.CENTER)
        self.employee_tree.column("name", width=220)
        self.employee_tree.column("identity", width=140)
        self.employee_tree.column("department", width=160)
        self.employee_tree.column("title", width=160)
        self.employee_tree.column("region", width=110)
        # Koyu tema zebra satırları
        self.employee_tree.tag_configure("odd", background="#252525", foreground="#E0E0E0")
        self.employee_tree.tag_configure("even", background="#1F1F1F", foreground="#E0E0E0")
        emp_xscroll = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.employee_tree.xview)
        emp_yscroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.employee_tree.yview)
        self.employee_tree.configure(xscrollcommand=emp_xscroll.set, yscrollcommand=emp_yscroll.set)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        self.employee_tree.grid(row=0, column=0, sticky="nsew")
        emp_yscroll.grid(row=0, column=1, sticky="ns")
        emp_xscroll.grid(row=1, column=0, sticky="ew")
        self.employee_tree.bind("<<TreeviewSelect>>", self.on_employee_select)

    def refresh_employees(self):
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        self.employee_map = {}
        self.employee_display_names = []
        self.employee_details = {}
        name_counts = {}
        for emp in db.list_employees(region=self._view_region()):
            emp_id, name, identity_no, department, title, region = emp
            name_counts[name] = name_counts.get(name, 0) + 1
            self.employee_map[(name, region or "")] = emp_id
            tag = "odd" if len(self.employee_tree.get_children()) % 2 else "even"
            self.employee_tree.insert("", tk.END, values=emp, tags=(tag,))
            self.employee_details[(name, region or "")] = {
                "department": department or "",
                "title": title or "",
                "identity_no": identity_no or "",
                "region": region or "",
            }
        self.employee_display_names = []
        for emp in db.list_employees(region=self._view_region()):
            _emp_id, name, _identity_no, _department, _title, region = emp
            display = name
            if name_counts.get(name, 0) > 1:
                display = f"{name} ({region or '-'})"
            self.employee_display_names.append(display)
        self._refresh_employee_comboboxes()

    def _refresh_employee_comboboxes(self):
        values = ["Tum Calisanlar"] + sorted(self.employee_display_names)
        if hasattr(self, "ts_employee_combo"):
            self.ts_employee_combo["values"] = values
        if hasattr(self, "ts_filter_combo"):
            self.ts_filter_combo["values"] = values
        if hasattr(self, "report_employee_combo"):
            self.report_employee_combo["values"] = values
        if hasattr(self, "admin_employee_combo"):
            self.admin_employee_combo["values"] = values
        if hasattr(self, "admin_department_combo"):
            departments = sorted({d["department"] for d in self.employee_details.values() if d["department"]})
            self.admin_department_combo["values"] = ["Tum Departmanlar"] + departments
        if hasattr(self, "admin_title_combo"):
            titles = sorted({d["title"] for d in self.employee_details.values() if d["title"]})
            self.admin_title_combo["values"] = ["Tum Unvanlar"] + titles

    def clear_employee_form(self):
        self.emp_id_var.set("")
        self.emp_name_var.set("")
        self.emp_identity_var.set("")
        self.emp_department_var.set("")
        self.emp_title_var.set("")

    def on_employee_select(self, _event=None):
        selected = self.employee_tree.selection()
        if not selected:
            return
        values = self.employee_tree.item(selected[0], "values")
        self.emp_id_var.set(values[0])
        self.emp_name_var.set(values[1])
        self.emp_identity_var.set(values[2])
        self.emp_department_var.set(values[3])
        self.emp_title_var.set(values[4])

    def add_or_update_employee(self):
        name = self.emp_name_var.get().strip()
        if not name:
            messagebox.showwarning("Uyari", "Ad Soyad zorunlu.")
            return
        identity_no = self.emp_identity_var.get().strip()
        department = self.emp_department_var.get().strip()
        title = self.emp_title_var.get().strip()

        emp_id = self.emp_id_var.get().strip()
        if emp_id:
            db.update_employee(
                parse_int(emp_id),
                name,
                identity_no,
                department,
                title,
                self._entry_region(),
            )
            self._log_action("employee_update", f"id={emp_id} name={name}")
        else:
            db.add_employee(name, identity_no, department, title, self._entry_region())
            self._log_action("employee_add", f"name={name}")
        self.refresh_employees()
        self.clear_employee_form()
        self.trigger_sync("employee")

    def delete_employee(self):
        emp_id = self.emp_id_var.get().strip()
        if not emp_id:
            messagebox.showwarning("Uyari", "Silmek icin calisan secin.")
            return
        if messagebox.askyesno("Onay", "Calisani silmek istiyor musunuz?"):
            db.delete_employee(parse_int(emp_id))
            self._log_action("employee_delete", f"id={emp_id}")
            self.refresh_employees()
            self.clear_employee_form()
            self.trigger_sync("employee_delete")

    def notify(self, message, sound=True, duration_ms=2500):
        self.status_var.set(message)
        if sound and winsound:
            try:
                winsound.MessageBeep(winsound.MB_ICONASTERISK)
            except Exception:
                pass
        self.after(duration_ms, lambda: self.status_var.set(""))

    # Timesheets tab
    def _build_timesheets_tab(self):
        form = ttk.LabelFrame(self.tab_timesheets_body, text="Puantaj Girisi", style="Section.TLabelframe")
        form.pack(fill=tk.X, padx=6, pady=6)

        self.ts_id_var = tk.StringVar()
        self.ts_employee_var = tk.StringVar()
        self.ts_date_var = tk.StringVar()
        self.ts_start_var = tk.StringVar(value="09:00")
        self.ts_end_var = tk.StringVar(value="18:00")
        self.ts_break_var = tk.StringVar(value="60")
        self.ts_notes_var = tk.StringVar()
        self.ts_template_var = tk.StringVar()
        self.ts_special_var = tk.IntVar(value=0)

        row1 = ttk.Frame(form)
        row1.pack(fill=tk.X, pady=4)
        ttk.Label(row1, text="Calisan").pack(side=tk.LEFT, padx=(0, 8))
        self.ts_employee_combo = ttk.Combobox(row1, textvariable=self.ts_employee_var, width=28, state="readonly")
        self.ts_employee_combo.pack(side=tk.LEFT)
        date_frame, self.ts_date_entry = create_labeled_date(row1, "Tarih", self.ts_date_var, 12)
        date_frame.pack(side=tk.LEFT, padx=6)
        start_frame = create_time_entry(row1, "Giris", self.ts_start_var, 8)
        start_frame.pack(side=tk.LEFT, padx=6)
        end_frame = create_time_entry(row1, "Cikis", self.ts_end_var, 8)
        end_frame.pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row1, "Mola dk", self.ts_break_var, 8).pack(side=tk.LEFT, padx=6)
        for child in start_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind("<FocusOut>", lambda _e: normalize_time_in_var(self.ts_start_var))
        for child in end_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind("<FocusOut>", lambda _e: normalize_time_in_var(self.ts_end_var))

        row2 = ttk.Frame(form)
        row2.pack(fill=tk.X, pady=4)
        create_labeled_entry(row2, "Not", self.ts_notes_var, 60).pack(side=tk.LEFT, padx=6)
        ttk.Checkbutton(row2, text="Ozel Gun", variable=self.ts_special_var).pack(side=tk.LEFT, padx=6)

        row3 = ttk.Frame(form)
        row3.pack(fill=tk.X, pady=4)
        ttk.Label(row3, text="Sablon").pack(side=tk.LEFT, padx=(0, 8))
        self.ts_template_combo = ttk.Combobox(row3, textvariable=self.ts_template_var, width=28, state="readonly")
        self.ts_template_combo.pack(side=tk.LEFT)
        ttk.Button(row3, text="Uygula", command=self.apply_shift_template).pack(side=tk.LEFT, padx=6)

        btn_row = ttk.Frame(form)
        btn_row.pack(fill=tk.X, pady=6)
        ttk.Button(btn_row, text="Kaydet", style="Accent.TButton", command=self.add_or_update_timesheet).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(btn_row, text="Sil", command=self.delete_timesheet).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="Temizle", command=self.clear_timesheet_form).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Excel/CSV Iceri Aktar", command=self.import_timesheets).pack(side=tk.LEFT, padx=6)

        filter_frame = ttk.LabelFrame(self.tab_timesheets_body, text="Filtre", style="Section.TLabelframe")
        filter_frame.pack(fill=tk.X, padx=6, pady=6)
        self.ts_filter_employee = tk.StringVar(value="Tum Calisanlar")
        self.ts_filter_start = tk.StringVar()
        self.ts_filter_end = tk.StringVar()

        ttk.Label(filter_frame, text="Calisan").pack(side=tk.LEFT, padx=(0, 8))
        self.ts_filter_combo = ttk.Combobox(filter_frame, textvariable=self.ts_filter_employee, width=28, state="readonly")
        self.ts_filter_combo.pack(side=tk.LEFT)
        start_frame, self.ts_filter_start_entry = create_labeled_date(
            filter_frame, "Baslangic", self.ts_filter_start, 12
        )
        start_frame.pack(side=tk.LEFT, padx=6)
        end_frame, self.ts_filter_end_entry = create_labeled_date(filter_frame, "Bitis", self.ts_filter_end, 12)
        end_frame.pack(side=tk.LEFT, padx=6)
        ttk.Button(filter_frame, text="Filtrele", command=self.refresh_timesheets).pack(side=tk.LEFT, padx=6)
        ttk.Button(filter_frame, text="Temizle", command=self.clear_timesheet_filter).pack(side=tk.LEFT)
        clear_date_entry(self.ts_filter_start_entry)
        clear_date_entry(self.ts_filter_end_entry)

        list_frame = ttk.Frame(self.tab_timesheets_body)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        columns = (
            "id",
            "employee",
            "date",
            "start",
            "end",
            "break",
            "worked",
            "scheduled",
            "overtime",
            "night",
            "overnight",
            "special",
            "special_normal",
            "special_overtime",
            "special_night",
            "notes",
            "region",
        )
        self.timesheet_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        self.timesheet_tree.heading("id", text="ID")
        self.timesheet_tree.heading("employee", text="Calisan")
        self.timesheet_tree.heading("date", text="Tarih")
        self.timesheet_tree.heading("start", text="Giris")
        self.timesheet_tree.heading("end", text="Cikis")
        self.timesheet_tree.heading("break", text="Mola")
        self.timesheet_tree.heading("worked", text="Calisilan")
        self.timesheet_tree.heading("scheduled", text="Plan")
        self.timesheet_tree.heading("overtime", text="Fazla Mesai")
        self.timesheet_tree.heading("night", text="Gece")
        self.timesheet_tree.heading("overnight", text="Geceye Tasan")
        self.timesheet_tree.heading("special", text="Ozel Gun")
        self.timesheet_tree.heading("special_normal", text="Ozel Gun Normal")
        self.timesheet_tree.heading("special_overtime", text="Ozel Gun Fazla")
        self.timesheet_tree.heading("special_night", text="Ozel Gun Gece")
        self.timesheet_tree.heading("notes", text="Not")
        self.timesheet_tree.heading("region", text="Bolge")
        self.timesheet_tree.column("id", width=60, anchor=tk.CENTER)
        self.timesheet_tree.column("employee", width=220)
        self.timesheet_tree.column("date", width=100)
        self.timesheet_tree.column("start", width=80)
        self.timesheet_tree.column("end", width=80)
        self.timesheet_tree.column("break", width=80)
        self.timesheet_tree.column("worked", width=90)
        self.timesheet_tree.column("scheduled", width=90)
        self.timesheet_tree.column("overtime", width=90)
        self.timesheet_tree.column("night", width=90)
        self.timesheet_tree.column("overnight", width=100)
        self.timesheet_tree.column("special", width=80)
        self.timesheet_tree.column("special_normal", width=110)
        self.timesheet_tree.column("special_overtime", width=110)
        self.timesheet_tree.column("special_night", width=110)
        self.timesheet_tree.column("notes", width=180)
        self.timesheet_tree.column("region", width=100)
        # Koyu tema zebra satırları
        self.timesheet_tree.tag_configure("odd", background="#252525", foreground="#E0E0E0")
        self.timesheet_tree.tag_configure("even", background="#1F1F1F", foreground="#E0E0E0")
        ts_xscroll = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.timesheet_tree.xview)
        ts_yscroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.timesheet_tree.yview)
        self.timesheet_tree.configure(xscrollcommand=ts_xscroll.set, yscrollcommand=ts_yscroll.set)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        self.timesheet_tree.grid(row=0, column=0, sticky="nsew")
        ts_yscroll.grid(row=0, column=1, sticky="ns")
        ts_xscroll.grid(row=1, column=0, sticky="ew")
        self.timesheet_tree.bind("<<TreeviewSelect>>", self.on_timesheet_select)
        self.timesheet_tree.bind("<Button-3>", self.on_timesheet_right_click)

        self.ts_menu = tk.Menu(self, tearoff=0)
        self.ts_menu.add_command(label="Duzenle", command=self.edit_selected_timesheet)
        self.ts_menu.add_command(label="Sil", command=self.delete_timesheet)

    def clear_timesheet_form(self):
        self.ts_id_var.set("")
        self.ts_employee_var.set("")
        self.ts_date_var.set("")
        self.ts_start_var.set("09:00")
        self.ts_end_var.set("18:00")
        self.ts_break_var.set("60")
        self.ts_notes_var.set("")
        self.ts_template_var.set("")
        self.ts_special_var.set(0)
        self.ts_original = None
        self.ts_editing_id = None

    def clear_timesheet_filter(self):
        self.ts_filter_employee.set("Tum Calisanlar")
        self.ts_filter_start.set("")
        self.ts_filter_end.set("")
        clear_date_entry(self.ts_filter_start_entry)
        clear_date_entry(self.ts_filter_end_entry)
        self.refresh_timesheets()

    def on_timesheet_select(self, _event=None):
        return

    def on_timesheet_right_click(self, event):
        row_id = self.timesheet_tree.identify_row(event.y)
        if row_id:
            self.timesheet_tree.selection_set(row_id)
            self.ts_menu.tk_popup(event.x_root, event.y_root)

    def edit_selected_timesheet(self):
        selected = self.timesheet_tree.selection()
        if not selected:
            return
        values = self.timesheet_tree.item(selected[0], "values")
        self.ts_editing_id = values[0]
        self.ts_id_var.set(values[0])
        region = values[16] if len(values) > 16 else ""
        display_name = values[1]
        if region:
            display_name = f"{values[1]} ({region})"
        self.ts_employee_var.set(display_name)
        self.ts_date_var.set(values[2])
        set_time_vars(values[3], self.ts_start_var)
        set_time_vars(values[4], self.ts_end_var)
        self.ts_break_var.set(values[5])
        self.ts_special_var.set(1 if values[11] == "Evet" else 0)
        self.ts_notes_var.set(values[15])
        self.ts_original = (values[1], values[2], values[3], values[4])

    def refresh_timesheets(self):
        for item in self.timesheet_tree.get_children():
            self.timesheet_tree.delete(item)

        employee_name = self.ts_filter_employee.get()
        employee_id = None
        if employee_name and employee_name != "Tum Calisanlar":
            base, region = split_display_name(employee_name, REGIONS)
            if region is None:
                employee_id = self.employee_map.get((base, "")) or self.employee_map.get(
                    (base, self._entry_region())
                )
            else:
                employee_id = self.employee_map.get((base, region))
        start_date = self.ts_filter_start.get().strip() or None
        end_date = self.ts_filter_end.get().strip() or None
        try:
            if start_date:
                start_date = normalize_date(start_date)
            if end_date:
                end_date = normalize_date(end_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return

        records = db.list_timesheets(
            employee_id=employee_id,
            start_date=start_date,
            end_date=end_date,
            region=self._view_region(),
        )
        for ts in records:
            ts_id, _emp_id, name, work_date, start_time, end_time, break_minutes, is_special, notes, region = ts
            try:
                (
                    worked,
                    scheduled,
                    overtime,
                    night_hours,
                    overnight_hours,
                    spec_norm,
                    spec_ot,
                    spec_night,
                ) = calc.calc_day_hours(
                    work_date,
                    start_time,
                    end_time,
                    break_minutes,
                    self.settings,
                    is_special,
                )
            except Exception:
                worked = scheduled = overtime = ""
                night_hours = overnight_hours = ""
                spec_norm = spec_ot = spec_night = ""
            tag = "odd" if len(self.timesheet_tree.get_children()) % 2 else "even"
            self.timesheet_tree.insert(
                "",
                tk.END,
                values=(
                    ts_id,
                    name,
                    work_date,
                    start_time,
                    end_time,
                    break_minutes,
                    worked,
                    scheduled,
                    overtime,
                    night_hours,
                    overnight_hours,
                    "Evet" if is_special else "Hayir",
                    spec_norm,
                    spec_ot,
                    spec_night,
                    notes or "",
                    region or "",
                ),
                tags=(tag,),
            )

        if hasattr(self, "ts_filter_combo"):
            self.ts_filter_combo["values"] = ["Tum Calisanlar"] + sorted(self.employee_display_names)

    def add_or_update_timesheet(self):
        name = self.ts_employee_var.get().strip()
        if not name:
            messagebox.showwarning("Uyari", "Calisan secin.")
            return
        employee_id = None
        base, region = split_display_name(name, REGIONS)
        if region is None:
            employee_id = self.employee_map.get((base, "")) or self.employee_map.get((base, self._entry_region()))
        else:
            employee_id = self.employee_map.get((base, region))
        if not employee_id:
            messagebox.showwarning("Uyari", "Calisan bulunamadi.")
            return
        work_date = self.ts_date_var.get().strip()
        if not work_date:
            messagebox.showwarning("Uyari", "Tarih zorunlu.")
            return
        try:
            work_date = normalize_date(work_date)
            start_time = normalize_time(self.ts_start_var.get())
            end_time = normalize_time(self.ts_end_var.get())
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        break_minutes = parse_int(self.ts_break_var.get(), 0)
        # Mola dakikasi validasyonu (0-480 dakika = 0-8 saat)
        if not (0 <= break_minutes <= 480):
            messagebox.showwarning("Uyari", "Mola dakikasi 0-480 arasinda olmalidir.")
            return
        notes = self.ts_notes_var.get().strip()
        is_special = 1 if self.ts_special_var.get() else 0
        # Bolge NULL kontrolu
        entry_region = self._entry_region()
        if not entry_region:
            messagebox.showwarning("Uyari", "Bolge tanimlanimis. Ayarlardan kontrol edin.")
            return

        ts_id = self.ts_editing_id
        if ts_id:
            db.update_timesheet(
                parse_int(ts_id),
                employee_id,
                work_date,
                start_time,
                end_time,
                break_minutes,
                is_special,
                notes,
                self._entry_region(),
            )
            self._log_action("timesheet_update", f"id={ts_id} date={work_date}")
        else:
            db.add_timesheet(
                employee_id,
                work_date,
                start_time,
                end_time,
                break_minutes,
                is_special,
                notes,
                self._entry_region(),
            )
            self._log_action("timesheet_add", f"employee_id={employee_id} date={work_date}")
        self.refresh_timesheets()
        self.clear_timesheet_form()
        self.trigger_sync("timesheet")
        self.notify("Puantaj kaydedildi.")

    def delete_timesheet(self):
        ts_id = self.ts_editing_id
        if not ts_id:
            selected = self.timesheet_tree.selection()
            if selected:
                values = self.timesheet_tree.item(selected[0], "values")
                ts_id = values[0]
        if not ts_id:
            messagebox.showwarning("Uyari", "Silmek icin puantaj secin.")
            return
        if messagebox.askyesno("Onay", "Puantaj kaydini silmek istiyor musunuz?"):
            db.delete_timesheet(parse_int(ts_id))
            self._log_action("timesheet_delete", f"id={ts_id}")
            self.refresh_timesheets()
            self.clear_timesheet_form()
            self.notify("Puantaj silindi.")
            self.trigger_sync("timesheet_delete")

    def refresh_shift_templates(self):
        templates = db.list_shift_templates()
        self.shift_template_map = {tpl[1]: tpl for tpl in templates}
        if hasattr(self, "ts_template_combo"):
            self.ts_template_combo["values"] = sorted(self.shift_template_map.keys())
            if not self.ts_template_var.get() and templates:
                self.ts_template_var.set(templates[0][1])
                self.apply_shift_template()
        if hasattr(self, "template_tree"):
            for item in self.template_tree.get_children():
                self.template_tree.delete(item)
            for tpl in templates:
                tag = "odd" if len(self.template_tree.get_children()) % 2 else "even"
                self.template_tree.insert("", tk.END, values=tpl, tags=(tag,))

    def apply_shift_template(self):
        name = self.ts_template_var.get().strip()
        if not name or name not in self.shift_template_map:
            messagebox.showwarning("Uyari", "Sablon secin.")
            return
        _tpl_id, _name, start_time, end_time, break_minutes = self.shift_template_map[name]
        set_time_vars(start_time, self.ts_start_var)
        set_time_vars(end_time, self.ts_end_var)
        self.ts_break_var.set(str(break_minutes))

    def save_shift_template(self):
        name = self.st_name_var.get().strip()
        if not name:
            messagebox.showwarning("Uyari", "Sablon adi zorunlu.")
            return
        try:
            start_time = normalize_time(self.st_start_var.get())
            end_time = normalize_time(self.st_end_var.get())
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        break_minutes = parse_int(self.st_break_var.get(), 0)
        db.upsert_shift_template(name, start_time, end_time, break_minutes)
        self._log_action("shift_template_save", f"name={name}")
        self.refresh_shift_templates()
        self.clear_shift_template_form()

    def delete_shift_template(self):
        tpl_id = self.st_id_var.get().strip()
        if not tpl_id:
            messagebox.showwarning("Uyari", "Silmek icin sablon secin.")
            return
        if messagebox.askyesno("Onay", "Sablonu silmek istiyor musunuz?"):
            db.delete_shift_template(parse_int(tpl_id))
            self._log_action("shift_template_delete", f"id={tpl_id}")
            self.refresh_shift_templates()
            self.clear_shift_template_form()

    def clear_shift_template_form(self):
        self.st_id_var.set("")
        self.st_name_var.set("")
        self.st_start_var.set("09:00")
        self.st_end_var.set("18:00")
        self.st_break_var.set("60")

    def on_template_select(self, _event=None):
        selected = self.template_tree.selection()
        if not selected:
            return
        values = self.template_tree.item(selected[0], "values")
        self.st_id_var.set(values[0])
        self.st_name_var.set(values[1])
        set_time_vars(values[2], self.st_start_var)
        set_time_vars(values[3], self.st_end_var)
        self.st_break_var.set(values[4])

    def import_employees(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV", "*.xlsx;*.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv"), ("All", "*.*")]
        )
        if not path:
            return
        try:
            rows = load_tabular_file(path)
        except Exception as exc:
            messagebox.showerror("Hata", f"Dosya okunamadi: {exc}")
            return
        if not rows:
            messagebox.showinfo("Bilgi", "Dosyada veri bulunamadi.")
            return

        header_map = map_headers(rows[0], EMP_HEADER_MAP)
        start_idx = 1 if "full_name" in header_map else 0
        existing_names = set(self.employee_map.keys())
        imported = 0
        skipped = 0
        for row in rows[start_idx:]:
            def cell(idx):
                return row[idx] if idx is not None and idx < len(row) else ""

            if "full_name" in header_map:
                name = str(cell(header_map.get("full_name"))).strip()
                identity_no = str(cell(header_map.get("identity_no"))).strip()
                department = str(cell(header_map.get("department"))).strip()
                title = str(cell(header_map.get("title"))).strip()
            else:
                name = str(cell(0)).strip()
                identity_no = str(cell(1)).strip() if len(row) > 1 else ""
                department = str(cell(2)).strip() if len(row) > 2 else ""
                title = str(cell(3)).strip() if len(row) > 3 else ""

            key = (name, self._entry_region())
            if not name or key in existing_names:
                skipped += 1
                continue
            db.add_employee(name, identity_no, department, title, self._entry_region())
            existing_names.add(key)
            imported += 1

        self.refresh_employees()
        self._log_action("employee_import", f"file={os.path.basename(path)} added={imported} skipped={skipped}")
        messagebox.showinfo("Bilgi", f"Iceri aktarma tamamlandi. Eklenen: {imported}, Atlanan: {skipped}")

    def import_timesheets(self):
        if not self.employee_map:
            messagebox.showwarning("Uyari", "Once calisan ekleyin.")
            return
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV", "*.xlsx;*.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv"), ("All", "*.*")]
        )
        if not path:
            return
        try:
            rows = load_tabular_file(path)
        except Exception as exc:
            messagebox.showerror("Hata", f"Dosya okunamadi: {exc}")
            return
        if not rows:
            messagebox.showinfo("Bilgi", "Dosyada veri bulunamadi.")
            return

        header_map = map_headers(rows[0], TS_HEADER_MAP)
        start_idx = 1 if "employee" in header_map else 0
        imported = 0
        skipped = 0
        missing_employee = 0

        for row in rows[start_idx:]:
            def cell(idx):
                return row[idx] if idx is not None and idx < len(row) else ""

            if "employee" in header_map:
                employee_name = str(cell(header_map.get("employee"))).strip()
                work_date = cell(header_map.get("work_date"))
                start_time = cell(header_map.get("start_time"))
                end_time = cell(header_map.get("end_time"))
                break_minutes = cell(header_map.get("break_minutes"))
                is_special = cell(header_map.get("is_special"))
                notes = str(cell(header_map.get("notes"))).strip()
            else:
                employee_name = str(cell(0)).strip()
                work_date = cell(1)
                start_time = cell(2)
                end_time = cell(3)
                break_minutes = cell(4) if len(row) > 4 else 0
                is_special = cell(5) if len(row) > 5 else 0
                notes = str(cell(6)).strip() if len(row) > 6 else ""

            if not employee_name:
                skipped += 1
                continue
            employee_id = None
            base, region = split_display_name(employee_name, REGIONS)
            if region is None:
                employee_id = self.employee_map.get((base, "")) or self.employee_map.get(
                    (base, self._entry_region())
                )
            else:
                employee_id = self.employee_map.get((base, region))
            if not employee_id:
                missing_employee += 1
                continue
            try:
                work_date = normalize_date_value(work_date)
                start_time = normalize_time_value(start_time)
                end_time = normalize_time_value(end_time)
            except ValueError:
                skipped += 1
                continue
            break_minutes = parse_int(break_minutes, 0)
            is_special = 1 if parse_bool(is_special) else 0

            db.add_timesheet(
                employee_id,
                work_date,
                start_time,
                end_time,
                break_minutes,
                is_special,
                notes,
                self._entry_region(),
            )
            imported += 1

        self.refresh_timesheets()
        self._log_action(
            "timesheet_import",
            f"file={os.path.basename(path)} added={imported} skipped={skipped} missing={missing_employee}",
        )
        messagebox.showinfo(
            "Bilgi",
            f"Iceri aktarma tamamlandi. Eklenen: {imported}, Atlanan: {skipped}, Calisan bulunamadi: {missing_employee}",
        )

    # Reports tab
    def _build_reports_tab(self):
        frame = ttk.LabelFrame(self.tab_reports_body, text="Excel Raporu", style="Section.TLabelframe")
        frame.pack(fill=tk.X, padx=6, pady=6)

        self.report_employee_var = tk.StringVar(value="Tum Calisanlar")
        self.report_start_var = tk.StringVar()
        self.report_end_var = tk.StringVar()
        self.report_use_dates = tk.BooleanVar(value=False)

        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=6)
        ttk.Label(row1, text="Calisan").pack(side=tk.LEFT, padx=(0, 8))
        self.report_employee_combo = ttk.Combobox(row1, textvariable=self.report_employee_var, width=28, state="readonly")
        self.report_employee_combo.pack(side=tk.LEFT)
        report_start_frame, self.report_start_entry = create_labeled_date(
            row1, "Baslangic", self.report_start_var, 12
        )
        report_start_frame.pack(side=tk.LEFT, padx=6)
        report_end_frame, self.report_end_entry = create_labeled_date(
            row1, "Bitis", self.report_end_var, 12
        )
        report_end_frame.pack(side=tk.LEFT, padx=6)
        ttk.Button(row1, text="Rapor Olustur", style="Accent.TButton", command=self.export_report).pack(
            side=tk.LEFT, padx=6
        )
        clear_date_entry(self.report_start_entry)
        clear_date_entry(self.report_end_entry)

        row2 = ttk.Frame(frame)
        row2.pack(fill=tk.X, pady=(0, 6))
        ttk.Checkbutton(row2, text="Tarih filtresi kullan", variable=self.report_use_dates).pack(
            side=tk.LEFT, padx=6
        )

        viewer = ttk.LabelFrame(self.tab_reports_body, text="Rapor Goruntule", style="Section.TLabelframe")
        viewer.pack(fill=tk.X, padx=6, pady=6)
        ttk.Button(viewer, text="XLSX Sec ve Goruntule", command=self.pick_and_preview_report).pack(
            side=tk.LEFT, padx=6, pady=6
        )

        archive = ttk.LabelFrame(self.tab_reports_body, text="Rapor Arsivi", style="Section.TLabelframe")
        archive.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.report_tree = ttk.Treeview(
            archive,
            columns=("id", "file", "created", "employee", "range"),
            show="headings",
            height=8,
        )
        self.report_tree.heading("id", text="ID")
        self.report_tree.heading("file", text="Dosya")
        self.report_tree.heading("created", text="Tarih")
        self.report_tree.heading("employee", text="Calisan")
        self.report_tree.heading("range", text="Aralik")
        self.report_tree.column("id", width=60, anchor=tk.CENTER)
        self.report_tree.column("file", width=360)
        self.report_tree.column("created", width=140)
        self.report_tree.column("employee", width=180)
        self.report_tree.column("range", width=180)
        report_xscroll = ttk.Scrollbar(archive, orient=tk.HORIZONTAL, command=self.report_tree.xview)
        report_yscroll = ttk.Scrollbar(archive, orient=tk.VERTICAL, command=self.report_tree.yview)
        self.report_tree.configure(xscrollcommand=report_xscroll.set, yscrollcommand=report_yscroll.set)
        archive.columnconfigure(0, weight=1)
        archive.rowconfigure(0, weight=1)
        self.report_tree.grid(row=0, column=0, sticky="nsew")
        report_yscroll.grid(row=0, column=1, sticky="ns")
        report_xscroll.grid(row=1, column=0, sticky="ew")

        archive.rowconfigure(0, weight=1)
        archive.columnconfigure(0, weight=1)

        btn_row = ttk.Frame(archive)
        btn_row.grid(row=2, column=0, sticky="ew", pady=6)
        ttk.Button(btn_row, text="Arsivi Yenile", command=self.refresh_report_archive).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Seciliyi Goruntule", command=self.preview_selected_report).pack(
            side=tk.LEFT, padx=6
        )

        note = ttk.Label(
            frame,
            text=f"Tarih formatlari: {DATE_FMT} / Saat formatlari: {TIME_FMT}",
            foreground="#444444",
        )
        note.pack(anchor=tk.W, padx=6, pady=(0, 6))

    def export_report(self):
        employee_name = self.report_employee_var.get().strip()
        employee_id = None
        if employee_name and employee_name != "Tum Calisanlar":
            base, region = split_display_name(employee_name, REGIONS)
            if region is None:
                employee_id = self.employee_map.get((base, "")) or self.employee_map.get(
                    (base, self._entry_region())
                )
            else:
                employee_id = self.employee_map.get((base, region))
        start_date = self.report_start_var.get().strip() or None
        end_date = self.report_end_var.get().strip() or None
        try:
            if self.report_use_dates.get():
                if start_date:
                    start_date = normalize_date(start_date)
                if end_date:
                    end_date = normalize_date(end_date)
            else:
                start_date = None
                end_date = None
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return

        records = db.list_timesheets(
            employee_id=employee_id,
            start_date=start_date,
            end_date=end_date,
            region=self._view_region(),
        )
        if not records:
            messagebox.showinfo("Bilgi", "Rapor icin veri bulunamadi.")
            return

        if employee_name and employee_name != "Tum Calisanlar":
            employee_slug = employee_name.replace(" ", "_")
        else:
            employee_slug = "tum_calisanlar"
        filename = f"puantaj_raporu_{employee_slug}_{start_date or 'tum'}_{end_date or 'tum'}.xlsx"
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=filename,
        )
        if not output_path:
            return

        date_text = f"Tarih Araligi: {start_date or '-'} - {end_date or '-'}"
        try:
            report.export_report(output_path, records, db.get_all_settings(), date_text)
        except ValueError as exc:
            messagebox.showerror("Hata", str(exc))
            return
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.add_report_log(output_path, created_at, employee_name, start_date, end_date)
        self.refresh_report_archive()
        self._log_action(
            "report_export",
            f"file={os.path.basename(output_path)} employee={employee_name or 'Tum'} range={start_date or '-'}-{end_date or '-'}",
        )
        messagebox.showinfo("Basarili", f"Rapor kaydedildi: {output_path}")

    def refresh_report_archive(self):
        if not hasattr(self, "report_tree"):
            return
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        for rep in db.list_report_logs():
            rep_id, file_path, created_at, employee, start_date, end_date = rep
            range_text = f"{start_date or '-'} - {end_date or '-'}"
            values = (rep_id, file_path, created_at, employee or "Tum Calisanlar", range_text)
            self.report_tree.insert("", tk.END, values=values)

    def pick_and_preview_report(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        self.preview_report(path)

    def preview_selected_report(self):
        selected = self.report_tree.selection()
        if not selected:
            messagebox.showwarning("Uyari", "Once rapor secin.")
            return
        values = self.report_tree.item(selected[0], "values")
        path = values[1]
        self.preview_report(path)

    def preview_report(self, path):
        if not path or not os.path.isfile(path):
            messagebox.showwarning("Uyari", "Dosya bulunamadi.")
            return
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
        except Exception as exc:
            messagebox.showerror("Hata", f"Dosya acilamadi: {exc}")
            return

        preview = tk.Toplevel(self)
        preview.title(f"Rapor Goruntule - {os.path.basename(path)}")
        preview.geometry("1100x700")

        frame = ttk.Frame(preview)
        frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        max_cols = min(ws.max_column, 25)
        max_rows = min(ws.max_row, 200)
        headers = [ws.cell(row=1, column=c).value or f"C{c}" for c in range(1, max_cols + 1)]
        tree = ttk.Treeview(frame, columns=list(range(max_cols)), show="headings")
        for idx, header in enumerate(headers):
            tree.heading(idx, text=str(header))
            tree.column(idx, width=120)
        tree.pack(fill=tk.BOTH, expand=True)

        yscroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        xscroll = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(xscrollcommand=xscroll.set)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)

        for r in range(2, max_rows + 1):
            row_vals = []
            for c in range(1, max_cols + 1):
                val = ws.cell(row=r, column=c).value
                row_vals.append("" if val is None else val)
            tree.insert("", tk.END, values=row_vals)

    def refresh_admin_summary(self):
        if not hasattr(self, "admin_tree"):
            return
        for item in self.admin_tree.get_children():
            self.admin_tree.delete(item)
        if hasattr(self, "admin_alert_tree"):
            for item in self.admin_alert_tree.get_children():
                self.admin_alert_tree.delete(item)
        if hasattr(self, "admin_anomaly_tree"):
            for item in self.admin_anomaly_tree.get_children():
                self.admin_anomaly_tree.delete(item)

        employee_name = self.admin_employee_var.get().strip()
        employee_id = None
        if employee_name and employee_name != "Tum Calisanlar":
            base, region = split_display_name(employee_name, REGIONS)
            if region is None:
                employee_id = self.employee_map.get((base, "")) or self.employee_map.get(
                    (base, self._entry_region())
                )
            else:
                employee_id = self.employee_map.get((base, region))
        department_filter = self.admin_department_var.get().strip()
        title_filter = self.admin_title_var.get().strip()
        search_text = self.admin_search_var.get().strip().lower()
        start_date = self.admin_start_var.get().strip() or None
        end_date = self.admin_end_var.get().strip() or None
        try:
            if start_date:
                start_date = normalize_date(start_date)
            if end_date:
                end_date = normalize_date(end_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return

        records = db.list_timesheets(
            employee_id=employee_id,
            start_date=start_date,
            end_date=end_date,
            region=self._view_region(),
        )
        totals = {}
        daily_overtime = {}
        dept_overtime = {}
        alerts = []
        work_days = {}
        for _ts_id, _emp_id, name, work_date, start_time, end_time, break_minutes, is_special, _notes, _region in records:
            details = self.employee_details.get((name, _region or ""), {})
            department = details.get("department", "")
            title = details.get("title", "")
            if department_filter and department_filter != "Tum Departmanlar" and department != department_filter:
                continue
            if title_filter and title_filter != "Tum Unvanlar" and title != title_filter:
                continue
            if search_text:
                hay = " ".join([name, department, title, str(_notes or "")]).lower()
                if search_text not in hay:
                    continue

            (
                worked,
                _scheduled,
                overtime,
                night_hours,
                overnight_hours,
                spec_norm,
                spec_ot,
                spec_night,
            ) = calc.calc_day_hours(
                work_date,
                start_time,
                end_time,
                break_minutes,
                self.settings,
                is_special,
            )
            key = (name, _region or "")
            if key not in totals:
                totals[key] = {
                    "name": name,
                    "region": _region or "",
                    "worked": 0.0,
                    "overtime": 0.0,
                    "night": 0.0,
                    "overnight": 0.0,
                    "special": 0.0,
                }
            totals[key]["worked"] += worked
            totals[key]["overtime"] += overtime
            totals[key]["night"] += night_hours
            totals[key]["overnight"] += overnight_hours
            totals[key]["special"] += spec_norm + spec_ot + spec_night

            daily_overtime[work_date] = daily_overtime.get(work_date, 0.0) + overtime
            dept_key = department or "Bilinmeyen"
            dept_overtime[dept_key] = dept_overtime.get(dept_key, 0.0) + overtime

            try:
                gross_hours = calc.hours_between(calc.parse_time(start_time), calc.parse_time(end_time))
            except Exception:
                gross_hours = 0.0
            if gross_hours >= 12:
                alerts.append((work_date, name, "Uzun Mesai", f"{gross_hours:.1f}s"))
            if overnight_hours > 0:
                alerts.append((work_date, name, "Geceye Tasan", f"{overnight_hours:.1f}s"))
            if is_special:
                alerts.append((work_date, name, "Ozel Gun", f"{worked:.1f}s"))

            work_days.setdefault(name, set()).add(work_date)

        total_worked = sum(v["worked"] for v in totals.values())
        total_overtime = sum(v["overtime"] for v in totals.values())
        total_night = sum(v["night"] for v in totals.values())
        total_overnight = sum(v["overnight"] for v in totals.values())
        total_special = sum(v["special"] for v in totals.values())

        self.admin_stats["total_records"].set(str(len(records)))
        self.admin_stats["total_employees"].set(str(len(totals)))
        self.admin_stats["total_worked"].set(f"{total_worked:.2f}")
        self.admin_stats["total_overtime"].set(f"{total_overtime:.2f}")
        self.admin_stats["total_night"].set(f"{total_night:.2f}")
        self.admin_stats["total_overnight"].set(f"{total_overnight:.2f}")
        self.admin_stats["total_special"].set(f"{total_special:.2f}")
        avg_ot = (total_overtime / len(totals)) if totals else 0.0
        self.admin_stats["avg_overtime"].set(f"{avg_ot:.2f}")
        max_daily = max(daily_overtime.values()) if daily_overtime else 0.0
        self.admin_stats["max_daily"].set(f"{max_daily:.2f}")

        for _key, data in sorted(totals.items(), key=lambda x: x[1]["name"]):
            display_name = data["name"]
            if data.get("region"):
                display_name = f"{data['name']} ({data['region']})"
            self.admin_tree.insert(
                "",
                tk.END,
                values=(
                    display_name,
                    round(data["worked"], 2),
                    round(data["overtime"], 2),
                    round(data["night"], 2),
                    round(data["overnight"], 2),
                    round(data["special"], 2),
                ),
            )

        if hasattr(self, "admin_alert_tree"):
            for work_date, name, issue, value in alerts[:200]:
                self.admin_alert_tree.insert("", tk.END, values=(work_date, name, issue, value))

        if hasattr(self, "admin_anomaly_tree"):
            anomalies = self._build_consecutive_day_anomalies(work_days)
            for name, period, issue in anomalies:
                self.admin_anomaly_tree.insert("", tk.END, values=(name, period, issue))

    def _build_consecutive_day_anomalies(self, work_days):
        anomalies = []
        for name, dates in work_days.items():
            try:
                sorted_dates = sorted(datetime.strptime(d, "%Y-%m-%d").date() for d in dates)
            except Exception:
                continue
            streak = 1
            start = sorted_dates[0] if sorted_dates else None
            for i in range(1, len(sorted_dates)):
                if (sorted_dates[i] - sorted_dates[i - 1]).days == 1:
                    streak += 1
                else:
                    if streak >= 6 and start:
                        anomalies.append((name, f"{start} - {sorted_dates[i-1]}", f"{streak} gun ust uste"))
                    streak = 1
                    start = sorted_dates[i]
            if streak >= 6 and start:
                anomalies.append((name, f"{start} - {sorted_dates[-1]}", f"{streak} gun ust uste"))
        return anomalies


    def package_monthly_reports(self):
        month_text = self.admin_month_var.get().strip()
        try:
            parse_month(month_text)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        logs = db.list_report_logs()
        files = [r[1] for r in logs if r[2].startswith(month_text)]
        files = [f for f in files if os.path.isfile(f)]
        if not files:
            messagebox.showinfo("Bilgi", "Secilen ay icin rapor bulunamadi.")
            return
        output_path = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("ZIP", "*.zip")],
            initialfile=f"raporlar_{month_text}.zip",
        )
        if not output_path:
            return
        try:
            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for file_path in files:
                    zf.write(file_path, arcname=os.path.basename(file_path))
        except Exception as exc:
            messagebox.showerror("Hata", f"Zip olusturulamadi: {exc}")
            return
        messagebox.showinfo("Basarili", f"Zip kaydedildi: {output_path}")

    def refresh_vehicles(self):
        if not hasattr(self, "vehicle_tree"):
            return
        for item in self.vehicle_tree.get_children():
            self.vehicle_tree.delete(item)
        self.vehicle_map = {}
        km = None
        oil_change_km = None
        oil_interval_km = None
        for vehicle in db.list_vehicles(region=self._view_region()):
            (
                vehicle_id,
                plate,
                brand,
                model,
                year,
                km,
                inspection_date,
                insurance_date,
                maintenance_date,
                oil_change_date,
                oil_change_km,
                oil_interval_km,
                _notes,
                region,
            ) = vehicle
            oil_status = "-"
            interval_km = oil_interval_km or DEFAULT_OIL_INTERVAL_KM
            if interval_km and oil_change_km is not None and km is not None:
                remaining = interval_km - (km - oil_change_km)
                oil_status = "Geldi" if remaining <= 0 else f"{remaining} km"
            self.vehicle_tree.insert(
                "",
                tk.END,
                values=(
                    vehicle_id,
                    plate,
                    brand,
                    model,
                    year,
                    km,
                    inspection_date,
                    insurance_date,
                    maintenance_date,
                    oil_status,
                    region or "",
                ),
            )
            self.vehicle_map[plate] = vehicle_id
        if hasattr(self, "inspect_vehicle_combo"):
            self.inspect_vehicle_combo["values"] = sorted(self.vehicle_map.keys())
        if hasattr(self, "fault_vehicle_combo"):
            self.fault_vehicle_combo["values"] = sorted(self.vehicle_map.keys())
        if hasattr(self, "service_vehicle_combo"):
            self.service_vehicle_combo["values"] = sorted(self.vehicle_map.keys())

    def refresh_drivers(self):
        if not hasattr(self, "driver_tree"):
            return
        for item in self.driver_tree.get_children():
            self.driver_tree.delete(item)
        self.driver_map = {}
        self.driver_display_names = []
        name_counts = {}
        drivers = db.list_drivers(region=self._view_region())
        for driver in drivers:
            driver_id, name, license_class, license_expiry, phone, _notes, region = driver
            name_counts[name] = name_counts.get(name, 0) + 1
            self.driver_tree.insert(
                "",
                tk.END,
                values=(driver_id, name, license_class, license_expiry, phone, region or ""),
            )
            self.driver_map[(name, region or "")] = driver_id
        for driver in drivers:
            _driver_id, name, _license_class, _license_expiry, _phone, _notes, region = driver
            display = name
            if name_counts.get(name, 0) > 1:
                display = f"{name} ({region or '-'})"
            self.driver_display_names.append(display)
        if hasattr(self, "inspect_driver_combo"):
            self.inspect_driver_combo["values"] = sorted(self.driver_display_names)

    def refresh_faults(self):
        if hasattr(self, "fault_tree"):
            for item in self.fault_tree.get_children():
                self.fault_tree.delete(item)
        self.fault_map = {}
        self.fault_display_by_id = {}
        for fault in db.list_vehicle_faults(region=self._view_region()):
            fault_id, vehicle_id, plate, title, desc, opened_date, closed_date, status, region = fault
            if hasattr(self, "fault_tree"):
                self.fault_tree.insert(
                    "",
                    tk.END,
                    values=(fault_id, plate, title, status, opened_date or "", closed_date or "", region or ""),
                )
            display = f"{plate} - {title} (#{fault_id})"
            self.fault_map[display] = fault_id
            self.fault_display_by_id[fault_id] = display
        if hasattr(self, "inspect_fault_combo"):
            values = [""] + list(self.fault_map.keys())
            self.inspect_fault_combo["values"] = values
        if hasattr(self, "service_fault_combo"):
            values = [""] + list(self.fault_map.keys())
            self.service_fault_combo["values"] = values

    def refresh_service_visits(self):
        if not hasattr(self, "service_tree"):
            return
        for item in self.service_tree.get_children():
            self.service_tree.delete(item)
        self.service_visit_map = {}
        for visit in db.list_vehicle_service_visits(region=self._view_region()):
            (
                visit_id,
                _vehicle_id,
                plate,
                _fault_id,
                fault_title,
                start_date,
                end_date,
                reason,
                cost,
                _notes,
                region,
            ) = visit
            self.service_tree.insert(
                "",
                tk.END,
                values=(
                    visit_id,
                    plate,
                    fault_title or "",
                    start_date,
                    end_date or ("Sanayide" if end_date is None or end_date == "" else ""),
                    f"{cost:.2f}" if cost is not None else "",
                    reason or "",
                    region or "",
                ),
            )
            self.service_visit_map[visit_id] = visit

    def clear_fault_form(self):
        self.fault_id_var.set("")
        self.fault_vehicle_var.set("")
        self.fault_title_var.set("")
        self.fault_desc_var.set("")
        self.fault_open_var.set("")
        self.fault_close_var.set("")
        self.fault_status_var.set("Acik")
        if hasattr(self, "fault_open_entry"):
            clear_date_entry(self.fault_open_entry)
        if hasattr(self, "fault_close_entry"):
            clear_date_entry(self.fault_close_entry)

    def add_or_update_fault(self):
        plate = self.fault_vehicle_var.get().strip()
        if not plate:
            messagebox.showwarning("Uyari", "Arac secin.")
            return
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        title = self.fault_title_var.get().strip()
        if not title:
            messagebox.showwarning("Uyari", "Baslik zorunlu.")
            return
        opened_date = self.fault_open_var.get().strip()
        closed_date = self.fault_close_var.get().strip()
        if opened_date:
            try:
                opened_date = normalize_date(opened_date)
            except ValueError as exc:
                messagebox.showwarning("Uyari", str(exc))
                return
        if closed_date:
            try:
                closed_date = normalize_date(closed_date)
            except ValueError as exc:
                messagebox.showwarning("Uyari", str(exc))
                return
        status = self.fault_status_var.get().strip() or "Acik"
        desc = self.fault_desc_var.get().strip()
        fault_id = parse_int(self.fault_id_var.get())
        if fault_id:
            db.update_vehicle_fault(
                fault_id,
                vehicle_id,
                title,
                desc,
                opened_date,
                closed_date,
                status,
                self._entry_region(),
            )
            self._log_action("fault_update", f"id={fault_id} plate={plate} status={status}")
        else:
            db.add_vehicle_fault(
                vehicle_id,
                title,
                desc,
                opened_date,
                closed_date,
                status,
                self._entry_region(),
            )
            self._log_action("fault_add", f"plate={plate} title={title} status={status}")
        self.refresh_faults()
        self.refresh_vehicle_dashboard()
        self.clear_fault_form()
        messagebox.showinfo("Basarili", "Ariza kaydi kaydedildi.")
        self.trigger_sync("fault")

    def delete_fault(self):
        fault_id = parse_int(self.fault_id_var.get())
        if not fault_id:
            selected = self.fault_tree.selection() if hasattr(self, "fault_tree") else None
            if selected:
                values = self.fault_tree.item(selected[0], "values")
                fault_id = parse_int(values[0])
        if not fault_id:
            messagebox.showwarning("Uyari", "Silinecek ariza secin.")
            return
        if not messagebox.askyesno("Onay", "Ariza kaydi silinsin mi?"):
            return
        db.delete_vehicle_fault(fault_id)
        self._log_action("fault_delete", f"id={fault_id}")
        self.refresh_faults()
        self.refresh_vehicle_dashboard()
        self.clear_fault_form()
        self.trigger_sync("fault_delete")

    def on_fault_select(self, _event=None):
        selected = self.fault_tree.selection()
        if not selected:
            return
        values = self.fault_tree.item(selected[0], "values")
        fault_id = parse_int(values[0])
        fault = db.get_vehicle_fault(fault_id)
        if not fault:
            return
        (
            _fid,
            vehicle_id,
            title,
            desc,
            opened_date,
            closed_date,
            status,
        ) = fault
        plate = None
        for plate_name, vid in self.vehicle_map.items():
            if vid == vehicle_id:
                plate = plate_name
                break
        if plate:
            self.fault_vehicle_var.set(plate)
        self.fault_id_var.set(fault_id)
        self.fault_title_var.set(title or "")
        self.fault_desc_var.set(desc or "")
        self.fault_open_var.set(opened_date or "")
        self.fault_close_var.set(closed_date or "")
        self.fault_status_var.set(status or "Acik")

    def clear_service_visit_form(self):
        self.service_id_var.set("")
        self.service_vehicle_var.set("")
        self.service_fault_var.set("")
        self.service_start_var.set("")
        self.service_end_var.set("")
        self.service_reason_var.set("")
        self.service_cost_var.set("")
        self.service_notes_var.set("")
        self.service_in_shop_var.set(False)
        self._toggle_service_end_date()
        if hasattr(self, "service_start_entry"):
            clear_date_entry(self.service_start_entry)
        if hasattr(self, "service_end_entry"):
            clear_date_entry(self.service_end_entry)

    def add_or_update_service_visit(self):
        plate = self.service_vehicle_var.get().strip()
        if not plate:
            messagebox.showwarning("Uyari", "Arac secin.")
            return
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        fault_display = self.service_fault_var.get().strip()
        fault_id = self.fault_map.get(fault_display) if fault_display else None
        start_date = self.service_start_var.get().strip()
        if not start_date:
            messagebox.showwarning("Uyari", "Gidis tarihi zorunlu.")
            return
        try:
            start_date = normalize_date(start_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        end_date = self.service_end_var.get().strip()
        if self.service_in_shop_var.get():
            end_date = ""
        elif end_date:
            try:
                end_date = normalize_date(end_date)
            except ValueError as exc:
                messagebox.showwarning("Uyari", str(exc))
                return
        reason = self.service_reason_var.get().strip()
        notes = self.service_notes_var.get().strip()
        cost = self.service_cost_var.get().strip()
        cost_val = None if cost == "" else parse_float(cost, 0.0)
        visit_id = parse_int(self.service_id_var.get())
        if visit_id:
            db.update_vehicle_service_visit(
                visit_id,
                vehicle_id,
                fault_id,
                start_date,
                end_date,
                reason,
                cost_val,
                notes,
                self._entry_region(),
            )
            self._log_action("service_visit_update", f"id={visit_id} plate={plate}")
        else:
            db.add_vehicle_service_visit(
                vehicle_id,
                fault_id,
                start_date,
                end_date,
                reason,
                cost_val,
                notes,
                self._entry_region(),
            )
            self._log_action("service_visit_add", f"plate={plate} start={start_date}")
        self.refresh_service_visits()
        self.refresh_vehicle_dashboard()
        self.clear_service_visit_form()
        messagebox.showinfo("Basarili", "Sanayi kaydi kaydedildi.")
        self.trigger_sync("service_visit")

    def delete_service_visit(self):
        visit_id = parse_int(self.service_id_var.get())
        if not visit_id:
            selected = self.service_tree.selection() if hasattr(self, "service_tree") else None
            if selected:
                values = self.service_tree.item(selected[0], "values")
                visit_id = parse_int(values[0])
        if not visit_id:
            messagebox.showwarning("Uyari", "Silinecek kaydi secin.")
            return
        if not messagebox.askyesno("Onay", "Sanayi kaydi silinsin mi?"):
            return
        db.delete_vehicle_service_visit(visit_id)
        self._log_action("service_visit_delete", f"id={visit_id}")
        self.refresh_service_visits()
        self.refresh_vehicle_dashboard()
        self.clear_service_visit_form()
        self.trigger_sync("service_visit_delete")

    def on_service_visit_select(self, _event=None):
        selected = self.service_tree.selection()
        if not selected:
            return
        values = self.service_tree.item(selected[0], "values")
        visit_id = parse_int(values[0])
        visit = db.get_vehicle_service_visit(visit_id)
        if not visit:
            return
        (
            _vid,
            vehicle_id,
            fault_id,
            start_date,
            end_date,
            reason,
            cost,
            notes,
        ) = visit
        plate = None
        for plate_name, vid in self.vehicle_map.items():
            if vid == vehicle_id:
                plate = plate_name
                break
        if plate:
            self.service_vehicle_var.set(plate)
        fault_display = ""
        if fault_id:
            for display, fid in self.fault_map.items():
                if fid == fault_id:
                    fault_display = display
                    break
        self.service_fault_var.set(fault_display)
        self.service_id_var.set(visit_id)
        self.service_start_var.set(start_date or "")
        self.service_end_var.set(end_date or "")
        self.service_in_shop_var.set(False if end_date else True)
        self._toggle_service_end_date()
        self.service_reason_var.set(reason or "")
        self.service_cost_var.set("" if cost is None else f"{cost:.2f}")
        self.service_notes_var.set(notes or "")

    def _toggle_service_end_date(self):
        state = "disabled" if self.service_in_shop_var.get() else "normal"
        if hasattr(self, "service_end_entry"):
            try:
                self.service_end_entry.configure(state=state)
            except tk.TclError:
                pass

    def refresh_vehicle_dashboard(self):
        if not hasattr(self, "vehicle_status_tree"):
            return
        for item in self.vehicle_status_tree.get_children():
            self.vehicle_status_tree.delete(item)
        for item in self.vehicle_alert_tree.get_children():
            self.vehicle_alert_tree.delete(item)
        for item in self.driver_alert_tree.get_children():
            self.driver_alert_tree.delete(item)

        vehicles = db.list_vehicles(region=self._view_region())
        drivers = db.list_drivers(region=self._view_region())
        driver_by_id = {row[0]: row for row in drivers}
        driver_latest = {}

        # Populate vehicle_map for alert clicks
        self.vehicle_map = {}

        oil_due = 0
        insp_due = 0
        ins_due = 0
        maint_due = 0
        lic_due = 0

        for vehicle in vehicles:
            (
                _vid,
                plate,
                brand,
                model,
                year,
                km,
                inspection_date,
                insurance_date,
                maintenance_date,
                oil_change_date,
                oil_change_km,
                oil_interval_km,
                _notes,
                region,
            ) = vehicle
            # Map plate to vehicle ID for alert clicks
            self.vehicle_map[plate] = _vid
            oil_status = "-"
            oil_flag = None
            interval_km = oil_interval_km or DEFAULT_OIL_INTERVAL_KM
            if interval_km and oil_change_km is not None and km is not None:
                remaining = interval_km - (km - oil_change_km)
                oil_status = "Geldi" if remaining <= 0 else f"{remaining} km"
                if remaining <= 0:
                    oil_due += 1
                    oil_flag = "oil_due"
                    self.vehicle_alert_tree.insert("", tk.END, values=(plate, "Yag Degisimi", "Geldi"))
                elif remaining <= DEFAULT_OIL_SOON_KM:
                    oil_flag = "oil_soon"

            insp_days = days_until(inspection_date)
            if insp_days is not None and insp_days <= 30:
                insp_due += 1
                detail = f"{inspection_date} ({insp_days} gun)"
                if insp_days < 0:
                    detail = f"{inspection_date} ({abs(insp_days)} gun gecikme)"
                self.vehicle_alert_tree.insert("", tk.END, values=(plate, "Muayene", detail))

            ins_days = days_until(insurance_date)
            if ins_days is not None and ins_days <= 30:
                ins_due += 1
                detail = f"{insurance_date} ({ins_days} gun)"
                if ins_days < 0:
                    detail = f"{insurance_date} ({abs(ins_days)} gun gecikme)"
                self.vehicle_alert_tree.insert("", tk.END, values=(plate, "Sigorta", detail))

            maint_days = days_until(maintenance_date)
            if maint_days is not None and maint_days <= 30:
                maint_due += 1
                detail = f"{maintenance_date} ({maint_days} gun)"
                if maint_days < 0:
                    detail = f"{maintenance_date} ({abs(maint_days)} gun gecikme)"
                self.vehicle_alert_tree.insert("", tk.END, values=(plate, "Bakim", detail))

            inspections = db.list_vehicle_inspections(vehicle_id=_vid, region=self._view_region())
            last_check = "-"
            last_driver = "-"
            if inspections:
                last_inspection = inspections[0]
                last_check = last_inspection[5]
                last_driver = last_inspection[4] or "-"
                driver_id = last_inspection[3]
                if driver_id:
                    current = driver_latest.get(driver_id)
                    if not current or last_inspection[5] > current[5]:
                        driver_latest[driver_id] = last_inspection

            if len(inspections) >= 2:
                current_inspection = inspections[0]
                previous_inspection = inspections[1]
                current_results = {
                    row[0]: normalize_vehicle_status(row[1])
                    for row in db.list_vehicle_inspection_results(current_inspection[0])
                }
                prev_results = {
                    row[0]: normalize_vehicle_status(row[1])
                    for row in db.list_vehicle_inspection_results(previous_inspection[0])
                }
                for item_key, label in VEHICLE_CHECKLIST:
                    if (
                        current_results.get(item_key) == "Olumsuz"
                        and prev_results.get(item_key) == "Olumsuz"
                    ):
                        self.vehicle_alert_tree.insert(
                            "",
                            tk.END,
                            values=(plate, "Tekrar Eden Sorun", f"{label} (2 hafta)"),
                        )

            self.vehicle_status_tree.insert(
                "",
                tk.END,
                values=(
                    plate,
                    km or "-",
                    oil_status,
                    inspection_date or "-",
                    insurance_date or "-",
                    maintenance_date or "-",
                    last_check,
                    last_driver,
                    region or "",
                ),
                tags=(oil_flag,) if oil_flag else (),
            )

        for driver in drivers:
            _did, name, _cls, license_expiry, _phone, _notes, _region = driver
            days = days_until(license_expiry)
            if days is not None and days <= 30:
                lic_due += 1
                detail = f"{license_expiry} ({days} gun)"
                if days < 0:
                    detail = f"{license_expiry} ({abs(days)} gun gecikme)"
                self.driver_alert_tree.insert("", tk.END, values=(name, "Ehliyet", detail))

        faults = db.list_vehicle_faults(region=self._view_region())
        now = datetime.now().date()
        faults_by_plate = {}
        for fault in faults:
            _fid, _vid, plate, title, _desc, opened_date, _closed_date, status, _region = fault
            faults_by_plate.setdefault(plate, []).append(fault)
            if status == "Acik":
                self.vehicle_alert_tree.insert("", tk.END, values=(plate, "Acik Ariza", title))

        for plate, items in faults_by_plate.items():
            title_counts = {}
            for fault in items:
                opened_date = fault[5]
                if not opened_date:
                    continue
                try:
                    opened_dt = datetime.strptime(opened_date, "%Y-%m-%d").date()
                except ValueError:
                    continue
                if (now - opened_dt).days <= 30:
                    title_counts[fault[3]] = title_counts.get(fault[3], 0) + 1
            for title, count in title_counts.items():
                if count >= 2:
                    self.vehicle_alert_tree.insert(
                        "",
                        tk.END,
                        values=(plate, "Tekrar Ariza (30 gun)", f"{title} x{count}"),
                    )

        fault_counts = {}
        for fault in faults:
            plate = fault[2]
            fault_counts[plate] = fault_counts.get(plate, 0) + 1
        top_faults = sorted(fault_counts.items(), key=lambda x: x[1], reverse=True)[:3]
        for plate, count in top_faults:
            self.vehicle_alert_tree.insert("", tk.END, values=(plate, "En Cok Ariza", f"{count} kayit"))

        for driver_id, inspection in driver_latest.items():
            driver = driver_by_id.get(driver_id)
            if not driver:
                continue
            name = driver[1]
            plate = inspection[2]
            results = db.list_vehicle_inspection_results(inspection[0])
            bad_items = []
            for item_key, label in VEHICLE_CHECKLIST:
                for res_key, status, _note in results:
                    if res_key == item_key and normalize_vehicle_status(status) == "Olumsuz":
                        bad_items.append(label)
                        break
            if bad_items:
                detail = f"{plate} - {len(bad_items)} olumsuz"
                self.driver_alert_tree.insert("", tk.END, values=(name, "Son Kontrol Sorun", detail))

        self.dashboard_stats["vehicles"].set(str(len(vehicles)))
        self.dashboard_stats["drivers"].set(str(len(drivers)))
        self.dashboard_stats["oil_due"].set(str(oil_due))
        self.dashboard_stats["inspection_due"].set(str(insp_due))
        self.dashboard_stats["insurance_due"].set(str(ins_due))
        self.dashboard_stats["maintenance_due"].set(str(maint_due))
        self.dashboard_stats["license_due"].set(str(lic_due))

    def clear_vehicle_form(self):
        self.vehicle_id_var.set("")
        self.vehicle_plate_var.set("")
        self.vehicle_brand_var.set("")
        self.vehicle_model_var.set("")
        self.vehicle_year_var.set("")
        self.vehicle_km_var.set("")
        self.vehicle_inspection_var.set("")
        self.vehicle_insurance_var.set("")
        self.vehicle_maintenance_var.set("")
        self.vehicle_oil_var.set("")
        self.vehicle_oil_km_var.set("")
        self.vehicle_oil_interval_var.set(str(DEFAULT_OIL_INTERVAL_KM))
        self.vehicle_notes_var.set("")
        self.vehicle_original_plate = None
        clear_date_entry(self.vehicle_insp_entry)
        clear_date_entry(self.vehicle_ins_entry)
        clear_date_entry(self.vehicle_maint_entry)
        clear_date_entry(self.vehicle_oil_entry)

    def on_vehicle_select(self, _event=None):
        selected = self.vehicle_tree.selection()
        if not selected:
            return
        values = self.vehicle_tree.item(selected[0], "values")
        vehicle_id = parse_int(values[0])
        row = db.get_vehicle(vehicle_id)
        if not row:
            return
        (
            _vid,
            plate,
            brand,
            model,
            year,
            km,
            inspection_date,
            insurance_date,
            maintenance_date,
            oil_change_date,
            oil_change_km,
            oil_interval_km,
            _notes,
            _region,
        ) = row
        self.vehicle_id_var.set(vehicle_id)
        self.vehicle_plate_var.set(plate)
        self.vehicle_original_plate = plate
        self.vehicle_brand_var.set(brand)
        self.vehicle_model_var.set(model)
        self.vehicle_year_var.set(year)
        self.vehicle_km_var.set(km or "")
        self.vehicle_inspection_var.set(inspection_date or "")
        self.vehicle_insurance_var.set(insurance_date or "")
        self.vehicle_maintenance_var.set(maintenance_date or "")
        self.vehicle_oil_var.set(oil_change_date or "")
        self.vehicle_oil_km_var.set(oil_change_km or "")
        self.vehicle_oil_interval_var.set(oil_interval_km or str(DEFAULT_OIL_INTERVAL_KM))

    def add_or_update_vehicle(self):
        plate = self.vehicle_plate_var.get().strip()
        if not plate:
            messagebox.showwarning("Uyari", "Plaka zorunlu.")
            return
        brand = self.vehicle_brand_var.get().strip()
        model = self.vehicle_model_var.get().strip()
        year = self.vehicle_year_var.get().strip()
        km = parse_int(self.vehicle_km_var.get(), 0)
        inspection_date = self.vehicle_inspection_var.get().strip()
        insurance_date = self.vehicle_insurance_var.get().strip()
        maintenance_date = self.vehicle_maintenance_var.get().strip()
        oil_change_date = self.vehicle_oil_var.get().strip()
        oil_change_km = parse_int(self.vehicle_oil_km_var.get(), 0)
        oil_interval_km = parse_int(self.vehicle_oil_interval_var.get(), 0)
        notes = self.vehicle_notes_var.get().strip()
        try:
            inspection_date = normalize_date(inspection_date) if inspection_date else ""
            insurance_date = normalize_date(insurance_date) if insurance_date else ""
            maintenance_date = normalize_date(maintenance_date) if maintenance_date else ""
            oil_change_date = normalize_date(oil_change_date) if oil_change_date else ""
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return

        vehicle_id = self.vehicle_id_var.get().strip()
        if vehicle_id and self.vehicle_original_plate and plate != self.vehicle_original_plate:
            add_new = messagebox.askyesno(
                "Onay",
                "Plaka degisti. Yeni arac olarak eklensin mi?\n"
                "Evet: yeni kayit, Hayir: mevcut araci guncelle.",
            )
            if add_new:
                vehicle_id = ""
        try:
            if vehicle_id:
                db.update_vehicle(
                    parse_int(vehicle_id),
                    plate,
                    brand,
                    model,
                    year,
                    km,
                    inspection_date,
                    insurance_date,
                    maintenance_date,
                    oil_change_date,
                    oil_change_km,
                    oil_interval_km,
                    notes,
                    self._entry_region(),
                )
                self._log_action("vehicle_update", f"id={vehicle_id} plate={plate}")
            else:
                db.add_vehicle(
                    plate,
                    brand,
                    model,
                    year,
                    km,
                    inspection_date,
                    insurance_date,
                    maintenance_date,
                    oil_change_date,
                    oil_change_km,
                    oil_interval_km,
                    notes,
                    self._entry_region(),
                )
                self._log_action("vehicle_add", f"plate={plate}")
        except Exception:
            messagebox.showwarning("Uyari", "Arac kaydi eklenemedi. Plaka zaten var olabilir.")
            return
        self.refresh_vehicles()
        self.clear_vehicle_form()
        self.trigger_sync("vehicle")

    def delete_vehicle(self):
        vehicle_id = self.vehicle_id_var.get().strip()
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Silmek icin arac secin.")
            return
        if messagebox.askyesno("Onay", "Araci silmek istiyor musunuz?"):
            db.delete_vehicle(parse_int(vehicle_id))
            self._log_action("vehicle_delete", f"id={vehicle_id}")
            self.refresh_vehicles()
            self.clear_vehicle_form()
            self.trigger_sync("vehicle_delete")

    def clear_driver_form(self):
        self.driver_id_var.set("")
        self.driver_name_var.set("")
        self.driver_license_var.set("")
        self.driver_license_exp_var.set("")
        self.driver_phone_var.set("")
        self.driver_notes_var.set("")
        clear_date_entry(self.driver_exp_entry)

    def on_driver_select(self, _event=None):
        selected = self.driver_tree.selection()
        if not selected:
            return
        values = self.driver_tree.item(selected[0], "values")
        self.driver_id_var.set(values[0])
        self.driver_name_var.set(values[1])
        self.driver_license_var.set(values[2])
        self.driver_license_exp_var.set(values[3] or "")
        self.driver_phone_var.set(values[4])

    def add_or_update_driver(self):
        name = self.driver_name_var.get().strip()
        if not name:
            messagebox.showwarning("Uyari", "Ad Soyad zorunlu.")
            return
        license_class = self.driver_license_var.get().strip()
        license_expiry = self.driver_license_exp_var.get().strip()
        phone = self.driver_phone_var.get().strip()
        notes = self.driver_notes_var.get().strip()
        try:
            license_expiry = normalize_date(license_expiry) if license_expiry else ""
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        driver_id = self.driver_id_var.get().strip()
        if driver_id:
            db.update_driver(
                parse_int(driver_id),
                name,
                license_class,
                license_expiry,
                phone,
                notes,
                self._entry_region(),
            )
            self._log_action("driver_update", f"id={driver_id} name={name}")
        else:
            db.add_driver(name, license_class, license_expiry, phone, notes, self._entry_region())
            self._log_action("driver_add", f"name={name}")
        self.refresh_drivers()
        self.clear_driver_form()
        self.trigger_sync("driver")

    def delete_driver(self):
        driver_id = self.driver_id_var.get().strip()
        if not driver_id:
            messagebox.showwarning("Uyari", "Silmek icin surucu secin.")
            return
        if messagebox.askyesno("Onay", "Surucuyu silmek istiyor musunuz?"):
            db.delete_driver(parse_int(driver_id))
            self._log_action("driver_delete", f"id={driver_id}")
            self.refresh_drivers()
            self.clear_driver_form()
            self.trigger_sync("driver_delete")

    def save_vehicle_inspection(self):
        plate = self.inspect_vehicle_var.get().strip()
        if not plate:
            messagebox.showwarning("Uyari", "Arac secin.")
            return
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        driver_name = self.inspect_driver_var.get().strip()
        driver_id = None
        if driver_name:
            base, region = split_display_name(driver_name, REGIONS)
            if region is None:
                driver_id = self.driver_map.get((base, "")) or self.driver_map.get((base, self._entry_region()))
            else:
                driver_id = self.driver_map.get((base, region))
        inspect_date = self.inspect_date_var.get().strip()
        if not inspect_date:
            messagebox.showwarning("Uyari", "Tarih zorunlu.")
            return
        inspect_km = parse_int(self.inspect_km_var.get(), 0)
        try:
            inspect_date = normalize_date(inspect_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        week_start = week_start_from_date(inspect_date)
        notes = self.inspect_notes_var.get().strip()
        fault_display = self.inspect_fault_var.get().strip()
        fault_id = self.fault_map.get(fault_display) if fault_display else None
        fault_status = self.inspect_fault_status_var.get().strip() if fault_id else None
        service_visit = 1 if self.inspect_service_var.get() else 0
        if service_visit and not fault_id:
            messagebox.showwarning("Uyari", "Sanayi icin ariza secin.")
            return
        inspection_id = db.add_vehicle_inspection(
            vehicle_id,
            driver_id,
            inspect_date,
            week_start,
            inspect_km,
            notes,
            fault_id=fault_id,
            fault_status=fault_status,
            service_visit=service_visit,
        )
        self._log_action(
            "vehicle_inspection_add",
            f"plate={plate} date={inspect_date} km={inspect_km}",
        )
        for item_key, _label in VEHICLE_CHECKLIST:
            status = self.inspect_item_vars[item_key].get()
            note = self.inspect_note_vars[item_key].get().strip()
            db.add_vehicle_inspection_result(inspection_id, item_key, status, note)
        if inspect_km:
            current = db.get_vehicle(vehicle_id)
            if current:
                (
                    _vid,
                    plate,
                    brand,
                    model,
                    year,
                    _km,
                    inspection_date,
                    insurance_date,
                    maintenance_date,
                    oil_change_date,
                    oil_change_km,
                    oil_interval_km,
                    notes,
                    _region,
                ) = current
                db.update_vehicle(
                    vehicle_id,
                    plate,
                    brand,
                    model,
                    year,
                    inspect_km,
                    inspection_date or "",
                    insurance_date or "",
                    maintenance_date or "",
                    oil_change_date or "",
                    oil_change_km or 0,
                    oil_interval_km or 0,
                    notes or "",
                    self._entry_region(),
                )
                self.refresh_vehicles()
        if fault_id and fault_status == "Kapandi":
            fault = db.get_vehicle_fault(fault_id)
            if fault:
                (
                    _fid,
                    f_vehicle_id,
                    title,
                    desc,
                    opened_date,
                    closed_date,
                    status,
                ) = fault
                if not closed_date:
                    closed_date = inspect_date
                db.update_vehicle_fault(
                    fault_id,
                    f_vehicle_id,
                    title,
                    desc,
                    opened_date,
                    closed_date,
                    "Kapandi",
                    self._entry_region(),
                )
                self.refresh_faults()
        messagebox.showinfo("Basarili", "Haftalik kontrol kaydedildi.")
        self.trigger_sync("vehicle_inspection")

    def on_inspect_vehicle_change(self, _event=None):
        plate = self.inspect_vehicle_var.get().strip()
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            return
        faults = db.list_open_vehicle_faults(vehicle_id=vehicle_id, region=self._view_region())
        if not faults:
            self.inspect_fault_var.set("")
            self.inspect_fault_status_var.set("Acik")
            return
        latest = faults[0]
        fault_id = latest[0]
        display = self.fault_display_by_id.get(fault_id)
        if display:
            self.inspect_fault_var.set(display)
            self.inspect_fault_status_var.set(latest[7] or "Acik")

    def _get_latest_inspection(self, vehicle_id, week_start):
        inspections = db.list_vehicle_inspections(
            vehicle_id=vehicle_id,
            week_start=week_start,
            region=self._view_region(),
        )
        if not inspections:
            return None
        inspections.sort(key=lambda x: x[5], reverse=True)
        return inspections[0]

    def compare_vehicle_week(self):
        for item in self.vehicle_compare_tree.get_children():
            self.vehicle_compare_tree.delete(item)
        plate = self.inspect_vehicle_var.get().strip()
        if not plate:
            messagebox.showwarning("Uyari", "Arac secin.")
            return
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        inspect_date = self.inspect_date_var.get().strip()
        if not inspect_date:
            messagebox.showwarning("Uyari", "Tarih secin.")
            return
        try:
            inspect_date = normalize_date(inspect_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        week_start = week_start_from_date(inspect_date)
        prev_week = (datetime.strptime(week_start, "%Y-%m-%d").date() - timedelta(days=7)).strftime("%Y-%m-%d")

        current = self._get_latest_inspection(vehicle_id, week_start)
        previous = self._get_latest_inspection(vehicle_id, prev_week)
        if not current:
            messagebox.showwarning("Uyari", "Bu hafta icin kontrol bulunamadi.")
            return
        current_results = {row[0]: normalize_vehicle_status(row[1]) for row in db.list_vehicle_inspection_results(current[0])}
        prev_results = (
            {row[0]: normalize_vehicle_status(row[1]) for row in db.list_vehicle_inspection_results(previous[0])}
            if previous
            else {}
        )

        for item_key, label in VEHICLE_CHECKLIST:
            prev_status = prev_results.get(item_key, "-")
            curr_status = current_results.get(item_key, "-")
            if prev_status == curr_status:
                change = "Ayni"
            elif prev_status == "Olumsuz" and curr_status == "Olumsuz":
                change = "Tekrar eden sorun"
            elif curr_status == "Olumsuz" and prev_status != "Olumsuz":
                change = "Kotulesti"
            elif prev_status == "Olumsuz" and curr_status != "Olumsuz":
                change = "Iyilesti"
            else:
                change = "Degisti"
            self.vehicle_compare_tree.insert("", tk.END, values=(label, prev_status, curr_status, change))
        prev_fault_status = previous[10] if previous else None
        curr_fault_status = current[10]
        if prev_fault_status or curr_fault_status:
            prev_fault_status = prev_fault_status or "-"
            curr_fault_status = curr_fault_status or "-"
            if prev_fault_status == curr_fault_status:
                change = "Ayni"
            elif curr_fault_status == "Kapandi":
                change = "Iyilesti"
            elif prev_fault_status == "Kapandi" and curr_fault_status != "Kapandi":
                change = "Kotulesti"
            else:
                change = "Degisti"
            self.vehicle_compare_tree.insert(
                "",
                tk.END,
                values=("Ariza Durumu", prev_fault_status, curr_fault_status, change),
            )
        prev_service = "Evet" if previous and previous[11] else "Hayir"
        curr_service = "Evet" if current[11] else "Hayir"
        change = "Ayni" if prev_service == curr_service else "Degisti"
        self.vehicle_compare_tree.insert(
            "",
            tk.END,
            values=("Sanayiye Gitti", prev_service, curr_service, change),
        )

    def export_vehicle_weekly_report(self):
        plate = self.inspect_vehicle_var.get().strip()
        if not plate:
            messagebox.showwarning("Uyari", "Arac secin.")
            return
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        inspect_date = self.inspect_date_var.get().strip()
        if not inspect_date:
            messagebox.showwarning("Uyari", "Tarih secin.")
            return
        try:
            inspect_date = normalize_date(inspect_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return
        week_start = week_start_from_date(inspect_date)
        prev_week = (datetime.strptime(week_start, "%Y-%m-%d").date() - timedelta(days=7)).strftime("%Y-%m-%d")
        current = self._get_latest_inspection(vehicle_id, week_start)
        previous = self._get_latest_inspection(vehicle_id, prev_week)
        if not current:
            messagebox.showwarning("Uyari", "Bu hafta icin kontrol bulunamadi.")
            return
        current_results = {row[0]: normalize_vehicle_status(row[1]) for row in db.list_vehicle_inspection_results(current[0])}
        prev_results = (
            {row[0]: normalize_vehicle_status(row[1]) for row in db.list_vehicle_inspection_results(previous[0])}
            if previous
            else {}
        )
        current_fault_id = current[9]
        current_fault_status = current[10] or ""
        current_service = bool(current[11])
        prev_fault_id = previous[9] if previous else None
        prev_fault_status = previous[10] if previous else ""
        prev_service = bool(previous[11]) if previous else False
        current_fault_title = ""
        prev_fault_title = ""
        if current_fault_id:
            fault = db.get_vehicle_fault(current_fault_id)
            if fault:
                current_fault_title = fault[2] or ""
        if prev_fault_id:
            fault = db.get_vehicle_fault(prev_fault_id)
            if fault:
                prev_fault_title = fault[2] or ""
        week_end = week_end_from_start(week_start)
        service_visits = db.list_vehicle_service_visits(
            vehicle_id=vehicle_id,
            start_date=week_start,
            end_date=week_end,
            region=self._view_region(),
        )

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"arac_kontrol_{plate}_{week_start}.xlsx",
        )
        if not output_path:
            return
        vehicle = None
        for row in db.list_vehicles(region=self._view_region()):
            if row[1] == plate:
                vehicle = row
                break
        report.export_vehicle_weekly_report(
            output_path,
            plate,
            week_start,
            prev_week if previous else None,
            VEHICLE_CHECKLIST,
            prev_results,
            current_results,
            current[7],
            previous[7] if previous else None,
            vehicle,
            {
                "title": current_fault_title,
                "status": current_fault_status,
                "service": current_service,
            },
            {
                "title": prev_fault_title,
                "status": prev_fault_status,
                "service": prev_service,
            },
            service_visits,
        )
        self._log_action("vehicle_weekly_report", f"plate={plate} week={week_start} file={os.path.basename(output_path)}")

    def export_vehicle_card(self, plate):
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        vehicle = db.get_vehicle(vehicle_id)
        if not vehicle:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        inspections = db.list_vehicle_inspections(vehicle_id=vehicle_id, region=self._view_region())
        faults = db.list_vehicle_faults(vehicle_id=vehicle_id, region=self._view_region())
        services = db.list_vehicle_service_visits(vehicle_id=vehicle_id, region=self._view_region())
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"arac_karti_{plate}.xlsx",
        )
        if not output_path:
            return
        report.export_vehicle_card_report(
            output_path,
            plate,
            vehicle,
            inspections,
            faults,
            services,
        )
        self._log_action("vehicle_card_report", f"plate={plate} file={os.path.basename(output_path)}")
        messagebox.showinfo("Basarili", "Arac karti Excel olusturuldu.")
        messagebox.showinfo("Basarili", f"Rapor kaydedildi: {output_path}")

    def on_admin_right_click(self, event):
        row_id = self.admin_tree.identify_row(event.y)
        if row_id:
            self.admin_tree.selection_set(row_id)
            self.admin_menu.tk_popup(event.x_root, event.y_root)

    def show_admin_employee_detail(self):
        selected = self.admin_tree.selection()
        if not selected:
            return
        values = self.admin_tree.item(selected[0], "values")
        employee_name = values[0]
        employee_id = None
        base, region = split_display_name(employee_name, REGIONS)
        if region is None:
            employee_id = self.employee_map.get((base, "")) or self.employee_map.get(
                (base, self._entry_region())
            )
        else:
            employee_id = self.employee_map.get((base, region))
        if not employee_id:
            messagebox.showwarning("Uyari", "Calisan bulunamadi.")
            return

        start_date = self.admin_start_var.get().strip() or None
        end_date = self.admin_end_var.get().strip() or None
        try:
            if start_date:
                start_date = normalize_date(start_date)
            if end_date:
                end_date = normalize_date(end_date)
        except ValueError as exc:
            messagebox.showwarning("Uyari", str(exc))
            return

        records = db.list_timesheets(
            employee_id=employee_id,
            start_date=start_date,
            end_date=end_date,
            region=self._view_region(),
        )
        detail = tk.Toplevel(self)
        detail.title(f"Mesai Detay - {employee_name}")
        detail.geometry("980x600")

        frame = ttk.Frame(detail)
        frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        cols = ("date", "start", "end", "worked", "overtime", "night", "overnight", "special", "note")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        tree.heading("date", text="Tarih")
        tree.heading("start", text="Giris")
        tree.heading("end", text="Cikis")
        tree.heading("worked", text="Calisilan")
        tree.heading("overtime", text="Fazla Mesai")
        tree.heading("night", text="Gece")
        tree.heading("overnight", text="Geceye Tasan")
        tree.heading("special", text="Ozel Gun")
        tree.heading("note", text="Not")
        tree.column("date", width=100)
        tree.column("start", width=70)
        tree.column("end", width=70)
        tree.column("worked", width=90)
        tree.column("overtime", width=90)
        tree.column("night", width=90)
        tree.column("overnight", width=110)
        tree.column("special", width=90)
        tree.column("note", width=240)
        tree.pack(fill=tk.BOTH, expand=True)

        yscroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)

        for _ts_id, _emp_id, _name, work_date, start_time, end_time, break_minutes, is_special, notes, _region in records:
            (
                worked,
                _scheduled,
                overtime,
                night_hours,
                overnight_hours,
                spec_norm,
                spec_ot,
                spec_night,
            ) = calc.calc_day_hours(
                work_date,
                start_time,
                end_time,
                break_minutes,
                self.settings,
                is_special,
            )
            special_total = round(spec_norm + spec_ot + spec_night, 2)
            tree.insert(
                "",
                tk.END,
                values=(
                    work_date,
                    start_time,
                    end_time,
                    worked,
                    overtime,
                    night_hours,
                    overnight_hours,
                    special_total,
                    notes or "",
                ),
            )

    # Settings tab
    def _build_settings_tab(self):
        frame = ttk.LabelFrame(self.tab_settings_body, text="Genel Ayarlar", style="Section.TLabelframe")
        frame.pack(fill=tk.X, padx=6, pady=6)

        self.company_name_var = tk.StringVar(value=self.settings.get("company_name", ""))
        self.report_title_var = tk.StringVar(value=self.settings.get("report_title", "Puantaj ve Mesai Raporu"))
        self.weekday_hours_var = tk.StringVar(value=self.settings.get("weekday_hours", "9"))
        self.sat_start_var = tk.StringVar(value=self.settings.get("saturday_start", "09:00"))
        self.sat_end_var = tk.StringVar(value=self.settings.get("saturday_end", "14:00"))
        self.logo_path_var = tk.StringVar(value=self.settings.get("logo_path", ""))
        self.sync_enabled_var = tk.BooleanVar(value=self.settings.get("sync_enabled", "0") == "1")
        self.sync_url_var = tk.StringVar(value=self.settings.get("sync_url", ""))
        self.sync_token_var = tk.StringVar(value=self.settings.get("sync_token", ""))
        self.admin_entry_region_var.set(self.settings.get("admin_entry_region", "Ankara"))
        view_region = self.settings.get("admin_view_region", "Tum Bolgeler")
        if view_region == "ALL":
            view_region = "Tum Bolgeler"
        self.admin_view_region_var.set(view_region)

        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=4)
        create_labeled_entry(row1, "Kurum Adi", self.company_name_var, 40).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row1, "Rapor Basligi", self.report_title_var, 32).pack(side=tk.LEFT, padx=6)

        row2 = ttk.Frame(frame)
        row2.pack(fill=tk.X, pady=4)
        create_labeled_entry(row2, "Hafta Ici Saat", self.weekday_hours_var, 10).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row2, "Cumartesi Baslangic", self.sat_start_var, 10).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(row2, "Cumartesi Bitis", self.sat_end_var, 10).pack(side=tk.LEFT, padx=6)
        if self.is_admin:
            ttk.Label(row2, text="Kayit Bolge").pack(side=tk.LEFT, padx=(12, 6))
            entry_region_combo = ttk.Combobox(
                row2,
                textvariable=self.admin_entry_region_var,
                values=REGIONS,
                width=12,
                state="readonly",
            )
            entry_region_combo.pack(side=tk.LEFT, padx=6)
            ttk.Label(row2, text="Goruntuleme Bolge").pack(side=tk.LEFT, padx=(12, 6))
            view_region_combo = ttk.Combobox(
                row2,
                textvariable=self.admin_view_region_var,
                values=VIEW_REGIONS,
                width=14,
                state="readonly",
            )
            view_region_combo.pack(side=tk.LEFT, padx=6)
            view_region_combo.bind("<<ComboboxSelected>>", lambda _e: self._refresh_region_views())

        row3 = ttk.Frame(frame)
        row3.pack(fill=tk.X, pady=4)
        ttk.Label(row3, text="Logo").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(row3, textvariable=self.logo_path_var, width=60).pack(side=tk.LEFT)
        ttk.Button(row3, text="Sec", command=self.select_logo).pack(side=tk.LEFT, padx=6)

        btn_row = ttk.Frame(frame)
        btn_row.pack(fill=tk.X, pady=6)
        ttk.Button(btn_row, text="Kaydet", style="Accent.TButton", command=self.save_settings).pack(
            side=tk.LEFT, padx=6
        )

        sync_frame = ttk.LabelFrame(self.tab_settings_body, text="Bulut Senkron", style="Section.TLabelframe")
        sync_frame.pack(fill=tk.X, padx=6, pady=6)
        srow1 = ttk.Frame(sync_frame)
        srow1.pack(fill=tk.X, pady=4)
        ttk.Checkbutton(srow1, text="Senkronu Ac", variable=self.sync_enabled_var).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(srow1, "Sunucu URL", self.sync_url_var, 40).pack(side=tk.LEFT, padx=6)
        srow2 = ttk.Frame(sync_frame)
        srow2.pack(fill=tk.X, pady=4)
        create_labeled_entry(srow2, "API Token", self.sync_token_var, 40).pack(side=tk.LEFT, padx=6)
        ttk.Button(srow2, text="Senkronu Dene", command=self.manual_sync).pack(side=tk.LEFT, padx=6)
        ttk.Label(
            sync_frame,
            text="Kayit sonrasi otomatik yukleme yapilir. URL ornek: https://seninapp.onrender.com",
            foreground="#5f6a72",
        ).pack(anchor="w", padx=8, pady=(2, 6))

        data_frame = ttk.LabelFrame(self.tab_settings_body, text="Veri Yonetimi", style="Section.TLabelframe")
        data_frame.pack(fill=tk.X, padx=6, pady=6)
        drow1 = ttk.Frame(data_frame)
        drow1.pack(fill=tk.X, pady=4)
        ttk.Button(drow1, text="Log Klasoru", command=self.open_log_folder).pack(side=tk.LEFT, padx=6)
        ttk.Button(drow1, text="Veri Klasoru", command=self.open_data_folder).pack(side=tk.LEFT, padx=6)
        ttk.Button(drow1, text="Yedek Al", command=self.backup_database).pack(side=tk.LEFT, padx=6)
        ttk.Button(drow1, text="Yedek Geri Yukle", command=self.restore_database).pack(side=tk.LEFT, padx=6)

        drow2 = ttk.Frame(data_frame)
        drow2.pack(fill=tk.X, pady=4)
        ttk.Button(drow2, text="Veri Disari Aktar (ZIP)", command=self.export_data_zip).pack(side=tk.LEFT, padx=6)
        ttk.Button(drow2, text="Veri Iceri Aktar (ZIP)", command=self.import_data_zip).pack(side=tk.LEFT, padx=6)
        ttk.Label(
            data_frame,
            text="Yedek ve tasima islemleri veritabani dosyasini kopyalar. Islemden sonra uygulamayi yeniden baslatin.",
            foreground="#5f6a72",
        ).pack(anchor="w", padx=8, pady=(2, 6))

        template_frame = ttk.LabelFrame(self.tab_settings_body, text="Vardiya Sablonlari", style="Section.TLabelframe")
        template_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        self.st_id_var = tk.StringVar()
        self.st_name_var = tk.StringVar()
        self.st_start_var = tk.StringVar(value="09:00")
        self.st_end_var = tk.StringVar(value="18:00")
        self.st_break_var = tk.StringVar(value="60")

        trow1 = ttk.Frame(template_frame)
        trow1.pack(fill=tk.X, pady=4)
        create_labeled_entry(trow1, "Sablon Adi", self.st_name_var, 26).pack(side=tk.LEFT, padx=6)
        start_tpl_frame = create_time_entry(trow1, "Giris", self.st_start_var, 8)
        start_tpl_frame.pack(side=tk.LEFT, padx=6)
        end_tpl_frame = create_time_entry(trow1, "Cikis", self.st_end_var, 8)
        end_tpl_frame.pack(side=tk.LEFT, padx=6)
        create_labeled_entry(trow1, "Mola dk", self.st_break_var, 8).pack(side=tk.LEFT, padx=6)
        for child in start_tpl_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind("<FocusOut>", lambda _e: normalize_time_in_var(self.st_start_var))
        for child in end_tpl_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind("<FocusOut>", lambda _e: normalize_time_in_var(self.st_end_var))

        trow2 = ttk.Frame(template_frame)
        trow2.pack(fill=tk.X, pady=4)
        ttk.Button(trow2, text="Kaydet", style="Accent.TButton", command=self.save_shift_template).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(trow2, text="Sil", command=self.delete_shift_template).pack(side=tk.LEFT)
        ttk.Button(trow2, text="Temizle", command=self.clear_shift_template_form).pack(side=tk.LEFT, padx=6)

        list_frame = ttk.Frame(template_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        columns = ("id", "name", "start", "end", "break")
        self.template_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        self.template_tree.heading("id", text="ID")
        self.template_tree.heading("name", text="Sablon")
        self.template_tree.heading("start", text="Giris")
        self.template_tree.heading("end", text="Cikis")
        self.template_tree.heading("break", text="Mola")
        self.template_tree.column("id", width=60, anchor=tk.CENTER)
        self.template_tree.column("name", width=220)
        self.template_tree.column("start", width=80)
        self.template_tree.column("end", width=80)
        self.template_tree.column("break", width=80)
        self.template_tree.tag_configure("odd", background="#252525", foreground="#E0E0E0")
        self.template_tree.tag_configure("even", background="#1F1F1F", foreground="#E0E0E0")
        tpl_xscroll = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.template_tree.xview)
        tpl_yscroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.template_tree.yview)
        self.template_tree.configure(xscrollcommand=tpl_xscroll.set, yscrollcommand=tpl_yscroll.set)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        self.template_tree.grid(row=0, column=0, sticky="nsew")
        tpl_yscroll.grid(row=0, column=1, sticky="ns")
        tpl_xscroll.grid(row=1, column=0, sticky="ew")
        self.template_tree.bind("<<TreeviewSelect>>", self.on_template_select)

    # Kullanım rehberi kaldırıldı - Modern ERP tasarımına geçildi

    def _build_logs_tab(self):
        frame = ttk.LabelFrame(self.tab_logs_body, text="Canli Log", style="Section.TLabelframe")
        frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        tools = ttk.Frame(frame)
        tools.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(tools, text="Temizle", command=self.clear_log_view).pack(side=tk.LEFT, padx=6)
        ttk.Button(tools, text="Log Dosyasi", command=self.open_log_file).pack(side=tk.LEFT, padx=6)
        ttk.Label(tools, text=LOG_PATH, foreground="#5f6a72").pack(side=tk.LEFT, padx=8)

        text_frame = ttk.Frame(frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(
            text_frame,
            height=18,
            wrap="none",
            bg="#0f172a",
            fg="#e2e8f0",
            insertbackground="#e2e8f0",
            font=("Consolas", 9),
        )
        yscroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        xscroll = ttk.Scrollbar(text_frame, orient=tk.HORIZONTAL, command=self.log_text.xview)
        self.log_text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        self.log_text.insert(tk.END, "Log ekranina hosgeldiniz.\n")
        self.log_text.configure(state=tk.DISABLED)
        self.after(300, self._drain_log_queue)

    def _build_vehicles_tab(self):
        vehicle_frame = ttk.LabelFrame(self.tab_vehicles_body, text="Arac Bilgisi", style="Section.TLabelframe")
        vehicle_frame.pack(fill=tk.X, padx=6, pady=6)

        self.vehicle_id_var = tk.StringVar()
        self.vehicle_plate_var = tk.StringVar()
        self.vehicle_brand_var = tk.StringVar()
        self.vehicle_model_var = tk.StringVar()
        self.vehicle_year_var = tk.StringVar()
        self.vehicle_km_var = tk.StringVar()
        self.vehicle_inspection_var = tk.StringVar()
        self.vehicle_insurance_var = tk.StringVar()
        self.vehicle_maintenance_var = tk.StringVar()
        self.vehicle_oil_var = tk.StringVar()
        self.vehicle_oil_km_var = tk.StringVar()
        self.vehicle_oil_interval_var = tk.StringVar(value=str(DEFAULT_OIL_INTERVAL_KM))
        self.vehicle_notes_var = tk.StringVar()

        vrow1 = ttk.Frame(vehicle_frame)
        vrow1.pack(fill=tk.X, pady=4)
        create_labeled_entry(vrow1, "Plaka", self.vehicle_plate_var, 12).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow1, "Marka", self.vehicle_brand_var, 14).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow1, "Model", self.vehicle_model_var, 14).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow1, "Yil", self.vehicle_year_var, 8).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow1, "KM", self.vehicle_km_var, 10).pack(side=tk.LEFT, padx=6)

        vrow2 = ttk.Frame(vehicle_frame)
        vrow2.pack(fill=tk.X, pady=4)
        insp_frame, self.vehicle_insp_entry = create_labeled_date(vrow2, "Muayene", self.vehicle_inspection_var, 12)
        insp_frame.pack(side=tk.LEFT, padx=6)
        ins_frame, self.vehicle_ins_entry = create_labeled_date(vrow2, "Sigorta", self.vehicle_insurance_var, 12)
        ins_frame.pack(side=tk.LEFT, padx=6)
        maint_frame, self.vehicle_maint_entry = create_labeled_date(
            vrow2, "Bakim", self.vehicle_maintenance_var, 12
        )
        maint_frame.pack(side=tk.LEFT, padx=6)
        oil_frame, self.vehicle_oil_entry = create_labeled_date(vrow2, "Yag", self.vehicle_oil_var, 12)
        oil_frame.pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow2, "Yag KM", self.vehicle_oil_km_var, 10).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow2, "Periyot KM", self.vehicle_oil_interval_var, 10).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(vrow2, "Not", self.vehicle_notes_var, 30).pack(side=tk.LEFT, padx=6)

        vbtn = ttk.Frame(vehicle_frame)
        vbtn.pack(fill=tk.X, pady=6)
        ttk.Button(vbtn, text="Kaydet", style="Accent.TButton", command=self.add_or_update_vehicle).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(vbtn, text="Sil", command=self.delete_vehicle).pack(side=tk.LEFT)
        ttk.Button(vbtn, text="Temizle", command=self.clear_vehicle_form).pack(side=tk.LEFT, padx=6)

        vlist_frame = ttk.Frame(self.tab_vehicles_body)
        vlist_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.vehicle_tree = ttk.Treeview(
            vlist_frame,
            columns=(
                "id",
                "plate",
                "brand",
                "model",
                "year",
                "km",
                "inspection",
                "insurance",
                "maintenance",
                "oil_due",
                "region",
            ),
            show="headings",
        )
        self.vehicle_tree.heading("id", text="ID")
        self.vehicle_tree.heading("plate", text="Plaka")
        self.vehicle_tree.heading("brand", text="Marka")
        self.vehicle_tree.heading("model", text="Model")
        self.vehicle_tree.heading("year", text="Yil")
        self.vehicle_tree.heading("km", text="KM")
        self.vehicle_tree.heading("inspection", text="Muayene")
        self.vehicle_tree.heading("insurance", text="Sigorta")
        self.vehicle_tree.heading("maintenance", text="Bakim")
        self.vehicle_tree.heading("oil_due", text="Yag Durum")
        self.vehicle_tree.heading("region", text="Bolge")
        self.vehicle_tree.column("id", width=60, anchor=tk.CENTER)
        self.vehicle_tree.column("plate", width=100)
        self.vehicle_tree.column("brand", width=120)
        self.vehicle_tree.column("model", width=120)
        self.vehicle_tree.column("year", width=70)
        self.vehicle_tree.column("km", width=80)
        self.vehicle_tree.column("inspection", width=100)
        self.vehicle_tree.column("insurance", width=100)
        self.vehicle_tree.column("maintenance", width=100)
        self.vehicle_tree.column("oil_due", width=120)
        self.vehicle_tree.column("region", width=100)
        v_xscroll = ttk.Scrollbar(vlist_frame, orient=tk.HORIZONTAL, command=self.vehicle_tree.xview)
        v_yscroll = ttk.Scrollbar(vlist_frame, orient=tk.VERTICAL, command=self.vehicle_tree.yview)
        self.vehicle_tree.configure(xscrollcommand=v_xscroll.set, yscrollcommand=v_yscroll.set)
        vlist_frame.columnconfigure(0, weight=1)
        vlist_frame.rowconfigure(0, weight=1)
        self.vehicle_tree.grid(row=0, column=0, sticky="nsew")
        v_yscroll.grid(row=0, column=1, sticky="ns")
        v_xscroll.grid(row=1, column=0, sticky="ew")
        self.vehicle_tree.bind("<<TreeviewSelect>>", self.on_vehicle_select)
        self.vehicle_tree.bind("<Double-1>", lambda _e: self.show_vehicle_card_from_list())

        vcard_row = ttk.Frame(self.tab_vehicles_body)
        vcard_row.pack(fill=tk.X, padx=6, pady=4)
        ttk.Button(vcard_row, text="Arac Karti", command=self.show_vehicle_card_from_list).pack(
            side=tk.LEFT, padx=6
        )

        driver_frame = ttk.LabelFrame(self.tab_vehicles_body, text="Surucu Bilgisi", style="Section.TLabelframe")
        driver_frame.pack(fill=tk.X, padx=6, pady=6)

        self.driver_id_var = tk.StringVar()
        self.driver_name_var = tk.StringVar()
        self.driver_license_var = tk.StringVar()
        self.driver_license_exp_var = tk.StringVar()
        self.driver_phone_var = tk.StringVar()
        self.driver_notes_var = tk.StringVar()

        drow1 = ttk.Frame(driver_frame)
        drow1.pack(fill=tk.X, pady=4)
        create_labeled_entry(drow1, "Ad Soyad", self.driver_name_var, 24).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(drow1, "Ehliyet Sinifi", self.driver_license_var, 12).pack(side=tk.LEFT, padx=6)
        d_exp_frame, self.driver_exp_entry = create_labeled_date(drow1, "Bitis", self.driver_license_exp_var, 12)
        d_exp_frame.pack(side=tk.LEFT, padx=6)
        create_labeled_entry(drow1, "Telefon", self.driver_phone_var, 14).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(drow1, "Not", self.driver_notes_var, 30).pack(side=tk.LEFT, padx=6)

        dbtn = ttk.Frame(driver_frame)
        dbtn.pack(fill=tk.X, pady=6)
        ttk.Button(dbtn, text="Kaydet", style="Accent.TButton", command=self.add_or_update_driver).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(dbtn, text="Sil", command=self.delete_driver).pack(side=tk.LEFT)
        ttk.Button(dbtn, text="Temizle", command=self.clear_driver_form).pack(side=tk.LEFT, padx=6)

        dlist_frame = ttk.Frame(self.tab_vehicles_body)
        dlist_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.driver_tree = ttk.Treeview(
            dlist_frame,
            columns=("id", "name", "license", "expiry", "phone", "region"),
            show="headings",
        )
        self.driver_tree.heading("id", text="ID")
        self.driver_tree.heading("name", text="Ad Soyad")
        self.driver_tree.heading("license", text="Ehliyet")
        self.driver_tree.heading("expiry", text="Bitis")
        self.driver_tree.heading("phone", text="Telefon")
        self.driver_tree.heading("region", text="Bolge")
        self.driver_tree.column("id", width=60, anchor=tk.CENTER)
        self.driver_tree.column("name", width=220)
        self.driver_tree.column("license", width=100)
        self.driver_tree.column("expiry", width=100)
        self.driver_tree.column("phone", width=120)
        self.driver_tree.column("region", width=100)
        d_xscroll = ttk.Scrollbar(dlist_frame, orient=tk.HORIZONTAL, command=self.driver_tree.xview)
        d_yscroll = ttk.Scrollbar(dlist_frame, orient=tk.VERTICAL, command=self.driver_tree.yview)
        self.driver_tree.configure(xscrollcommand=d_xscroll.set, yscrollcommand=d_yscroll.set)
        dlist_frame.columnconfigure(0, weight=1)
        dlist_frame.rowconfigure(0, weight=1)
        self.driver_tree.grid(row=0, column=0, sticky="nsew")
        d_yscroll.grid(row=0, column=1, sticky="ns")
        d_xscroll.grid(row=1, column=0, sticky="ew")
        self.driver_tree.bind("<<TreeviewSelect>>", self.on_driver_select)
        self.driver_tree.bind("<Double-1>", lambda _e: self.show_driver_detail_from_list())

        inspect_frame = ttk.LabelFrame(self.tab_vehicles_body, text="Haftalik Kontrol", style="Section.TLabelframe")
        inspect_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        self.inspect_vehicle_var = tk.StringVar()
        self.inspect_driver_var = tk.StringVar()
        self.inspect_date_var = tk.StringVar()
        self.inspect_km_var = tk.StringVar()
        self.inspect_notes_var = tk.StringVar()
        self.inspect_fault_var = tk.StringVar()
        self.inspect_fault_status_var = tk.StringVar(value="Acik")
        self.inspect_service_var = tk.BooleanVar(value=False)

        irow1 = ttk.Frame(inspect_frame)
        irow1.pack(fill=tk.X, pady=4)
        ttk.Label(irow1, text="Arac").pack(side=tk.LEFT, padx=(0, 6))
        self.inspect_vehicle_combo = ttk.Combobox(irow1, textvariable=self.inspect_vehicle_var, width=18, state="readonly")
        self.inspect_vehicle_combo.pack(side=tk.LEFT)
        self.inspect_vehicle_combo.bind("<<ComboboxSelected>>", self.on_inspect_vehicle_change)
        ttk.Label(irow1, text="Surucu").pack(side=tk.LEFT, padx=(12, 6))
        self.inspect_driver_combo = ttk.Combobox(irow1, textvariable=self.inspect_driver_var, width=18, state="readonly")
        self.inspect_driver_combo.pack(side=tk.LEFT)
        date_frame, self.inspect_date_entry = create_labeled_date(irow1, "Tarih", self.inspect_date_var, 12)
        date_frame.pack(side=tk.LEFT, padx=6)
        create_labeled_entry(irow1, "KM", self.inspect_km_var, 10).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(irow1, "Not", self.inspect_notes_var, 40).pack(side=tk.LEFT, padx=6)

        irow2 = ttk.Frame(inspect_frame)
        irow2.pack(fill=tk.X, pady=4)
        ttk.Label(irow2, text="Ariza").pack(side=tk.LEFT, padx=(0, 6))
        self.inspect_fault_combo = ttk.Combobox(
            irow2,
            textvariable=self.inspect_fault_var,
            width=40,
            state="readonly",
            values=[""],
        )
        self.inspect_fault_combo.pack(side=tk.LEFT)
        ttk.Label(irow2, text="Durum").pack(side=tk.LEFT, padx=(12, 6))
        ttk.Combobox(
            irow2,
            textvariable=self.inspect_fault_status_var,
            values=["Acik", "Kapandi", "Takip"],
            width=10,
            state="readonly",
        ).pack(side=tk.LEFT)
        ttk.Checkbutton(irow2, text="Sanayiye Gitti", variable=self.inspect_service_var).pack(
            side=tk.LEFT, padx=12
        )
        ttk.Label(
            inspect_frame,
            text="Ariza opsiyonel. Sanayiye gidis icin ariza secin; donus tarihi bilinmiyorsa bos kalabilir.",
            foreground="#5f6a72",
        ).pack(anchor="w", padx=8, pady=(2, 6))

        self.inspect_item_vars = {}
        self.inspect_note_vars = {}
        status_values = ["Olumlu", "Olumsuz", "Bilinmiyor"]
        for item_key, label in VEHICLE_CHECKLIST:
            row = ttk.Frame(inspect_frame)
            row.pack(fill=tk.X, pady=2)
            ttk.Label(row, text=label, width=22).pack(side=tk.LEFT, padx=6)
            status_var = tk.StringVar(value="Olumlu")
            note_var = tk.StringVar()
            ttk.Combobox(row, textvariable=status_var, values=status_values, width=10, state="readonly").pack(
                side=tk.LEFT
            )
            ttk.Entry(row, textvariable=note_var, width=60).pack(side=tk.LEFT, padx=6)
            self.inspect_item_vars[item_key] = status_var
            self.inspect_note_vars[item_key] = note_var

        ibtn = ttk.Frame(inspect_frame)
        ibtn.pack(fill=tk.X, pady=6)
        ttk.Button(ibtn, text="Kontrolu Kaydet", style="Accent.TButton", command=self.save_vehicle_inspection).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(ibtn, text="Haftalik Karsilastir", command=self.compare_vehicle_week).pack(side=tk.LEFT, padx=6)
        ttk.Button(ibtn, text="Excel Raporu", command=self.export_vehicle_weekly_report).pack(
            side=tk.LEFT, padx=6
        )

        compare_frame = ttk.LabelFrame(self.tab_vehicles_body, text="Haftalik Karsilastirma", style="Section.TLabelframe")
        compare_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.vehicle_compare_tree = ttk.Treeview(
            compare_frame,
            columns=("item", "prev", "current", "change"),
            show="headings",
            height=8,
        )
        self.vehicle_compare_tree.heading("item", text="Kontrol")
        self.vehicle_compare_tree.heading("prev", text="Onceki Hafta")
        self.vehicle_compare_tree.heading("current", text="Bu Hafta")
        self.vehicle_compare_tree.heading("change", text="Durum")
        self.vehicle_compare_tree.column("item", width=220)
        self.vehicle_compare_tree.column("prev", width=120)
        self.vehicle_compare_tree.column("current", width=120)
        self.vehicle_compare_tree.column("change", width=160)
        vc_xscroll = ttk.Scrollbar(compare_frame, orient=tk.HORIZONTAL, command=self.vehicle_compare_tree.xview)
        vc_yscroll = ttk.Scrollbar(compare_frame, orient=tk.VERTICAL, command=self.vehicle_compare_tree.yview)
        self.vehicle_compare_tree.configure(xscrollcommand=vc_xscroll.set, yscrollcommand=vc_yscroll.set)
        compare_frame.columnconfigure(0, weight=1)
        compare_frame.rowconfigure(0, weight=1)
        self.vehicle_compare_tree.grid(row=0, column=0, sticky="nsew")
        vc_yscroll.grid(row=0, column=1, sticky="ns")
        vc_xscroll.grid(row=1, column=0, sticky="ew")

    def _build_dashboard_tab(self):
        summary = ttk.LabelFrame(self.tab_dashboard_body, text="Genel Ozet", style="Section.TLabelframe")
        summary.pack(fill=tk.X, padx=6, pady=6)

        self.dashboard_stats = {
            "vehicles": tk.StringVar(value="0"),
            "drivers": tk.StringVar(value="0"),
            "oil_due": tk.StringVar(value="0"),
            "inspection_due": tk.StringVar(value="0"),
            "insurance_due": tk.StringVar(value="0"),
            "maintenance_due": tk.StringVar(value="0"),
            "license_due": tk.StringVar(value="0"),
        }

        row1 = ttk.Frame(summary)
        row1.pack(fill=tk.X, pady=4)
        ttk.Label(row1, text="Arac").pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, textvariable=self.dashboard_stats["vehicles"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, text="Surucu").pack(side=tk.LEFT, padx=18)
        ttk.Label(row1, textvariable=self.dashboard_stats["drivers"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, text="Yag Degisimi").pack(side=tk.LEFT, padx=18)
        ttk.Label(row1, textvariable=self.dashboard_stats["oil_due"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, text="Muayene").pack(side=tk.LEFT, padx=18)
        ttk.Label(row1, textvariable=self.dashboard_stats["inspection_due"]).pack(side=tk.LEFT, padx=6)

        row2 = ttk.Frame(summary)
        row2.pack(fill=tk.X, pady=4)
        ttk.Label(row2, text="Sigorta").pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, textvariable=self.dashboard_stats["insurance_due"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, text="Bakim").pack(side=tk.LEFT, padx=18)
        ttk.Label(row2, textvariable=self.dashboard_stats["maintenance_due"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, text="Ehliyet").pack(side=tk.LEFT, padx=18)
        ttk.Label(row2, textvariable=self.dashboard_stats["license_due"]).pack(side=tk.LEFT, padx=6)

        vehicle_status = ttk.LabelFrame(self.tab_dashboard_body, text="Arac Durumu", style="Section.TLabelframe")
        vehicle_status.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.vehicle_status_tree = ttk.Treeview(
            vehicle_status,
            columns=(
                "plate",
                "km",
                "oil",
                "inspection",
                "insurance",
                "maintenance",
                "last_check",
                "driver",
                "region",
            ),
            show="headings",
            height=10,
        )
        self.vehicle_status_tree.heading("plate", text="Plaka")
        self.vehicle_status_tree.heading("km", text="KM")
        self.vehicle_status_tree.heading("oil", text="Yag")
        self.vehicle_status_tree.heading("inspection", text="Muayene")
        self.vehicle_status_tree.heading("insurance", text="Sigorta")
        self.vehicle_status_tree.heading("maintenance", text="Bakim")
        self.vehicle_status_tree.heading("last_check", text="Son Kontrol")
        self.vehicle_status_tree.heading("driver", text="Surucu")
        self.vehicle_status_tree.heading("region", text="Bolge")
        self.vehicle_status_tree.column("plate", width=120)
        self.vehicle_status_tree.column("km", width=80)
        self.vehicle_status_tree.column("oil", width=110)
        self.vehicle_status_tree.column("inspection", width=110)
        self.vehicle_status_tree.column("insurance", width=110)
        self.vehicle_status_tree.column("maintenance", width=110)
        self.vehicle_status_tree.column("last_check", width=110)
        self.vehicle_status_tree.column("driver", width=160)
        self.vehicle_status_tree.column("region", width=100)
        self.vehicle_status_tree.tag_configure("oil_due", background="#fde68a")
        self.vehicle_status_tree.tag_configure("oil_soon", background="#fff7ed")
        vs_xscroll = ttk.Scrollbar(vehicle_status, orient=tk.HORIZONTAL, command=self.vehicle_status_tree.xview)
        vs_yscroll = ttk.Scrollbar(vehicle_status, orient=tk.VERTICAL, command=self.vehicle_status_tree.yview)
        self.vehicle_status_tree.configure(xscrollcommand=vs_xscroll.set, yscrollcommand=vs_yscroll.set)
        vehicle_status.columnconfigure(0, weight=1)
        vehicle_status.rowconfigure(0, weight=1)
        self.vehicle_status_tree.grid(row=0, column=0, sticky="nsew")
        vs_yscroll.grid(row=0, column=1, sticky="ns")
        vs_xscroll.grid(row=1, column=0, sticky="ew")
        self.vehicle_status_menu = tk.Menu(self, tearoff=0)
        self.vehicle_status_menu.add_command(label="Detay", command=self.show_vehicle_detail)
        self.vehicle_status_tree.bind("<Button-3>", self.on_vehicle_status_right_click)
        self.vehicle_status_tree.bind("<Double-1>", lambda _e: self.show_vehicle_detail())

        vehicle_alerts = ttk.LabelFrame(self.tab_dashboard_body, text="Arac Uyarilari", style="Section.TLabelframe")
        vehicle_alerts.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.vehicle_alert_tree = ttk.Treeview(
            vehicle_alerts,
            columns=("plate", "issue", "detail"),
            show="headings",
            height=6,
        )
        self.vehicle_alert_tree.heading("plate", text="Plaka")
        self.vehicle_alert_tree.heading("issue", text="Uyari")
        self.vehicle_alert_tree.heading("detail", text="Detay")
        self.vehicle_alert_tree.column("plate", width=120)
        self.vehicle_alert_tree.column("issue", width=200)
        self.vehicle_alert_tree.column("detail", width=240)
        self.vehicle_alert_tree.bind("<Double-1>", lambda _e: self._open_vehicle_card_from_alert())
        va_xscroll = ttk.Scrollbar(vehicle_alerts, orient=tk.HORIZONTAL, command=self.vehicle_alert_tree.xview)
        va_yscroll = ttk.Scrollbar(vehicle_alerts, orient=tk.VERTICAL, command=self.vehicle_alert_tree.yview)
        self.vehicle_alert_tree.configure(xscrollcommand=va_xscroll.set, yscrollcommand=va_yscroll.set)
        vehicle_alerts.columnconfigure(0, weight=1)
        vehicle_alerts.rowconfigure(0, weight=1)
        self.vehicle_alert_tree.grid(row=0, column=0, sticky="nsew")
        va_yscroll.grid(row=0, column=1, sticky="ns")
        va_xscroll.grid(row=1, column=0, sticky="ew")

        driver_alerts = ttk.LabelFrame(self.tab_dashboard_body, text="Surucu Uyarilari", style="Section.TLabelframe")
        driver_alerts.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.driver_alert_tree = ttk.Treeview(
            driver_alerts,
            columns=("driver", "issue", "detail"),
            show="headings",
            height=6,
        )
        self.driver_alert_tree.heading("driver", text="Surucu")
        self.driver_alert_tree.heading("issue", text="Uyari")
        self.driver_alert_tree.heading("detail", text="Detay")
        self.driver_alert_tree.column("driver", width=200)
        self.driver_alert_tree.column("issue", width=200)
        self.driver_alert_tree.column("detail", width=240)
        da_xscroll = ttk.Scrollbar(driver_alerts, orient=tk.HORIZONTAL, command=self.driver_alert_tree.xview)
        da_yscroll = ttk.Scrollbar(driver_alerts, orient=tk.VERTICAL, command=self.driver_alert_tree.yview)
        self.driver_alert_tree.configure(xscrollcommand=da_xscroll.set, yscrollcommand=da_yscroll.set)
        driver_alerts.columnconfigure(0, weight=1)
        driver_alerts.rowconfigure(0, weight=1)
        self.driver_alert_tree.grid(row=0, column=0, sticky="nsew")
        da_yscroll.grid(row=0, column=1, sticky="ns")
        da_xscroll.grid(row=1, column=0, sticky="ew")

    def _build_service_tab(self):
        fault_frame = ttk.LabelFrame(self.tab_service_body, text="Ariza Kaydi", style="Section.TLabelframe")
        fault_frame.pack(fill=tk.X, padx=6, pady=6)

        self.fault_id_var = tk.StringVar()
        self.fault_vehicle_var = tk.StringVar()
        self.fault_title_var = tk.StringVar()
        self.fault_desc_var = tk.StringVar()
        self.fault_open_var = tk.StringVar()
        self.fault_close_var = tk.StringVar()
        self.fault_status_var = tk.StringVar(value="Acik")

        frow1 = ttk.Frame(fault_frame)
        frow1.pack(fill=tk.X, pady=4)
        ttk.Label(frow1, text="Arac").pack(side=tk.LEFT, padx=(0, 6))
        self.fault_vehicle_combo = ttk.Combobox(
            frow1, textvariable=self.fault_vehicle_var, width=18, state="readonly"
        )
        self.fault_vehicle_combo.pack(side=tk.LEFT)
        create_labeled_entry(frow1, "Baslik", self.fault_title_var, 24).pack(side=tk.LEFT, padx=6)

        frow2 = ttk.Frame(fault_frame)
        frow2.pack(fill=tk.X, pady=4)
        create_labeled_entry(frow2, "Aciklama", self.fault_desc_var, 60).pack(side=tk.LEFT, padx=6)

        frow3 = ttk.Frame(fault_frame)
        frow3.pack(fill=tk.X, pady=4)
        open_frame, self.fault_open_entry = create_labeled_date(frow3, "Acilis", self.fault_open_var, 12)
        open_frame.pack(side=tk.LEFT, padx=6)
        close_frame, self.fault_close_entry = create_labeled_date(frow3, "Kapanis", self.fault_close_var, 12)
        close_frame.pack(side=tk.LEFT, padx=6)
        ttk.Label(frow3, text="Durum").pack(side=tk.LEFT, padx=(12, 6))
        ttk.Combobox(
            frow3,
            textvariable=self.fault_status_var,
            values=["Acik", "Kapandi", "Takip"],
            width=10,
            state="readonly",
        ).pack(side=tk.LEFT)

        fbtn = ttk.Frame(fault_frame)
        fbtn.pack(fill=tk.X, pady=6)
        ttk.Button(fbtn, text="Kaydet", style="Accent.TButton", command=self.add_or_update_fault).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(fbtn, text="Sil", command=self.delete_fault).pack(side=tk.LEFT)
        ttk.Button(fbtn, text="Temizle", command=self.clear_fault_form).pack(side=tk.LEFT, padx=6)

        fault_list = ttk.Frame(self.tab_service_body)
        fault_list.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.fault_tree = ttk.Treeview(
            fault_list,
            columns=("id", "plate", "title", "status", "opened", "closed", "region"),
            show="headings",
            height=8,
        )
        self.fault_tree.heading("id", text="ID")
        self.fault_tree.heading("plate", text="Plaka")
        self.fault_tree.heading("title", text="Baslik")
        self.fault_tree.heading("status", text="Durum")
        self.fault_tree.heading("opened", text="Acilis")
        self.fault_tree.heading("closed", text="Kapanis")
        self.fault_tree.heading("region", text="Bolge")
        self.fault_tree.column("id", width=60, anchor=tk.CENTER)
        self.fault_tree.column("plate", width=100)
        self.fault_tree.column("title", width=220)
        self.fault_tree.column("status", width=100)
        self.fault_tree.column("opened", width=120)
        self.fault_tree.column("closed", width=120)
        self.fault_tree.column("region", width=100)
        f_xscroll = ttk.Scrollbar(fault_list, orient=tk.HORIZONTAL, command=self.fault_tree.xview)
        f_yscroll = ttk.Scrollbar(fault_list, orient=tk.VERTICAL, command=self.fault_tree.yview)
        self.fault_tree.configure(xscrollcommand=f_xscroll.set, yscrollcommand=f_yscroll.set)
        fault_list.columnconfigure(0, weight=1)
        fault_list.rowconfigure(0, weight=1)
        self.fault_tree.grid(row=0, column=0, sticky="nsew")
        f_yscroll.grid(row=0, column=1, sticky="ns")
        f_xscroll.grid(row=1, column=0, sticky="ew")
        self.fault_tree.bind("<<TreeviewSelect>>", self.on_fault_select)

        service_frame = ttk.LabelFrame(self.tab_service_body, text="Sanayi Kaydi", style="Section.TLabelframe")
        service_frame.pack(fill=tk.X, padx=6, pady=6)

        self.service_id_var = tk.StringVar()
        self.service_vehicle_var = tk.StringVar()
        self.service_fault_var = tk.StringVar()
        self.service_start_var = tk.StringVar()
        self.service_end_var = tk.StringVar()
        self.service_reason_var = tk.StringVar()
        self.service_cost_var = tk.StringVar()
        self.service_notes_var = tk.StringVar()
        self.service_in_shop_var = tk.BooleanVar(value=False)

        srow1 = ttk.Frame(service_frame)
        srow1.pack(fill=tk.X, pady=4)
        ttk.Label(srow1, text="Arac").pack(side=tk.LEFT, padx=(0, 6))
        self.service_vehicle_combo = ttk.Combobox(
            srow1, textvariable=self.service_vehicle_var, width=18, state="readonly"
        )
        self.service_vehicle_combo.pack(side=tk.LEFT)
        ttk.Label(srow1, text="Ariza").pack(side=tk.LEFT, padx=(12, 6))
        self.service_fault_combo = ttk.Combobox(
            srow1, textvariable=self.service_fault_var, width=40, state="readonly", values=[""]
        )
        self.service_fault_combo.pack(side=tk.LEFT)

        srow2 = ttk.Frame(service_frame)
        srow2.pack(fill=tk.X, pady=4)
        start_frame, self.service_start_entry = create_labeled_date(srow2, "Gidis", self.service_start_var, 12)
        start_frame.pack(side=tk.LEFT, padx=6)
        end_frame, self.service_end_entry = create_labeled_date(srow2, "Donus", self.service_end_var, 12)
        end_frame.pack(side=tk.LEFT, padx=6)
        ttk.Checkbutton(
            srow2,
            text="Sanayide",
            variable=self.service_in_shop_var,
            command=self._toggle_service_end_date,
        ).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(srow2, "Masraf", self.service_cost_var, 10).pack(side=tk.LEFT, padx=6)

        srow3 = ttk.Frame(service_frame)
        srow3.pack(fill=tk.X, pady=4)
        create_labeled_entry(srow3, "Neden", self.service_reason_var, 40).pack(side=tk.LEFT, padx=6)
        create_labeled_entry(srow3, "Not", self.service_notes_var, 40).pack(side=tk.LEFT, padx=6)

        sbtn = ttk.Frame(service_frame)
        sbtn.pack(fill=tk.X, pady=6)
        ttk.Button(sbtn, text="Kaydet", style="Accent.TButton", command=self.add_or_update_service_visit).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(sbtn, text="Sil", command=self.delete_service_visit).pack(side=tk.LEFT)
        ttk.Button(sbtn, text="Temizle", command=self.clear_service_visit_form).pack(side=tk.LEFT, padx=6)

        service_list = ttk.Frame(self.tab_service_body)
        service_list.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.service_tree = ttk.Treeview(
            service_list,
            columns=("id", "plate", "fault", "start", "end", "cost", "reason", "region"),
            show="headings",
            height=8,
        )
        self.service_tree.heading("id", text="ID")
        self.service_tree.heading("plate", text="Plaka")
        self.service_tree.heading("fault", text="Ariza")
        self.service_tree.heading("start", text="Gidis")
        self.service_tree.heading("end", text="Donus")
        self.service_tree.heading("cost", text="Masraf")
        self.service_tree.heading("reason", text="Neden")
        self.service_tree.heading("region", text="Bolge")
        self.service_tree.column("id", width=60, anchor=tk.CENTER)
        self.service_tree.column("plate", width=100)
        self.service_tree.column("fault", width=220)
        self.service_tree.column("start", width=120)
        self.service_tree.column("end", width=120)
        self.service_tree.column("cost", width=90)
        self.service_tree.column("reason", width=180)
        self.service_tree.column("region", width=100)
        s_xscroll = ttk.Scrollbar(service_list, orient=tk.HORIZONTAL, command=self.service_tree.xview)
        s_yscroll = ttk.Scrollbar(service_list, orient=tk.VERTICAL, command=self.service_tree.yview)
        self.service_tree.configure(xscrollcommand=s_xscroll.set, yscrollcommand=s_yscroll.set)
        service_list.columnconfigure(0, weight=1)
        service_list.rowconfigure(0, weight=1)
        self.service_tree.grid(row=0, column=0, sticky="nsew")
        s_yscroll.grid(row=0, column=1, sticky="ns")
        s_xscroll.grid(row=1, column=0, sticky="ew")
        self.service_tree.bind("<<TreeviewSelect>>", self.on_service_visit_select)

    def on_vehicle_status_right_click(self, event):
        row_id = self.vehicle_status_tree.identify_row(event.y)
        if not row_id:
            return
        self.vehicle_status_tree.selection_set(row_id)
        try:
            self.vehicle_status_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.vehicle_status_menu.grab_release()

    def _open_vehicle_card_from_alert(self):
        """Open vehicle card from alert tree row double-click"""
        selected = self.vehicle_alert_tree.selection()
        if not selected:
            return
        values = self.vehicle_alert_tree.item(selected[0], "values")
        if not values or len(values) < 1:
            return
        plate = values[0]  # First column is plate
        self._open_vehicle_card(plate)

    def show_vehicle_detail(self):
        selected = self.vehicle_status_tree.selection()
        if not selected:
            return
        values = self.vehicle_status_tree.item(selected[0], "values")
        if not values:
            return
        plate = values[0]
        self._open_vehicle_card(plate)

    def show_vehicle_card_from_list(self):
        selected = self.vehicle_tree.selection()
        if not selected:
            messagebox.showwarning("Uyari", "Arac secin.")
            return
        values = self.vehicle_tree.item(selected[0], "values")
        if not values:
            return
        plate = values[1]
        self._open_vehicle_card(plate)

    def show_driver_detail_from_list(self):
        selected = self.driver_tree.selection()
        if not selected:
            messagebox.showwarning("Uyari", "Surucu secin.")
            return
        values = self.driver_tree.item(selected[0], "values")
        if not values:
            return
        driver_id = parse_int(values[0])
        self._open_driver_card(driver_id)

    def _open_vehicle_card(self, plate):
        vehicle_id = self.vehicle_map.get(plate)
        if not vehicle_id:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return
        vehicle = db.get_vehicle(vehicle_id)
        if not vehicle:
            messagebox.showwarning("Uyari", "Arac bulunamadi.")
            return

        (
            _vid,
            _plate,
            brand,
            model,
            year,
            km,
            inspection_date,
            insurance_date,
            maintenance_date,
            oil_change_date,
            oil_change_km,
            oil_interval_km,
            notes,
            _region,
        ) = vehicle

        detail_win = tk.Toplevel(self)
        detail_win.title(f"Arac Karti - {plate}")
        detail_win.geometry("980x700")

        content = self._make_window_scrollable(detail_win)

        info = ttk.LabelFrame(content, text="Arac Bilgisi", style="Section.TLabelframe")
        info.pack(fill=tk.X, padx=10, pady=8)
        info_row1 = ttk.Frame(info)
        info_row1.pack(fill=tk.X, pady=4)
        ttk.Label(info_row1, text=f"Plaka: {plate}").pack(side=tk.LEFT, padx=6)
        ttk.Label(info_row1, text=f"Marka/Model: {brand} {model}").pack(side=tk.LEFT, padx=12)
        ttk.Label(info_row1, text=f"Yil: {year}").pack(side=tk.LEFT, padx=12)
        info_row2 = ttk.Frame(info)
        info_row2.pack(fill=tk.X, pady=4)
        ttk.Label(info_row2, text=f"KM: {km or '-'}").pack(side=tk.LEFT, padx=6)
        ttk.Label(info_row2, text=f"Yag Degisim: {oil_change_date or '-'}").pack(side=tk.LEFT, padx=12)
        ttk.Label(info_row2, text=f"Yag KM: {oil_change_km or '-'}").pack(side=tk.LEFT, padx=12)
        interval_km = oil_interval_km or DEFAULT_OIL_INTERVAL_KM
        oil_status = "-"
        if interval_km and oil_change_km is not None and km is not None:
            remaining = interval_km - (km - oil_change_km)
            oil_status = "Geldi" if remaining <= 0 else f"{remaining} km"
        ttk.Label(info_row2, text=f"Periyot: {interval_km or '-'} km").pack(side=tk.LEFT, padx=12)
        ttk.Label(info_row2, text=f"Yag Durum: {oil_status}").pack(side=tk.LEFT, padx=12)
        info_row3 = ttk.Frame(info)
        info_row3.pack(fill=tk.X, pady=4)
        ttk.Label(info_row3, text=f"Muayene: {inspection_date or '-'}").pack(side=tk.LEFT, padx=6)
        ttk.Label(info_row3, text=f"Sigorta: {insurance_date or '-'}").pack(side=tk.LEFT, padx=12)
        ttk.Label(info_row3, text=f"Bakim: {maintenance_date or '-'}").pack(side=tk.LEFT, padx=12)
        if notes:
            info_row4 = ttk.Frame(info)
            info_row4.pack(fill=tk.X, pady=4)
            ttk.Label(info_row4, text=f"Not: {notes}").pack(side=tk.LEFT, padx=6)

        btn_row = ttk.Frame(content)
        btn_row.pack(fill=tk.X, padx=10, pady=4)
        ttk.Button(btn_row, text="Excel Arac Karti", command=lambda: self.export_vehicle_card(plate)).pack(
            side=tk.LEFT, padx=6
        )

        inspections = db.list_vehicle_inspections(vehicle_id=vehicle_id, region=self._view_region())
        if inspections:
            last_driver = inspections[0][4] or "-"
            last_date = inspections[0][5] or "-"
            info_row5 = ttk.Frame(info)
            info_row5.pack(fill=tk.X, pady=4)
            ttk.Label(info_row5, text=f"Son Surucu: {last_driver}").pack(side=tk.LEFT, padx=6)
            ttk.Label(info_row5, text=f"Son Kontrol: {last_date}").pack(side=tk.LEFT, padx=12)
        faults = db.list_vehicle_faults(vehicle_id=vehicle_id, region=self._view_region())
        services = db.list_vehicle_service_visits(vehicle_id=vehicle_id, region=self._view_region())

        inspect_frame = ttk.LabelFrame(content, text="Kontroller", style="Section.TLabelframe")
        inspect_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        inspect_tree = ttk.Treeview(
            inspect_frame,
            columns=("date", "week", "driver", "km", "fault", "status", "service", "note"),
            show="headings",
            height=6,
        )
        inspect_tree.heading("date", text="Tarih")
        inspect_tree.heading("week", text="Hafta")
        inspect_tree.heading("driver", text="Surucu")
        inspect_tree.heading("km", text="KM")
        inspect_tree.heading("fault", text="Ariza")
        inspect_tree.heading("status", text="Durum")
        inspect_tree.heading("service", text="Sanayi")
        inspect_tree.heading("note", text="Not")
        inspect_tree.column("date", width=110)
        inspect_tree.column("week", width=110)
        inspect_tree.column("driver", width=160)
        inspect_tree.column("km", width=70)
        inspect_tree.column("fault", width=160)
        inspect_tree.column("status", width=90)
        inspect_tree.column("service", width=80)
        inspect_tree.column("note", width=180)
        i_xscroll = ttk.Scrollbar(inspect_frame, orient=tk.HORIZONTAL, command=inspect_tree.xview)
        i_yscroll = ttk.Scrollbar(inspect_frame, orient=tk.VERTICAL, command=inspect_tree.yview)
        inspect_tree.configure(xscrollcommand=i_xscroll.set, yscrollcommand=i_yscroll.set)
        inspect_frame.columnconfigure(0, weight=1)
        inspect_frame.rowconfigure(0, weight=1)
        inspect_tree.grid(row=0, column=0, sticky="nsew")
        i_yscroll.grid(row=0, column=1, sticky="ns")
        i_xscroll.grid(row=1, column=0, sticky="ew")

        for row in inspections[:20]:
            (
                _iid,
                _veh_id,
                _plate,
                _driver_id,
                driver_name,
                inspect_date,
                week_start,
                km_val,
                note_val,
                fault_id,
                fault_status,
                service_visit,
            ) = row
            fault_title = ""
            if fault_id:
                fault = db.get_vehicle_fault(fault_id)
                if fault:
                    fault_title = fault[2] or ""
            inspect_tree.insert(
                "",
                tk.END,
                values=(
                    inspect_date,
                    week_start,
                    driver_name or "-",
                    km_val or "-",
                    fault_title,
                    fault_status or "",
                    "Evet" if service_visit else "",
                    note_val or "",
                ),
            )

        fault_frame = ttk.LabelFrame(content, text="Ariza Kayitlari", style="Section.TLabelframe")
        fault_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        fault_tree = ttk.Treeview(
            fault_frame,
            columns=("title", "status", "opened", "closed"),
            show="headings",
            height=5,
        )
        fault_tree.heading("title", text="Baslik")
        fault_tree.heading("status", text="Durum")
        fault_tree.heading("opened", text="Acilis")
        fault_tree.heading("closed", text="Kapanis")
        fault_tree.column("title", width=240)
        fault_tree.column("status", width=100)
        fault_tree.column("opened", width=120)
        fault_tree.column("closed", width=120)
        f_xscroll = ttk.Scrollbar(fault_frame, orient=tk.HORIZONTAL, command=fault_tree.xview)
        f_yscroll = ttk.Scrollbar(fault_frame, orient=tk.VERTICAL, command=fault_tree.yview)
        fault_tree.configure(xscrollcommand=f_xscroll.set, yscrollcommand=f_yscroll.set)
        fault_frame.columnconfigure(0, weight=1)
        fault_frame.rowconfigure(0, weight=1)
        fault_tree.grid(row=0, column=0, sticky="nsew")
        f_yscroll.grid(row=0, column=1, sticky="ns")
        f_xscroll.grid(row=1, column=0, sticky="ew")
        for fault in faults:
            _fid, _vid, _plate, title, _desc, opened_date, closed_date, status, _region = fault
            fault_tree.insert("", tk.END, values=(title, status, opened_date or "", closed_date or ""))

        service_frame = ttk.LabelFrame(content, text="Sanayi Kayitlari", style="Section.TLabelframe")
        service_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        service_tree = ttk.Treeview(
            service_frame,
            columns=("fault", "start", "end", "cost", "reason"),
            show="headings",
            height=5,
        )
        service_tree.heading("fault", text="Ariza")
        service_tree.heading("start", text="Gidis")
        service_tree.heading("end", text="Donus")
        service_tree.heading("cost", text="Masraf")
        service_tree.heading("reason", text="Neden")
        service_tree.column("fault", width=220)
        service_tree.column("start", width=120)
        service_tree.column("end", width=120)
        service_tree.column("cost", width=90)
        service_tree.column("reason", width=200)
        s_xscroll = ttk.Scrollbar(service_frame, orient=tk.HORIZONTAL, command=service_tree.xview)
        s_yscroll = ttk.Scrollbar(service_frame, orient=tk.VERTICAL, command=service_tree.yview)
        service_tree.configure(xscrollcommand=s_xscroll.set, yscrollcommand=s_yscroll.set)
        service_frame.columnconfigure(0, weight=1)
        service_frame.rowconfigure(0, weight=1)
        service_tree.grid(row=0, column=0, sticky="nsew")
        s_yscroll.grid(row=0, column=1, sticky="ns")
        s_xscroll.grid(row=1, column=0, sticky="ew")
        for visit in services:
            _sid, _vid, _plate, _fid, title, start_date, end_date, reason, cost, _notes, _region = visit
            end_value = end_date or "Sanayide"
            cost_value = f"{cost:.2f}" if cost is not None else ""
            service_tree.insert("", tk.END, values=(title or "", start_date, end_value, cost_value, reason or ""))

        compare_frame = ttk.LabelFrame(content, text="Haftalik Karsilastirma", style="Section.TLabelframe")
        compare_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        compare_tree = ttk.Treeview(
            compare_frame,
            columns=("item", "prev", "current", "change"),
            show="headings",
            height=6,
        )
        compare_tree.heading("item", text="Kontrol")
        compare_tree.heading("prev", text="Onceki Hafta")
        compare_tree.heading("current", text="Bu Hafta")
        compare_tree.heading("change", text="Durum")
        compare_tree.column("item", width=220)
        compare_tree.column("prev", width=120)
        compare_tree.column("current", width=120)
        compare_tree.column("change", width=160)
        c_xscroll = ttk.Scrollbar(compare_frame, orient=tk.HORIZONTAL, command=compare_tree.xview)
        c_yscroll = ttk.Scrollbar(compare_frame, orient=tk.VERTICAL, command=compare_tree.yview)
        compare_tree.configure(xscrollcommand=c_xscroll.set, yscrollcommand=c_yscroll.set)
        compare_frame.columnconfigure(0, weight=1)
        compare_frame.rowconfigure(0, weight=1)
        compare_tree.grid(row=0, column=0, sticky="nsew")
        c_yscroll.grid(row=0, column=1, sticky="ns")
        c_xscroll.grid(row=1, column=0, sticky="ew")

        if len(inspections) >= 2:
            current_inspection = inspections[0]
            previous_inspection = inspections[1]
            current_results = {
                row[0]: normalize_vehicle_status(row[1])
                for row in db.list_vehicle_inspection_results(current_inspection[0])
            }
            prev_results = {
                row[0]: normalize_vehicle_status(row[1])
                for row in db.list_vehicle_inspection_results(previous_inspection[0])
            }
            for item_key, label in VEHICLE_CHECKLIST:
                prev_status = prev_results.get(item_key, "-")
                curr_status = current_results.get(item_key, "-")
                if prev_status == curr_status:
                    change = "Ayni"
                elif prev_status == "Olumsuz" and curr_status == "Olumsuz":
                    change = "Tekrar eden sorun"
                elif curr_status == "Olumsuz" and prev_status != "Olumsuz":
                    change = "Kotulesti"
                elif prev_status == "Olumsuz" and curr_status != "Olumsuz":
                    change = "Iyilesti"
                else:
                    change = "Degisti"
                compare_tree.insert("", tk.END, values=(label, prev_status, curr_status, change))
            prev_fault_status = previous_inspection[10] or "-"
            curr_fault_status = current_inspection[10] or "-"
            if prev_fault_status or curr_fault_status:
                if prev_fault_status == curr_fault_status:
                    change = "Ayni"
                elif curr_fault_status == "Kapandi":
                    change = "Iyilesti"
                elif prev_fault_status == "Kapandi" and curr_fault_status != "Kapandi":
                    change = "Kotulesti"
                else:
                    change = "Degisti"
                compare_tree.insert("", tk.END, values=("Ariza Durumu", prev_fault_status, curr_fault_status, change))
            prev_service = "Evet" if previous_inspection[11] else "Hayir"
            curr_service = "Evet" if current_inspection[11] else "Hayir"
            change = "Ayni" if prev_service == curr_service else "Degisti"
            compare_tree.insert("", tk.END, values=("Sanayiye Gitti", prev_service, curr_service, change))

    def _open_driver_card(self, driver_id):
        driver = db.get_driver(driver_id)
        if not driver:
            messagebox.showwarning("Uyari", "Surucu bulunamadi.")
            return
        _did, name, license_class, license_expiry, phone, notes = driver
        inspections = db.list_driver_inspections(driver_id, region=self._view_region())

        detail_win = tk.Toplevel(self)
        detail_win.title(f"Surucu Karti - {name}")
        detail_win.geometry("960x680")

        content = self._make_window_scrollable(detail_win)

        info = ttk.LabelFrame(content, text="Surucu Bilgisi", style="Section.TLabelframe")
        info.pack(fill=tk.X, padx=10, pady=8)
        info_row1 = ttk.Frame(info)
        info_row1.pack(fill=tk.X, pady=4)
        ttk.Label(info_row1, text=f"Ad Soyad: {name}").pack(side=tk.LEFT, padx=6)
        ttk.Label(info_row1, text=f"Ehliyet: {license_class or '-'}").pack(side=tk.LEFT, padx=12)
        ttk.Label(info_row1, text=f"Bitis: {license_expiry or '-'}").pack(side=tk.LEFT, padx=12)
        info_row2 = ttk.Frame(info)
        info_row2.pack(fill=tk.X, pady=4)
        ttk.Label(info_row2, text=f"Telefon: {phone or '-'}").pack(side=tk.LEFT, padx=6)
        if notes:
            ttk.Label(info_row2, text=f"Not: {notes}").pack(side=tk.LEFT, padx=12)

        vehicle_frame = ttk.LabelFrame(content, text="Surucunun Araclari", style="Section.TLabelframe")
        vehicle_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        vehicle_tree = ttk.Treeview(
            vehicle_frame,
            columns=("plate", "last_date", "count"),
            show="headings",
            height=5,
        )
        vehicle_tree.heading("plate", text="Plaka")
        vehicle_tree.heading("last_date", text="Son Kontrol")
        vehicle_tree.heading("count", text="Kontrol Sayisi")
        vehicle_tree.column("plate", width=140)
        vehicle_tree.column("last_date", width=140)
        vehicle_tree.column("count", width=120, anchor=tk.CENTER)
        v_xscroll = ttk.Scrollbar(vehicle_frame, orient=tk.HORIZONTAL, command=vehicle_tree.xview)
        v_yscroll = ttk.Scrollbar(vehicle_frame, orient=tk.VERTICAL, command=vehicle_tree.yview)
        vehicle_tree.configure(xscrollcommand=v_xscroll.set, yscrollcommand=v_yscroll.set)
        vehicle_frame.columnconfigure(0, weight=1)
        vehicle_frame.rowconfigure(0, weight=1)
        vehicle_tree.grid(row=0, column=0, sticky="nsew")
        v_yscroll.grid(row=0, column=1, sticky="ns")
        v_xscroll.grid(row=1, column=0, sticky="ew")

        vehicle_summary = {}
        for row in inspections:
            plate = row[2]
            inspect_date = row[5]
            entry = vehicle_summary.setdefault(plate, {"last": inspect_date, "count": 0})
            entry["count"] += 1
            if inspect_date and (not entry["last"] or inspect_date > entry["last"]):
                entry["last"] = inspect_date
        for plate, info_row in sorted(vehicle_summary.items()):
            vehicle_tree.insert("", tk.END, values=(plate, info_row["last"] or "-", info_row["count"]))

        history_frame = ttk.LabelFrame(content, text="Kontrol Gecmisi", style="Section.TLabelframe")
        history_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        history_tree = ttk.Treeview(
            history_frame,
            columns=("date", "week", "plate", "km", "fault", "status", "service", "note"),
            show="headings",
            height=8,
        )
        history_tree.heading("date", text="Tarih")
        history_tree.heading("week", text="Hafta")
        history_tree.heading("plate", text="Plaka")
        history_tree.heading("km", text="KM")
        history_tree.heading("fault", text="Ariza")
        history_tree.heading("status", text="Durum")
        history_tree.heading("service", text="Sanayi")
        history_tree.heading("note", text="Not")
        history_tree.column("date", width=110)
        history_tree.column("week", width=110)
        history_tree.column("plate", width=120)
        history_tree.column("km", width=70)
        history_tree.column("fault", width=160)
        history_tree.column("status", width=90)
        history_tree.column("service", width=80)
        history_tree.column("note", width=200)
        h_xscroll = ttk.Scrollbar(history_frame, orient=tk.HORIZONTAL, command=history_tree.xview)
        h_yscroll = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=history_tree.yview)
        history_tree.configure(xscrollcommand=h_xscroll.set, yscrollcommand=h_yscroll.set)
        history_frame.columnconfigure(0, weight=1)
        history_frame.rowconfigure(0, weight=1)
        history_tree.grid(row=0, column=0, sticky="nsew")
        h_yscroll.grid(row=0, column=1, sticky="ns")
        h_xscroll.grid(row=1, column=0, sticky="ew")

        for row in inspections[:50]:
            (
                _iid,
                _veh_id,
                plate,
                _driver_id,
                _driver_name,
                inspect_date,
                week_start,
                km_val,
                note_val,
                fault_id,
                fault_status,
                service_visit,
            ) = row
            fault_title = ""
            if fault_id:
                fault = db.get_vehicle_fault(fault_id)
                if fault:
                    fault_title = fault[2] or ""
            history_tree.insert(
                "",
                tk.END,
                values=(
                    inspect_date,
                    week_start,
                    plate,
                    km_val or "-",
                    fault_title,
                    fault_status or "",
                    "Evet" if service_visit else "",
                    note_val or "",
                ),
            )

    def _build_admin_tab(self):
        content = self.tab_admin_body

        filter_frame = ttk.LabelFrame(content, text="Filtre", style="Section.TLabelframe")
        filter_frame.pack(fill=tk.X, padx=6, pady=6)

        self.admin_employee_var = tk.StringVar(value="Tum Calisanlar")
        self.admin_department_var = tk.StringVar(value="Tum Departmanlar")
        self.admin_title_var = tk.StringVar(value="Tum Unvanlar")
        self.admin_start_var = tk.StringVar()
        self.admin_end_var = tk.StringVar()
        self.admin_search_var = tk.StringVar()

        ttk.Label(filter_frame, text="Calisan").pack(side=tk.LEFT, padx=(0, 6))
        self.admin_employee_combo = ttk.Combobox(
            filter_frame, textvariable=self.admin_employee_var, width=24, state="readonly"
        )
        self.admin_employee_combo.pack(side=tk.LEFT)
        ttk.Label(filter_frame, text="Departman").pack(side=tk.LEFT, padx=(12, 6))
        self.admin_department_combo = ttk.Combobox(
            filter_frame, textvariable=self.admin_department_var, width=18, state="readonly"
        )
        self.admin_department_combo.pack(side=tk.LEFT)
        ttk.Label(filter_frame, text="Unvan").pack(side=tk.LEFT, padx=(12, 6))
        self.admin_title_combo = ttk.Combobox(filter_frame, textvariable=self.admin_title_var, width=18, state="readonly")
        self.admin_title_combo.pack(side=tk.LEFT)

        row2 = ttk.Frame(filter_frame)
        row2.pack(fill=tk.X, pady=6)
        start_frame, self.admin_start_entry = create_labeled_date(row2, "Baslangic", self.admin_start_var, 12)
        start_frame.pack(side=tk.LEFT, padx=6)
        end_frame, self.admin_end_entry = create_labeled_date(row2, "Bitis", self.admin_end_var, 12)
        end_frame.pack(side=tk.LEFT, padx=6)
        clear_date_entry(self.admin_start_entry)
        clear_date_entry(self.admin_end_entry)
        ttk.Label(row2, text="Ara").pack(side=tk.LEFT, padx=(12, 6))
        ttk.Entry(row2, textvariable=self.admin_search_var, width=24).pack(side=tk.LEFT)
        ttk.Button(row2, text="Guncelle", command=self.refresh_admin_summary).pack(side=tk.LEFT, padx=12)

        summary = ttk.LabelFrame(content, text="Ozet", style="Section.TLabelframe")
        summary.pack(fill=tk.X, padx=6, pady=6)

        self.admin_stats = {
            "total_records": tk.StringVar(value="0"),
            "total_employees": tk.StringVar(value="0"),
            "total_worked": tk.StringVar(value="0"),
            "total_overtime": tk.StringVar(value="0"),
            "total_night": tk.StringVar(value="0"),
            "total_overnight": tk.StringVar(value="0"),
            "total_special": tk.StringVar(value="0"),
            "avg_overtime": tk.StringVar(value="0"),
            "max_daily": tk.StringVar(value="0"),
        }

        row1 = ttk.Frame(summary)
        row1.pack(fill=tk.X, pady=4)
        ttk.Label(row1, text="Kayit").pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, textvariable=self.admin_stats["total_records"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, text="Calisan").pack(side=tk.LEFT, padx=18)
        ttk.Label(row1, textvariable=self.admin_stats["total_employees"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, text="Toplam Calisilan").pack(side=tk.LEFT, padx=18)
        ttk.Label(row1, textvariable=self.admin_stats["total_worked"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row1, text="Toplam Fazla Mesai").pack(side=tk.LEFT, padx=18)
        ttk.Label(row1, textvariable=self.admin_stats["total_overtime"]).pack(side=tk.LEFT, padx=6)

        row2 = ttk.Frame(summary)
        row2.pack(fill=tk.X, pady=4)
        ttk.Label(row2, text="Toplam Gece").pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, textvariable=self.admin_stats["total_night"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, text="Geceye Tasan").pack(side=tk.LEFT, padx=18)
        ttk.Label(row2, textvariable=self.admin_stats["total_overnight"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, text="Ozel Gun").pack(side=tk.LEFT, padx=18)
        ttk.Label(row2, textvariable=self.admin_stats["total_special"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, text="Ortalama Fazla Mesai").pack(side=tk.LEFT, padx=18)
        ttk.Label(row2, textvariable=self.admin_stats["avg_overtime"]).pack(side=tk.LEFT, padx=6)
        ttk.Label(row2, text="En Yuksek Gun").pack(side=tk.LEFT, padx=18)
        ttk.Label(row2, textvariable=self.admin_stats["max_daily"]).pack(side=tk.LEFT, padx=6)

        table_frame = ttk.LabelFrame(content, text="Calisan Ozeti", style="Section.TLabelframe")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        self.admin_tree = ttk.Treeview(
            table_frame,
            columns=("employee", "worked", "overtime", "night", "overnight", "special"),
            show="headings",
        )
        self.admin_tree.heading("employee", text="Calisan")
        self.admin_tree.heading("worked", text="Calisilan")
        self.admin_tree.heading("overtime", text="Fazla Mesai")
        self.admin_tree.heading("night", text="Gece")
        self.admin_tree.heading("overnight", text="Geceye Tasan")
        self.admin_tree.heading("special", text="Ozel Gun")
        self.admin_tree.column("employee", width=220)
        self.admin_tree.column("worked", width=90)
        self.admin_tree.column("overtime", width=90)
        self.admin_tree.column("night", width=90)
        self.admin_tree.column("overnight", width=110)
        self.admin_tree.column("special", width=90)
        admin_xscroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.admin_tree.xview)
        admin_yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.admin_tree.yview)
        self.admin_tree.configure(xscrollcommand=admin_xscroll.set, yscrollcommand=admin_yscroll.set)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        self.admin_tree.grid(row=0, column=0, sticky="nsew")
        admin_yscroll.grid(row=0, column=1, sticky="ns")
        admin_xscroll.grid(row=1, column=0, sticky="ew")
        self.admin_tree.bind("<Button-3>", self.on_admin_right_click)
        self.admin_menu = tk.Menu(self, tearoff=0)
        self.admin_menu.add_command(label="Detay", command=self.show_admin_employee_detail)

        alert_frame = ttk.LabelFrame(content, text="Uyarilar", style="Section.TLabelframe")
        alert_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.admin_alert_tree = ttk.Treeview(
            alert_frame,
            columns=("date", "employee", "issue", "value"),
            show="headings",
            height=6,
        )
        self.admin_alert_tree.heading("date", text="Tarih")
        self.admin_alert_tree.heading("employee", text="Calisan")
        self.admin_alert_tree.heading("issue", text="Uyari")
        self.admin_alert_tree.heading("value", text="Deger")
        self.admin_alert_tree.column("date", width=100)
        self.admin_alert_tree.column("employee", width=220)
        self.admin_alert_tree.column("issue", width=180)
        self.admin_alert_tree.column("value", width=100)
        alert_xscroll = ttk.Scrollbar(alert_frame, orient=tk.HORIZONTAL, command=self.admin_alert_tree.xview)
        alert_yscroll = ttk.Scrollbar(alert_frame, orient=tk.VERTICAL, command=self.admin_alert_tree.yview)
        self.admin_alert_tree.configure(xscrollcommand=alert_xscroll.set, yscrollcommand=alert_yscroll.set)
        alert_frame.columnconfigure(0, weight=1)
        alert_frame.rowconfigure(0, weight=1)
        self.admin_alert_tree.grid(row=0, column=0, sticky="nsew")
        alert_yscroll.grid(row=0, column=1, sticky="ns")
        alert_xscroll.grid(row=1, column=0, sticky="ew")

        anomaly_frame = ttk.LabelFrame(content, text="Anomali Listesi", style="Section.TLabelframe")
        anomaly_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.admin_anomaly_tree = ttk.Treeview(
            anomaly_frame,
            columns=("employee", "period", "issue"),
            show="headings",
            height=6,
        )
        self.admin_anomaly_tree.heading("employee", text="Calisan")
        self.admin_anomaly_tree.heading("period", text="Donem")
        self.admin_anomaly_tree.heading("issue", text="Durum")
        self.admin_anomaly_tree.column("employee", width=220)
        self.admin_anomaly_tree.column("period", width=160)
        self.admin_anomaly_tree.column("issue", width=220)
        anom_xscroll = ttk.Scrollbar(anomaly_frame, orient=tk.HORIZONTAL, command=self.admin_anomaly_tree.xview)
        anom_yscroll = ttk.Scrollbar(anomaly_frame, orient=tk.VERTICAL, command=self.admin_anomaly_tree.yview)
        self.admin_anomaly_tree.configure(xscrollcommand=anom_xscroll.set, yscrollcommand=anom_yscroll.set)
        anomaly_frame.columnconfigure(0, weight=1)
        anomaly_frame.rowconfigure(0, weight=1)
        self.admin_anomaly_tree.grid(row=0, column=0, sticky="nsew")
        anom_yscroll.grid(row=0, column=1, sticky="ns")
        anom_xscroll.grid(row=1, column=0, sticky="ew")

        pack_frame = ttk.LabelFrame(content, text="Rapor Paketleme", style="Section.TLabelframe")
        pack_frame.pack(fill=tk.X, padx=6, pady=6)
        ttk.Label(pack_frame, text="Ay (YYYY-MM)").pack(side=tk.LEFT, padx=6)
        self.admin_month_var = tk.StringVar()
        ttk.Entry(pack_frame, textvariable=self.admin_month_var, width=10).pack(side=tk.LEFT)
        ttk.Button(pack_frame, text="Raporlari Zip Yap", command=self.package_monthly_reports).pack(
            side=tk.LEFT, padx=6
        )

        self.refresh_admin_summary()

    def select_logo(self):
        path = filedialog.askopenfilename(
            filetypes=[("Images", "*.png;*.jpg;*.jpeg;*.bmp"), ("All", "*.*")]
        )
        if path:
            self.logo_path_var.set(path)

    def save_settings(self):
        prev_entry_region = self.settings.get("admin_entry_region", "Ankara")
        prev_view_region = self.settings.get("admin_view_region", "Tum Bolgeler")
        db.set_setting("company_name", self.company_name_var.get().strip())
        db.set_setting("report_title", self.report_title_var.get().strip())
        db.set_setting("weekday_hours", self.weekday_hours_var.get().strip())
        db.set_setting("saturday_start", self.sat_start_var.get().strip())
        db.set_setting("saturday_end", self.sat_end_var.get().strip())
        db.set_setting("logo_path", self.logo_path_var.get().strip())
        db.set_setting("sync_enabled", "1" if self.sync_enabled_var.get() else "0")
        db.set_setting("sync_url", self.sync_url_var.get().strip())
        db.set_setting("sync_token", self.sync_token_var.get().strip())
        if self.is_admin:
            new_entry_region = self.admin_entry_region_var.get().strip() or "Ankara"
            new_view_region = self.admin_view_region_var.get().strip() or "Tum Bolgeler"
            db.set_setting("admin_entry_region", new_entry_region)
            db.set_setting("admin_view_region", new_view_region)
        self.settings = db.get_all_settings()
        if self.is_admin and prev_view_region != (self.admin_view_region_var.get().strip() or "Tum Bolgeler"):
            self._refresh_region_views()
        self._log_action("settings_save")
        messagebox.showinfo("Basarili", "Ayarlar kaydedildi.")
        self.trigger_sync("settings")

    def open_log_folder(self):
        try:
            os.makedirs(LOG_DIR, exist_ok=True)
            os.startfile(LOG_DIR)
        except Exception:
            messagebox.showerror("Hata", "Log klasoru acilamadi.")

    def open_log_file(self):
        try:
            if not os.path.isfile(LOG_PATH):
                with open(LOG_PATH, "a", encoding="utf-8"):
                    pass
            os.startfile(LOG_PATH)
        except Exception:
            messagebox.showerror("Hata", "Log dosyasi acilamadi.")

    def open_data_folder(self):
        try:
            os.makedirs(db.DB_DIR, exist_ok=True)
            os.startfile(db.DB_DIR)
        except Exception:
            messagebox.showerror("Hata", "Veri klasoru acilamadi.")

    def backup_database(self):
        try:
            default_name = f"puantaj_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            path = filedialog.asksaveasfilename(
                defaultextension=".db",
                initialfile=default_name,
                filetypes=[("SQLite DB", "*.db"), ("All Files", "*.*")],
            )
            if not path:
                return
            backup_path = db.create_backup(path)
            if self.logger:
                self.logger.info("Manual backup created: %s", backup_path)
            self._log_action("backup_create", f"file={os.path.basename(backup_path)}")
            messagebox.showinfo("Basarili", f"Yedek olusturuldu: {backup_path}")
        except Exception as exc:
            if self.logger:
                self.logger.exception("Backup failed")
            messagebox.showerror("Hata", f"Yedek alinamadi: {exc}")

    def restore_database(self):
        if not messagebox.askyesno(
            "Onay",
            "Bu islem mevcut veritabaniyi degistirecek. Devam etmek istiyor musun?",
        ):
            return
        try:
            path = filedialog.askopenfilename(
                filetypes=[("SQLite DB", "*.db"), ("All Files", "*.*")]
            )
            if not path:
                return
            db.restore_backup(path)
            if self.logger:
                self.logger.info("Database restored from: %s", path)
            self._log_action("backup_restore", f"file={os.path.basename(path)}")
            messagebox.showinfo("Basarili", "Yedek geri yuklendi. Uygulamayi yeniden baslatin.")
        except Exception as exc:
            if self.logger:
                self.logger.exception("Restore failed")
            messagebox.showerror("Hata", f"Geri yukleme basarisiz: {exc}")

    def export_data_zip(self):
        try:
            default_name = f"rainstaff_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            path = filedialog.asksaveasfilename(
                defaultextension=".zip",
                initialfile=default_name,
                filetypes=[("ZIP", "*.zip"), ("All Files", "*.*")],
            )
            if not path:
                return
            export_path = db.export_data_zip(path)
            if self.logger:
                self.logger.info("Data exported: %s", export_path)
            self._log_action("data_export", f"file={os.path.basename(export_path)}")
            messagebox.showinfo("Basarili", f"Disari aktarildi: {export_path}")
        except Exception as exc:
            if self.logger:
                self.logger.exception("Export failed")
            messagebox.showerror("Hata", f"Disari aktarma basarisiz: {exc}")

    def import_data_zip(self):
        if not messagebox.askyesno(
            "Onay",
            "Bu islem mevcut veritabaniyi degistirecek. Devam etmek istiyor musun?",
        ):
            return
        try:
            path = filedialog.askopenfilename(
                filetypes=[("ZIP", "*.zip"), ("All Files", "*.*")]
            )
            if not path:
                return
            db.import_data_zip(path)
            if self.logger:
                self.logger.info("Data imported from: %s", path)
            self._log_action("data_import", f"file={os.path.basename(path)}")
            messagebox.showinfo("Basarili", "Iceri aktarma tamamlandi. Uygulamayi yeniden baslatin.")
        except Exception as exc:
            if self.logger:
                self.logger.exception("Import failed")
            messagebox.showerror("Hata", f"Iceri aktarma basarisiz: {exc}")

    def _build_stock_tab(self):
        """Build stock inventory management tab"""
        upload_frame = ttk.LabelFrame(self.tab_stock_body, text="Excel Yukle", style="Section.TLabelframe")
        upload_frame.pack(fill=tk.X, padx=6, pady=6)

        self.stock_file_var = tk.StringVar(value="Dosya secilmedi")
        self.stock_region_var = tk.StringVar(value="Ankara")
        self.stock_file_path = None

        row1 = ttk.Frame(upload_frame)
        row1.pack(fill=tk.X, pady=4)
        ttk.Label(row1, text="Dosya").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(row1, textvariable=self.stock_file_var, foreground="#5B9BD5").pack(side=tk.LEFT, padx=6)
        ttk.Button(row1, text="Excel Sec", command=self.select_stock_file).pack(side=tk.LEFT, padx=6)

        row2 = ttk.Frame(upload_frame)
        row2.pack(fill=tk.X, pady=4)
        ttk.Label(row2, text="Bolge").pack(side=tk.LEFT, padx=(0, 8))
        region_combo = ttk.Combobox(row2, textvariable=self.stock_region_var, values=REGIONS, width=14, state="readonly")
        region_combo.pack(side=tk.LEFT, padx=6)
        ttk.Button(row2, text="Yukle", style="Accent.TButton", command=self.upload_stock_file).pack(side=tk.LEFT, padx=6)

        self.stock_status_var = tk.StringVar(value="")
        status_label = ttk.Label(upload_frame, textvariable=self.stock_status_var, foreground="#B0B0B0")
        status_label.pack(anchor="w", padx=6, pady=4)

        # Stock list
        list_frame = ttk.LabelFrame(self.tab_stock_body, text="Stok Envanteri", style="Section.TLabelframe")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        filter_row = ttk.Frame(list_frame)
        filter_row.pack(fill=tk.X, pady=4)
        ttk.Label(filter_row, text="Bolge").pack(side=tk.LEFT, padx=6)
        self.stock_filter_bolge = tk.StringVar(value="ALL")
        bolge_filter = ttk.Combobox(filter_row, textvariable=self.stock_filter_bolge, 
                                     values=["ALL"] + REGIONS, width=12, state="readonly")
        bolge_filter.pack(side=tk.LEFT, padx=6)
        bolge_filter.bind("<<ComboboxSelected>>", lambda _: self.refresh_stock_list())

        ttk.Label(filter_row, text="Durum").pack(side=tk.LEFT, padx=(12, 6))
        self.stock_filter_durum = tk.StringVar(value="ALL")
        durum_filter = ttk.Combobox(filter_row, textvariable=self.stock_filter_durum,
                                     values=["ALL", "VAR", "YOK", "FAZLA"], width=10, state="readonly")
        durum_filter.pack(side=tk.LEFT, padx=6)
        durum_filter.bind("<<ComboboxSelected>>", lambda _: self.refresh_stock_list())

        ttk.Label(filter_row, text="Ara").pack(side=tk.LEFT, padx=(12, 6))
        self.stock_search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_row, textvariable=self.stock_search_var, width=20)
        search_entry.pack(side=tk.LEFT, padx=6)
        search_entry.bind("<Return>", lambda _: self.refresh_stock_list())
        ttk.Button(filter_row, text="Filtrele", command=self.refresh_stock_list).pack(side=tk.LEFT, padx=6)

        # Treeview container
        tree_container = ttk.Frame(list_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)
        
        # Create header frame for column titles - Parent info
        header_frame = tk.Frame(tree_container, bg="#252525", height=25)
        header_frame.grid(row=0, column=0, sticky="ew", columnspan=2)
        header_frame.grid_propagate(False)
        
        headers = ["Stok Kodu", "Ürün Adı", "Seri Sayısı"]
        widths = [120, 280, 100]
        
        for i, (header, width) in enumerate(zip(headers, widths)):
            lbl = tk.Label(header_frame, text=header, bg="#252525", fg="#FFD700", 
                          font=("Segoe UI", 10, "bold"), anchor="w", padx=5)
            lbl.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Treeview columns - 3 column layout
        columns = ("stok_adi", "seri_sayisi")
        self.stock_tree = ttk.Treeview(tree_container, columns=columns, show="tree headings", height=20)

        self.stock_tree.heading("#0", text="Stok Kodu")
        self.stock_tree.heading("stok_adi", text="Ürün Adı")
        self.stock_tree.heading("seri_sayisi", text="Seri Sayısı")
        
        self.stock_tree.column("#0", width=120, anchor="w")
        self.stock_tree.column("stok_adi", width=280, anchor="w")
        self.stock_tree.column("seri_sayisi", width=100, anchor="center")

        self.stock_tree.tag_configure("parent", background="#252525", foreground="#FFD700", font=("Segoe UI", 10, "bold"))
        self.stock_tree.tag_configure("child", background="#1f1f1f", foreground="#e0e0e0", font=("Segoe UI", 9))

        stock_xscroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.stock_tree.xview)
        stock_yscroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.stock_tree.yview)
        self.stock_tree.configure(xscrollcommand=stock_xscroll.set, yscrollcommand=stock_yscroll.set)

        tree_container.columnconfigure(0, weight=1)
        tree_container.rowconfigure(1, weight=1)
        self.stock_tree.grid(row=1, column=0, sticky="nsew")
        stock_yscroll.grid(row=1, column=1, sticky="ns")
        stock_xscroll.grid(row=2, column=0, sticky="ew")
        
        # Click handler for expand/collapse
        self.stock_tree.bind("<Button-1>", self._on_stock_tree_click)

    def select_stock_file(self):
        """Select Excel file for stock upload"""
        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx;*.xls"), ("XLSX", "*.xlsx"), ("XLS", "*.xls"), ("All", "*.*")]
        )
        if path:
            self.stock_file_path = path
            filename = os.path.basename(path)
            self.stock_file_var.set(filename)
            self.stock_status_var.set("")

    def upload_stock_file(self):
        """Upload stock Excel file to server"""
        if not self.stock_file_path:
            messagebox.showwarning("Uyari", "Once dosya secin.")
            return

        if not os.path.isfile(self.stock_file_path):
            messagebox.showwarning("Uyari", "Dosya bulunamadi.")
            return

        bolge = self.stock_region_var.get().strip()
        if not bolge:
            messagebox.showwarning("Uyari", "Bolge secin.")
            return

        self.stock_status_var.set("Dosya isleniyor...")
        self.after(100, self._stock_upload_worker, self.stock_file_path, bolge)

    def _stock_upload_worker(self, file_path, bolge):
        """Process stock file upload in background"""
        try:
            # Read Excel locally first
            rows = load_tabular_file(file_path)
            if not rows:
                self.stock_status_var.set("Dosyada veri bulunamadi!")
                messagebox.showwarning("Uyari", "Dosyada veri bulunamadi.")
                return

            # Parse headers (flexible)
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            
            stok_kod_idx = next((i for i, h in enumerate(headers) if 'stok' in h and 'kod' in h), 0)
            stok_adi_idx = next((i for i, h in enumerate(headers) if 'stok' in h and ('adi' in h or 'ad' in h)), 1)
            seri_no_idx = next((i for i, h in enumerate(headers) if 'seri' in h and 'no' in h), 2)
            seri_sayi_idx = next((i for i, h in enumerate(headers) if 'seri' in h and 'say' in h), 3)

            # Parse nested Excel structure
            # Format: Stok header row followed by seri_no child rows (with empty stok_kod)
            imported = 0
            with db.get_conn() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM stock_inventory WHERE bolge = ?", (bolge,))
                
                i = 1
                while i < len(rows):
                    row = rows[i]
                    
                    # Check if this is a product header (stok_kod not empty)
                    stok_kod = str(row[stok_kod_idx]).strip() if stok_kod_idx < len(row) else ''
                    
                    if stok_kod and stok_kod not in ['', 'nan', 'None', None]:
                        # This is a product header
                        stok_adi = str(row[stok_adi_idx]).strip() if stok_adi_idx < len(row) else ''
                        seri_sayi = 0
                        try:
                            seri_sayi = int(row[seri_sayi_idx]) if seri_sayi_idx < len(row) and row[seri_sayi_idx] else 0
                        except (ValueError, TypeError):
                            pass
                        
                        # Collect all following seri_no rows (child rows where stok_kod is empty)
                        i += 1
                        seri_count = 0
                        while i < len(rows):
                            child_row = rows[i]
                            child_stok_kod = str(child_row[stok_kod_idx]).strip() if stok_kod_idx < len(child_row) else ''
                            
                            # If stok_kod is empty/nan, this is a seri_no row
                            if not child_stok_kod or child_stok_kod in ['', 'nan', 'None', None]:
                                try:
                                    seri_no = str(child_row[seri_no_idx]).strip() if seri_no_idx < len(child_row) else ''
                                    
                                    # Skip if seri_no is empty or just a number (serial position)
                                    if seri_no and seri_no not in ['', 'nan', 'None', None]:
                                        cursor.execute(
                                            """INSERT INTO stock_inventory 
                                               (stok_kod, stok_adi, seri_no, durum, tarih, girdi_yapan, bolge, adet, updated_at) 
                                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                            (stok_kod, stok_adi, seri_no, "OK", datetime.now().strftime("%Y-%m-%d"), 
                                             "system", bolge, 1, datetime.now().isoformat())
                                        )
                                        imported += 1
                                        seri_count += 1
                                    
                                    i += 1
                                except Exception as e:
                                    if self.logger:
                                        self.logger.debug(f"Stock seri row error: {e}")
                                    i += 1
                                    break
                            else:
                                # Next product header found
                                break
                    else:
                        i += 1

            self.stock_status_var.set(f"✓ {imported} kayit yuklendi ({bolge})")
            self._log_action("stock_upload", f"file={os.path.basename(file_path)} region={bolge} count={imported}")
            
            # Refresh view
            self.refresh_stock_list()
            messagebox.showinfo("Basarili", f"{imported} stok kaydı yüklendi.")
            
            # Trigger sync
            self.trigger_sync("stock_upload")

        except Exception as e:
            self.stock_status_var.set(f"✗ Hata: {str(e)[:50]}")
            if self.logger:
                self.logger.error(f"Stock upload error: {e}")
            messagebox.showerror("Hata", f"Yukleme basarisiz: {str(e)[:100]}")

    def refresh_stock_list(self):
        """Refresh stock inventory list"""
        if not hasattr(self, "stock_tree"):
            return

        for item in self.stock_tree.get_children():
            self.stock_tree.delete(item)

        try:
            with db.get_conn() as conn:
                cursor = conn.cursor()

                query = "SELECT * FROM stock_inventory WHERE 1=1"
                params = []

                bolge_filter = self.stock_filter_bolge.get()
                if bolge_filter and bolge_filter != "ALL":
                    query += " AND bolge = ?"
                    params.append(bolge_filter)

                durum_filter = self.stock_filter_durum.get()
                if durum_filter and durum_filter != "ALL":
                    query += " AND durum = ?"
                    params.append(durum_filter)

                search = self.stock_search_var.get().strip()
                if search:
                    query += " AND (stok_kod LIKE ? OR stok_adi LIKE ? OR seri_no LIKE ?)"
                    search_term = f"%{search}%"
                    params.extend([search_term, search_term, search_term])

                query += " ORDER BY stok_kod, seri_no"
                cursor.execute(query, params)
                rows = cursor.fetchall()

            # Group by stok_kod
            grouped = {}
            for row in rows:
                stok_kod, stok_adi, seri_no, durum, tarih, girdi_yapan, bolge, adet, *_ = row
                if stok_kod not in grouped:
                    grouped[stok_kod] = {
                        'stok_adi': stok_adi,
                        'items': []
                    }
                grouped[stok_kod]['items'].append({
                    'seri_no': seri_no,
                    'durum': durum,
                    'tarih': tarih,
                    'girdi_yapan': girdi_yapan,
                    'bolge': bolge,
                    'adet': adet
                })

            # Insert hierarchical list - parent headers with children
            for stok_kod, data in grouped.items():
                # Parent row: stok_kod | stok_adi | seri_sayisi
                parent_id = self.stock_tree.insert("", tk.END, text=stok_kod,
                    values=(
                        data['stok_adi'] or "",
                        f"{len(data['items'])} seri"
                    ),
                    tags=("parent",),
                    open=False
                )
                
                # Child rows: ONLY seri_no (shown in #0 column via text)
                for item in data['items']:
                    self.stock_tree.insert(parent_id, tk.END, text=item['seri_no'] or "",
                        values=("", ""),
                        tags=("child",)
                    )

        except Exception as e:
            if self.logger:
                self.logger.error(f"Stock list refresh error: {e}")

    def _on_stock_tree_click(self, event):
        """Handle treeview click for expand/collapse"""
        item = self.stock_tree.identify('item', event.x, event.y)
        if not item:
            return
        
        # Check if this is a parent item (has children)
        children = self.stock_tree.get_children(item)
        if children:
            # Toggle open state
            current_open = self.stock_tree.item(item, 'open')
            self.stock_tree.item(item, open=not current_open)


if __name__ == "__main__":
    ensure_app_dirs()
    db.init_db()
    app = PuantajApp()
    app.mainloop()
