import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

from calc import calc_day_hours

logger = logging.getLogger("rainstaff")


HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
TITLE_FILL = PatternFill("solid", fgColor="EAF2FB")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_report(output_path, records, settings, date_range_text):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rainstaff"

    company_name = settings.get("company_name", "")
    report_title = settings.get("report_title", "Rainstaff Puantaj ve Mesai Raporu")
    logo_path = settings.get("logo_path", "")

    row = 1
    header_cols = 16
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18

    for r in range(1, 4):
        for c in range(1, header_cols + 1):
            ws.cell(row=r, column=c).fill = TITLE_FILL

    if logo_path and os.path.isfile(logo_path):
        try:
            img = Image(logo_path)
            img.width = 110
            img.height = 60
            ws.add_image(img, "A1")
        except Exception as e:
            logger.warning("Logo yukleme basarısız (%s): %s", logo_path, str(e))

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=header_cols)
    title_cell = ws.cell(row=row, column=2, value=company_name)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(vertical="center", horizontal="left")
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=header_cols)
    subtitle_cell = ws.cell(row=row, column=2, value=report_title)
    subtitle_cell.font = Font(size=11, bold=True)
    subtitle_cell.alignment = Alignment(vertical="center", horizontal="left")
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=header_cols)
    date_cell = ws.cell(row=row, column=2, value=date_range_text or "Tarih Araligi: -")
    date_cell.font = Font(size=10)
    date_cell.alignment = Alignment(vertical="center", horizontal="left")
    row += 2

    headers = [
        "Calisan",
        "Bolge",
        "Tarih",
        "Giris",
        "Cikis",
        "Mola (dk)",
        "Calisilan (s)",
        "Plan (s)",
        "Fazla Mesai (s)",
        "Gece (s)",
        "Geceye Tasan (s)",
        "Ozel Gun",
        "Ozel Gun Normal (s)",
        "Ozel Gun Fazla (s)",
        "Ozel Gun Gece (s)",
        "Not",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    row += 1

    totals = {}
    special_records = []
    overnight_records = []
    for _, emp_id, name, work_date, start_time, end_time, break_minutes, is_special, notes, region in records:
        (
            worked,
            scheduled,
            overtime,
            night_hours,
            overnight_hours,
            special_normal,
            special_overtime,
            special_night,
        ) = calc_day_hours(
            work_date,
            start_time,
            end_time,
            break_minutes,
            settings,
            is_special,
        )
        ws.cell(row=row, column=1, value=name).border = BORDER
        ws.cell(row=row, column=2, value=region or "").border = BORDER
        ws.cell(row=row, column=3, value=work_date).border = BORDER
        ws.cell(row=row, column=4, value=start_time).border = BORDER
        ws.cell(row=row, column=5, value=end_time).border = BORDER
        ws.cell(row=row, column=6, value=break_minutes).border = BORDER
        ws.cell(row=row, column=7, value=worked).border = BORDER
        ws.cell(row=row, column=8, value=scheduled).border = BORDER
        ws.cell(row=row, column=9, value=overtime).border = BORDER
        ws.cell(row=row, column=10, value=night_hours).border = BORDER
        ws.cell(row=row, column=11, value=overnight_hours).border = BORDER
        ws.cell(row=row, column=12, value="Evet" if is_special else "Hayir").border = BORDER
        ws.cell(row=row, column=13, value=special_normal).border = BORDER
        ws.cell(row=row, column=14, value=special_overtime).border = BORDER
        ws.cell(row=row, column=15, value=special_night).border = BORDER
        ws.cell(row=row, column=16, value=notes or "").border = BORDER

        if emp_id not in totals:
            totals[emp_id] = {
                "name": name,
                "worked": 0.0,
                "scheduled": 0.0,
                "overtime": 0.0,
                "night": 0.0,
                "overnight": 0.0,
                "special_normal": 0.0,
                "special_overtime": 0.0,
                "special_night": 0.0,
            }
        totals[emp_id]["worked"] += worked
        totals[emp_id]["scheduled"] += scheduled
        totals[emp_id]["overtime"] += overtime
        totals[emp_id]["night"] += night_hours
        totals[emp_id]["overnight"] += overnight_hours
        totals[emp_id]["special_normal"] += special_normal
        totals[emp_id]["special_overtime"] += special_overtime
        totals[emp_id]["special_night"] += special_night
        if is_special:
            special_records.append((name, work_date, special_normal, special_overtime, special_night))
        if overnight_hours > 0:
            overnight_records.append((name, work_date, overnight_hours))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Ozet").font = Font(bold=True)
    row += 1
    summary_headers = [
        "Calisan",
        "Toplam Calisilan (s)",
        "Toplam Plan (s)",
        "Toplam Fazla Mesai (s)",
        "Toplam Gece (s)",
        "Toplam Geceye Tasan (s)",
    ]
    for col, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for _, data in sorted(totals.items(), key=lambda x: x[1]["name"]):
        ws.cell(row=row, column=1, value=data["name"]).border = BORDER
        ws.cell(row=row, column=2, value=round(data["worked"], 2)).border = BORDER
        ws.cell(row=row, column=3, value=round(data["scheduled"], 2)).border = BORDER
        ws.cell(row=row, column=4, value=round(data["overtime"], 2)).border = BORDER
        ws.cell(row=row, column=5, value=round(data["night"], 2)).border = BORDER
        ws.cell(row=row, column=6, value=round(data["overnight"], 2)).border = BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Ozel Gun Calismalari").font = Font(bold=True)
    row += 1
    special_headers = [
        "Calisan",
        "Tarih",
        "Ozel Gun Normal (s)",
        "Ozel Gun Fazla (s)",
        "Ozel Gun Gece (s)",
    ]
    for col, header in enumerate(special_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    if special_records:
        for name, work_date, spec_norm, spec_ot, spec_night in special_records:
            ws.cell(row=row, column=1, value=name).border = BORDER
            ws.cell(row=row, column=2, value=work_date).border = BORDER
            ws.cell(row=row, column=3, value=spec_norm).border = BORDER
            ws.cell(row=row, column=4, value=spec_ot).border = BORDER
            ws.cell(row=row, column=5, value=spec_night).border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Kayit yok").border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Geceye Tasan Mesailer").font = Font(bold=True)
    row += 1
    overnight_headers = ["Calisan", "Tarih", "Geceye Tasan (s)"]
    for col, header in enumerate(overnight_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    if overnight_records:
        for name, work_date, overnight in overnight_records:
            ws.cell(row=row, column=1, value=name).border = BORDER
            ws.cell(row=row, column=2, value=work_date).border = BORDER
            ws.cell(row=row, column=3, value=overnight).border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Kayit yok").border = BORDER
        row += 1

    for col in range(1, 17):
        ws.column_dimensions[chr(64 + col)].width = 16

    ws.freeze_panes = "A6"
    wb.save(output_path)


def export_vehicle_weekly_report(
    output_path,
    plate,
    week_start,
    prev_week,
    checklist,
    prev_results,
    current_results,
    current_km,
    prev_km,
    vehicle_row,
    current_fault,
    prev_fault,
    service_visits,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Arac Kontrol"

    ws.cell(row=1, column=1, value="Arac Haftalik Kontrol Raporu").font = Font(size=12, bold=True)
    ws.cell(row=2, column=1, value=f"Plaka: {plate}")
    ws.cell(row=3, column=1, value=f"Hafta: {week_start}")
    ws.cell(row=4, column=1, value=f"Onceki Hafta: {prev_week or '-'}")
    ws.cell(row=5, column=1, value=f"KM (Bu Hafta): {current_km or '-'}")
    ws.cell(row=6, column=1, value=f"KM (Onceki): {prev_km or '-'}")

    if vehicle_row:
        (
            _vid,
            _plate,
            brand,
            model,
            year,
            km,
            inspection_date,
            insurance_date,
            maintenance_date,
            oil_change_date,
            oil_change_km,
            oil_interval_km,
            _notes,
            _region,
        ) = vehicle_row
        today = datetime.now().date()
        if inspection_date:
            insp_dt = datetime.strptime(inspection_date, "%Y-%m-%d").date()
            diff = (insp_dt - today).days
            insp_text = f"{inspection_date} ({diff} gun)" if diff >= 0 else f"{inspection_date} ({abs(diff)} gun gecikme)"
        else:
            insp_text = "-"
        ws.cell(row=7, column=1, value=f"Muayene: {insp_text}")
        ws.cell(row=8, column=1, value=f"Son Bakim: {maintenance_date or '-'}")
        ws.cell(row=9, column=1, value=f"Son Yag Degisimi: {oil_change_date or '-'}")
        interval_km = oil_interval_km or 14000
        if interval_km and oil_change_km is not None and current_km:
            remaining = interval_km - (current_km - oil_change_km)
            oil_text = "Geldi" if remaining <= 0 else f"{remaining} km"
        else:
            oil_text = "-"
        ws.cell(row=10, column=1, value=f"Yag Periyodu: {interval_km or '-'} km, Kalan: {oil_text}")

    row = 12
    ws.cell(row=row, column=1, value="Ariza Bilgisi").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Bu Hafta Ariza: {current_fault.get('title') or '-'}")
    row += 1
    ws.cell(row=row, column=1, value=f"Bu Hafta Durum: {current_fault.get('status') or '-'}")
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Sanayiye Gitti: {'Evet' if current_fault.get('service') else 'Hayir'}",
    )
    row += 1
    ws.cell(row=row, column=1, value=f"Onceki Hafta Ariza: {prev_fault.get('title') or '-'}")
    row += 1
    ws.cell(row=row, column=1, value=f"Onceki Hafta Durum: {prev_fault.get('status') or '-'}")
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Onceki Hafta Sanayi: {'Evet' if prev_fault.get('service') else 'Hayir'}",
    )
    row += 2

    ws.cell(row=row, column=1, value="Sanayi Kayitlari (Hafta)").font = Font(bold=True)
    row += 1
    svc_headers = ["Gidis", "Donus", "Neden", "Masraf", "Not"]
    for col, header in enumerate(svc_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    if service_visits:
        for visit in service_visits:
            _sid, _vid, _plate, _fid, _title, start_date, end_date, reason, cost, notes, _region = visit
            ws.cell(row=row, column=1, value=start_date or "-").border = BORDER
            ws.cell(row=row, column=2, value=end_date or "Sanayide").border = BORDER
            ws.cell(row=row, column=3, value=reason or "-").border = BORDER
            ws.cell(row=row, column=4, value=cost if cost is not None else "-").border = BORDER
            ws.cell(row=row, column=5, value=notes or "-").border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Kayit yok").border = BORDER
        row += 1
    row += 1

    headers = ["Kontrol", "Onceki Hafta", "Bu Hafta", "Durum"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    def normalize_status(value):
        text = str(value or "").strip()
        mapping = {
            "OK": "Olumlu",
            "Issue": "Olumsuz",
            "NA": "Bilinmiyor",
            "Olumlu": "Olumlu",
            "Olumsuz": "Olumsuz",
            "Bilinmiyor": "Bilinmiyor",
        }
        return mapping.get(text, text or "-")

    for item_key, label in checklist:
        prev_status = normalize_status(prev_results.get(item_key, "-"))
        curr_status = normalize_status(current_results.get(item_key, "-"))
        if prev_status == curr_status:
            change = "Ayni"
        elif prev_status == "Olumsuz" and curr_status == "Olumsuz":
            change = "Tekrar eden sorun"
        elif curr_status == "Olumsuz" and prev_status != "Olumsuz":
            change = "Kotulesti"
        elif prev_status == "Olumsuz" and curr_status != "Olumsuz":
            change = "Iyilesti"
        else:
            change = "Degisti"

        ws.cell(row=row, column=1, value=label).border = BORDER
        ws.cell(row=row, column=2, value=prev_status).border = BORDER
        ws.cell(row=row, column=3, value=curr_status).border = BORDER
        ws.cell(row=row, column=4, value=change).border = BORDER
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 22

    wb.save(output_path)


def export_vehicle_card_report(output_path, plate, vehicle_row, inspections, faults, services):
    wb = Workbook()
    ws = wb.active
    ws.title = "Arac Karti"

    ws.cell(row=1, column=1, value="Arac Karti").font = Font(size=12, bold=True)
    ws.cell(row=2, column=1, value=f"Plaka: {plate}")

    if vehicle_row:
        (
            _vid,
            _plate,
            brand,
            model,
            year,
            km,
            inspection_date,
            insurance_date,
            maintenance_date,
            oil_change_date,
            oil_change_km,
            oil_interval_km,
            notes,
            _region,
        ) = vehicle_row
        ws.cell(row=3, column=1, value=f"Marka/Model: {brand} {model}")
        ws.cell(row=4, column=1, value=f"Yil: {year}")
        ws.cell(row=5, column=1, value=f"KM: {km or '-'}")
        ws.cell(row=6, column=1, value=f"Muayene: {inspection_date or '-'}")
        ws.cell(row=7, column=1, value=f"Sigorta: {insurance_date or '-'}")
        ws.cell(row=8, column=1, value=f"Bakim: {maintenance_date or '-'}")
        ws.cell(row=9, column=1, value=f"Yag Degisim: {oil_change_date or '-'}")
        ws.cell(row=10, column=1, value=f"Yag KM: {oil_change_km or '-'}")
        ws.cell(row=11, column=1, value=f"Yag Periyot: {(oil_interval_km or 14000)} km")
        if notes:
            ws.cell(row=12, column=1, value=f"Not: {notes}")

    fault_title_map = {row[0]: row[3] for row in faults} if faults else {}

    row = 14
    ws.cell(row=row, column=1, value="Ariza Kayitlari").font = Font(bold=True)
    row += 1
    headers = ["Baslik", "Durum", "Acilis", "Kapanis", "Aciklama"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    if faults:
        for fault in faults:
            _fid, _vid, _plate, title, desc, opened_date, closed_date, status, _region = fault
            ws.cell(row=row, column=1, value=title or "-").border = BORDER
            ws.cell(row=row, column=2, value=status or "-").border = BORDER
            ws.cell(row=row, column=3, value=opened_date or "-").border = BORDER
            ws.cell(row=row, column=4, value=closed_date or "-").border = BORDER
            ws.cell(row=row, column=5, value=desc or "-").border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Kayit yok").border = BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Sanayi Kayitlari").font = Font(bold=True)
    row += 1
    headers = ["Ariza", "Gidis", "Donus", "Masraf", "Neden", "Not"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    if services:
        for visit in services:
            _sid, _vid, _plate, _fid, title, start_date, end_date, reason, cost, notes, _region = visit
            ws.cell(row=row, column=1, value=title or "-").border = BORDER
            ws.cell(row=row, column=2, value=start_date or "-").border = BORDER
            ws.cell(row=row, column=3, value=end_date or "Sanayide").border = BORDER
            ws.cell(row=row, column=4, value=cost if cost is not None else "-").border = BORDER
            ws.cell(row=row, column=5, value=reason or "-").border = BORDER
            ws.cell(row=row, column=6, value=notes or "-").border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Kayit yok").border = BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Kontroller").font = Font(bold=True)
    row += 1
    headers = ["Tarih", "Hafta", "Surucu", "KM", "Ariza", "Durum", "Sanayi", "Not"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    if inspections:
        for row_data in inspections:
            (
                _iid,
                _veh_id,
                _plate,
                _driver_id,
                driver_name,
                inspect_date,
                week_start,
                km_val,
                note_val,
                fault_id,
                fault_status,
                service_visit,
            ) = row_data
            ws.cell(row=row, column=1, value=inspect_date or "-").border = BORDER
            ws.cell(row=row, column=2, value=week_start or "-").border = BORDER
            ws.cell(row=row, column=3, value=driver_name or "-").border = BORDER
            ws.cell(row=row, column=4, value=km_val or "-").border = BORDER
            ws.cell(row=row, column=5, value=fault_title_map.get(fault_id, "-")).border = BORDER
            ws.cell(row=row, column=6, value=fault_status or "-").border = BORDER
            ws.cell(row=row, column=7, value="Evet" if service_visit else "Hayir").border = BORDER
            ws.cell(row=row, column=8, value=note_val or "-").border = BORDER
            row += 1
    else:
        ws.cell(row=row, column=1, value="Kayit yok").border = BORDER
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    wb.save(output_path)
