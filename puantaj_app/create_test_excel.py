#!/usr/bin/env python3
"""Generate test Excel file matching the real warehouse structure"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Stok"

# Headers
headers = ["Stok Kod", "Stok Adı", "Adet", "Seri Sayım"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    cell.font = Font(bold=True)

# Product 1: 17 - TABLET TUZ (1 serial)
ws.cell(row=2, column=1, value=17)
ws.cell(row=2, column=2, value="TABLET TUZ 25 KG LUK ÇUVAL")
ws.cell(row=2, column=3, value=0)
ws.cell(row=2, column=4, value=1)

ws.cell(row=3, column=2, value=1)  # Seri no order
ws.cell(row=3, column=3, value="A084")  # ACTUAL serial number

# Product 2: 754 - POS KARBON (11 serials)
ws.cell(row=4, column=1, value=754)
ws.cell(row=4, column=2, value="POS KARBON REVERSE OSMOS")
ws.cell(row=4, column=3, value=7)
ws.cell(row=4, column=4, value=11)

# 11 serial numbers for product 754
serials = [
    "ADSBGASDG",
    "ST87085",
    "754xST87083",
    "754xST87088",
    "754xST87081",
    "754xST87086",
    "754xST84710",
    "754xST87084",
    "754xST87082",
    "754xST87087",
    "754xST87080"
]

for idx, serial in enumerate(serials, 1):
    row = 4 + idx
    ws.cell(row=row, column=2, value=idx)  # Serial order number
    ws.cell(row=row, column=3, value=serial)  # ACTUAL serial number

# Adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15

# Save
output_path = r"C:\Users\rainwater\Desktop\test_stok.xlsx"
wb.save(output_path)
print(f"✓ Test Excel created: {output_path}")
print(f"  - Ürün 1: Stok 17 (1 serial)")
print(f"  - Ürün 2: Stok 754 (11 serials)")
print(f"\nNow upload this in the app!")
