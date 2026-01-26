"""Test deneme.xlsx with fixed parser logic"""
import sys
sys.path.insert(0, 'puantaj_app')

from openpyxl import load_workbook

# Load the file
file_path = r'C:\Users\rainwater\Desktop\deneme.xlsx'

wb = load_workbook(file_path, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))

print(f"📊 Total rows: {len(rows)}")
print("\n🔍 PARSING WITH FIX:\n")

# Check if first row is empty or has no valid headers (all None)
first_row_empty = all(cell is None or str(cell).strip() == '' for cell in rows[0])

# Also check if headers are all None/empty
headers = [str(h).strip().lower() if h else '' for h in rows[0]]
has_valid_headers = any('stok' in h or 'seri' in h for h in headers)

if first_row_empty or not has_valid_headers:
    print("✓ No valid headers detected - using default column indices")
    print(f"  First row: {rows[0]}")
    stok_kod_idx = 0
    stok_adi_idx = 1
    seri_no_idx = 2
    seri_sayi_idx = 3
    start_row = 0  # NO HEADERS - start from row 0 (data starts immediately)
else:
    print("✓ Valid headers found - parsing headers")
    stok_kod_idx = next((i for i, h in enumerate(headers) if 'stok' in h and 'kod' in h), 0)
    stok_adi_idx = next((i for i, h in enumerate(headers) if 'stok' in h and ('adi' in h or 'ad' in h)), 1)
    seri_no_idx = next((i for i, h in enumerate(headers) if 'seri' in h and 'no' in h), 2)
    seri_sayi_idx = next((i for i, h in enumerate(headers) if 'seri' in h and 'say' in h), 3)
    start_row = 1

print(f"Column indices: stok_kod={stok_kod_idx}, stok_adi={stok_adi_idx}, seri_no={seri_no_idx}, seri_sayi={seri_sayi_idx}")
print(f"Starting from row: {start_row}\n")

# DEBUG: Show first few rows
print("DEBUG - First 3 rows:")
for idx in range(min(3, len(rows))):
    print(f"  Row {idx}: {rows[idx]}")
print()

parsed_items = []
i = start_row

while i < len(rows):
    row = rows[i]
    
    # Check if this is a product header (stok_kod not empty)
    stok_kod_value = row[stok_kod_idx] if stok_kod_idx < len(row) else None
    stok_kod = str(stok_kod_value).strip() if stok_kod_value is not None else ''
    
    print(f"Row {i}: stok_kod_value={stok_kod_value}, stok_kod='{stok_kod}'")
    
    if stok_kod and stok_kod not in ['', 'nan', 'None', 'None']:
        # This is a product header
        stok_adi = str(row[stok_adi_idx]).strip() if stok_adi_idx < len(row) and row[stok_adi_idx] else ''
        print(f"\n✓ Found product: {stok_kod} - {stok_adi}")
        
        # Collect all following seri_no rows
        i += 1
        seri_count = 0
        
        while i < len(rows):
            child_row = rows[i]
            child_stok_kod_value = child_row[stok_kod_idx] if stok_kod_idx < len(child_row) else None
            child_stok_kod = str(child_stok_kod_value).strip() if child_stok_kod_value is not None else ''
            
            print(f"  Child row {i}: stok_kod={child_stok_kod_value}")
            
            # If stok_kod is empty/None, this is a seri_no row
            if child_stok_kod_value is None or not child_stok_kod or child_stok_kod in ['', 'nan', 'None']:
                # Seri no is in column 2 for child rows (based on check_deneme output)
                seri_no_value = child_row[seri_no_idx] if seri_no_idx < len(child_row) else None
                seri_no = str(seri_no_value).strip() if seri_no_value is not None else ''
                
                print(f"    seri_no_value={seri_no_value}, seri_no='{seri_no}'")
                
                # Extract actual serial number (remove numbering like "1 ST87088")
                # BUT: Only if second part is NOT purely numeric (to preserve numeric serials)
                if seri_no:
                    parts = seri_no.split(maxsplit=1)
                    if len(parts) == 2 and parts[0].isdigit():
                        # Check if second part is purely numeric - BU KONTROL KALDIRILDI
                        # Kullanıcı isteği: Rakam bile olsa baştaki "1 " gibi kısımlar atılmalı
                        # if not parts[1].isdigit():
                        seri_no = parts[1]  # "ST87088" veya "8697236914625" - remove numbering
                        # else: keep as-is - BU DA İPTAL
                
                # Skip if seri_no is empty (but allow all other values including pure numbers)
                if seri_no and seri_no not in ['', 'nan', 'None', 'None']:
                    parsed_items.append({
                        'stok_kod': stok_kod,
                        'stok_adi': stok_adi,
                        'seri_no': seri_no
                    })
                    seri_count += 1
                    print(f"  └─ Serial {seri_count}: {seri_no}")
                
                i += 1
            else:
                # Next product header found
                break
    else:
        i += 1

print(f"\n{'='*60}")
print(f"✅ RESULT: Parsed {len(parsed_items)} serials")
for item in parsed_items:
    print(f"  {item['stok_kod']} | {item['stok_adi']} | {item['seri_no']}")

if len(parsed_items) == 5:
    print(f"\n✅ TEST PASSED: All 5 serials parsed correctly!")
else:
    print(f"\n❌ TEST FAILED: Expected 5 serials, got {len(parsed_items)}")
