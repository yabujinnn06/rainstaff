# Stock Inventory Routes
# Stok tablosu ve Excel upload/download işlemleri

import os
import sqlite3
from datetime import datetime
from flask import request, jsonify, send_file
from openpyxl import load_workbook

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'puantaj.db')


def init_stock_routes(app):
    """Register stock inventory routes"""
    
    @app.route('/stock/list', methods=['GET'])
    def list_stock():
        """Get stock inventory (filterable)"""
        try:
            bolge = request.args.get('bolge', '')
            durum = request.args.get('durum', '')  # VAR, YOK, FAZLA
            search = request.args.get('search', '')
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT * FROM stock_inventory WHERE 1=1"
            params = []
            
            if bolge:
                query += " AND bolge = ?"
                params.append(bolge)
            
            if durum:
                query += " AND durum = ?"
                params.append(durum)
            
            if search:
                query += " AND (stok_kod LIKE ? OR stok_adi LIKE ? OR seri_no LIKE ?)"
                search_term = f"%{search}%"
                params.extend([search_term, search_term, search_term])
            
            query += " ORDER BY stok_kod, seri_no"
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # Group by stok_kod for hierarchical display
            grouped = {}
            for row in rows:
                stok_kod = row['stok_kod']
                if stok_kod not in grouped:
                    grouped[stok_kod] = {
                        'stok_kod': stok_kod,
                        'stok_adi': row['stok_adi'],
                        'adet': row['adet'],
                        'bolge': row['bolge'],
                        'items': []
                    }
                
                grouped[stok_kod]['items'].append({
                    'id': row['id'],
                    'seri_no': row['seri_no'],
                    'durum': row['durum'],
                    'tarih': row['tarih'],
                    'girdi_yapan': row['girdi_yapan']
                })
            
            conn.close()
            
            return jsonify({'success': True, 'data': list(grouped.values())}), 200
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    
    @app.route('/stock/upload', methods=['POST'])
    def upload_stock_excel():
        """Upload stock Excel file and replace inventory"""
        try:
            if 'file' not in request.files:
                return jsonify({'success': False, 'error': 'No file provided'}), 400
            
            file = request.files['file']
            bolge = request.form.get('bolge', 'Ankara')
            
            if not file.filename.endswith(('.xlsx', '.xls')):
                return jsonify({'success': False, 'error': 'Only Excel files allowed'}), 400
            
            # Read Excel
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    rows.append(row)
            
            if not rows:
                return jsonify({'success': False, 'error': 'No data in Excel'}), 400
            
            # Parse headers (flexible)
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            
            stok_kod_idx = next((i for i, h in enumerate(headers) if 'stok' in h and 'kod' in h), 0)
            stok_adi_idx = next((i for i, h in enumerate(headers) if 'stok' in h and 'adi' in h), 1)
            seri_no_idx = next((i for i, h in enumerate(headers) if 'seri' in h), 2)
            durum_idx = next((i for i, h in enumerate(headers) if 'durum' in h), None)
            tarih_idx = next((i for i, h in enumerate(headers) if 'tarih' in h), None)
            girdi_yapan_idx = next((i for i, h in enumerate(headers) if 'girdi' in h), None)
            adet_idx = next((i for i, h in enumerate(headers) if 'adet' in h), None)
            
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # Clear existing records for this region
            cursor.execute("DELETE FROM stock_inventory WHERE bolge = ?", (bolge,))
            
            imported = 0
            i = 1  # Start after headers
            
            while i < len(rows):
                row = rows[i]
                stok_kod_value = row[stok_kod_idx] if stok_kod_idx < len(row) else None
                
                # Check if this is a main product row (has stok_kod)
                if stok_kod_value:
                    # Main product row
                    stok_kod = str(stok_kod_value).strip()
                    stok_adi = str(row[stok_adi_idx]).strip() if stok_adi_idx < len(row) and row[stok_adi_idx] else ''
                    adet = int(row[adet_idx]) if adet_idx and adet_idx < len(row) and row[adet_idx] else None
                    
                    i += 1
                    
                    # Collect child rows (serial numbers)
                    while i < len(rows):
                        child_row = rows[i]
                        child_stok_kod = child_row[stok_kod_idx] if stok_kod_idx < len(child_row) else None
                        
                        # If next row has stok_kod, it's a new product
                        if child_stok_kod:
                            break
                        
                        # This is a serial row
                        try:
                            serial_value = str(child_row[stok_adi_idx]).strip() if stok_adi_idx < len(child_row) and child_row[stok_adi_idx] else ''
                            
                            if serial_value:
                                # Extract actual serial number (remove numbering like "1 ST87088")
                                parts = serial_value.split(maxsplit=1)
                                if len(parts) == 2 and parts[0].isdigit():
                                    seri_no = parts[1]  # "ST87088"
                                else:
                                    seri_no = serial_value  # Use as-is
                                
                                # Get optional fields
                                durum = str(child_row[durum_idx]).strip() if durum_idx and durum_idx < len(child_row) and child_row[durum_idx] else 'OK'
                                tarih = str(child_row[tarih_idx]).strip() if tarih_idx and tarih_idx < len(child_row) and child_row[tarih_idx] else datetime.now().strftime('%Y-%m-%d')
                                girdi_yapan = str(child_row[girdi_yapan_idx]).strip() if girdi_yapan_idx and girdi_yapan_idx < len(child_row) and child_row[girdi_yapan_idx] else 'system'
                                
                                cursor.execute(
                                    """INSERT INTO stock_inventory 
                                       (stok_kod, stok_adi, seri_no, durum, tarih, girdi_yapan, bolge, adet, updated_at) 
                                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                    (stok_kod, stok_adi, seri_no, durum, tarih, girdi_yapan, bolge, adet, datetime.now().isoformat())
                                )
                                imported += 1
                        
                        except Exception as e:
                            pass  # Skip problematic rows
                        
                        i += 1
                else:
                    # Orphan row without parent, skip
                    i += 1
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': f'{imported} kayıt yüklendi',
                'imported': imported,
                'bolge': bolge
            }), 200
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    
    @app.route('/stock/export', methods=['GET'])
    def export_stock_excel():
        """Export current inventory as Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            bolge = request.args.get('bolge', '')
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT * FROM stock_inventory WHERE 1=1"
            params = []
            
            if bolge:
                query += " AND bolge = ?"
                params.append(bolge)
            
            query += " ORDER BY stok_kod, seri_no"
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Stok"
            
            # Headers
            headers = ['Stok Kod', 'Stok Adı', 'Seri No', 'Durum', 'Tarih', 'Girdi Yapan', 'Adet', 'Bölge']
            ws.append(headers)
            
            # Header formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Data
            for row in rows:
                ws.append([
                    row['stok_kod'],
                    row['stok_adi'],
                    row['seri_no'],
                    row['durum'],
                    row['tarih'],
                    row['girdi_yapan'],
                    row['adet'],
                    row['bolge']
                ])
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            conn.close()
            
            # Save to temp file
            temp_path = '/tmp/stok_envanter.xlsx'
            wb.save(temp_path)
            
            return send_file(temp_path, as_attachment=True, download_name='stok_envanter.xlsx'), 200
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
