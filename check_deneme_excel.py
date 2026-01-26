"""Check deneme.xlsx file structure"""
import sys
sys.path.insert(0, 'puantaj_app')

from openpyxl import load_workbook

# Load the file
file_path = r'C:\Users\rainwater\Desktop\deneme.xlsx'

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"📊 Excel Dosyası: {file_path}")
    print(f"📄 Aktif Sayfa: {ws.title}")
    print(f"📏 Satır Sayısı: {ws.max_row}")
    print(f"📏 Sütun Sayısı: {ws.max_column}")
    print("\n" + "="*80)
    
    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    
    print(f"\n🔍 İLK 20 SATIR:\n")
    for i, row in enumerate(rows[:20]):
        print(f"Satır {i}: {row}")
    
    print("\n" + "="*80)
    
    # Check headers
    if rows:
        headers = rows[0]
        print(f"\n📋 BAŞLIKLAR:")
        for i, h in enumerate(headers):
            print(f"  Sütun {i}: '{h}'")
        
        # Analyze structure
        print(f"\n🔍 YAPI ANALİZİ:")
        
        # Check for stok_kod column
        stok_kod_idx = None
        for i, h in enumerate(headers):
            h_str = str(h).strip().lower() if h else ''
            if 'stok' in h_str and 'kod' in h_str:
                stok_kod_idx = i
                print(f"  ✓ Stok Kod sütunu bulundu: Sütun {i} ('{headers[i]}')")
                break
        
        if stok_kod_idx is None:
            print(f"  ✗ HATA: 'Stok Kod' sütunu bulunamadı!")
            print(f"  Beklenen: 'Stok Kod', 'stok kod', 'STOK KOD' gibi")
        
        # Check for stok_adi column
        stok_adi_idx = None
        for i, h in enumerate(headers):
            h_str = str(h).strip().lower() if h else ''
            if 'stok' in h_str and ('adi' in h_str or 'ad' in h_str):
                stok_adi_idx = i
                print(f"  ✓ Stok Adı sütunu bulundu: Sütun {i} ('{headers[i]}')")
                break
        
        if stok_adi_idx is None:
            print(f"  ✗ HATA: 'Stok Adı' sütunu bulunamadı!")
            print(f"  Beklenen: 'Stok Adı', 'stok adi', 'STOK ADI' gibi")
        
        # Count products and serials
        product_count = 0
        serial_count = 0
        
        for i in range(1, len(rows)):
            row = rows[i]
            if stok_kod_idx is not None and stok_kod_idx < len(row):
                stok_kod = row[stok_kod_idx]
                if stok_kod and str(stok_kod).strip() not in ['', 'nan', 'None']:
                    product_count += 1
                else:
                    serial_count += 1
        
        print(f"\n📊 VERİ İSTATİSTİKLERİ:")
        print(f"  Ürün satırları (Stok Kod dolu): {product_count}")
        print(f"  Seri satırları (Stok Kod boş): {serial_count}")
        print(f"  Toplam veri satırı: {len(rows) - 1}")
        
        if product_count == 0:
            print(f"\n⚠️  UYARI: Hiç ürün satırı bulunamadı!")
            print(f"  Stok Kod sütununda değer olan satır yok.")
        
        if serial_count == 0:
            print(f"\n⚠️  UYARI: Hiç seri satırı bulunamadı!")
            print(f"  Stok Kod sütunu boş olan satır yok.")

except FileNotFoundError:
    print(f"❌ HATA: Dosya bulunamadı: {file_path}")
except Exception as e:
    print(f"❌ HATA: {e}")
    import traceback
    traceback.print_exc()
