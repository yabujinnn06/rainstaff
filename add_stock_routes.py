#!/usr/bin/env python3
"""Add stock routes to server/app.py"""

import os

# Read the app.py file
with open('server/app.py', 'r', encoding='utf-8') as f:
    content = f.read()

# Stock routes code
stock_routes = """
# Stock Management Routes
@app.route('/stock')
def stock_dashboard():
    \"\"\"Stock inventory dashboard\"\"\"
    if 'user' not in session:
        return redirect(url_for('login'))
    
    return render_template('stock.html')


@app.route('/stock/list')
def stock_list():
    \"\"\"Get stock list with filters\"\"\"
    if 'user' not in session:
        return {'success': False, 'error': 'Unauthorized'}, 401
    
    try:
        bolge = request.args.get('bolge', '')
        durum = request.args.get('durum', '')
        search = request.args.get('search', '')
        
        if not db_exists():
            return {'success': True, 'data': []}
        
        with get_conn() as conn:
            query = 'SELECT * FROM stock_inventory WHERE 1=1'
            params = []
            
            if bolge:
                query += ' AND bolge = ?'
                params.append(bolge)
            
            if durum:
                query += ' AND durum = ?'
                params.append(durum)
            
            if search:
                query += ' AND (stok_kod LIKE ? OR seri_no LIKE ? OR stok_adi LIKE ?)'
                search_term = f'%{search}%'
                params.extend([search_term, search_term, search_term])
            
            query += ' ORDER BY stok_kod, seri_no'
            
            stocks = conn.execute(query, params).fetchall()
            
            # Convert rows to dicts
            if stocks:
                cursor = conn.execute(query, params)
                columns = [desc[0] for desc in cursor.description]
            else:
                columns = ['id', 'stok_kod', 'stok_adi', 'seri_no', 'durum', 'tarih', 'girdi_yapan', 'bolge', 'adet', 'created_at', 'updated_at']
            
            data = []
            for row in stocks:
                stock_dict = dict(zip(columns, row))
                data.append(stock_dict)
            
            return {'success': True, 'data': data}
    
    except Exception as e:
        return {'success': False, 'error': str(e)}, 500


@app.route('/stock/export')
def stock_export():
    \"\"\"Export stock list to Excel\"\"\"
    if 'user' not in session:
        return redirect(url_for('login'))
    
    try:
        bolge = request.args.get('bolge', '')
        
        if not db_exists():
            return abort(404)
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Stok Envanteri'
        
        # Headers
        headers = ['Stok Kod', 'Stok Adı', 'Seri No', 'Durum', 'Tarih', 'Girdi Yapan', 'Adet', 'Bölge']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        with get_conn() as conn:
            query = 'SELECT stok_kod, stok_adi, seri_no, durum, tarih, girdi_yapan, adet, bolge FROM stock_inventory'
            params = []
            
            if bolge:
                query += ' WHERE bolge = ?'
                params.append(bolge)
            
            query += ' ORDER BY stok_kod, seri_no'
            
            stocks = conn.execute(query, params).fetchall()
            
            for stock in stocks:
                ws.append(stock)
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 12
        
        # Save to temp file
        import tempfile
        temp_dir = tempfile.gettempdir()
        filename = os.path.join(temp_dir, 'stok_envanter.xlsx')
        wb.save(filename)
        
        from flask import send_file
        return send_file(filename, as_attachment=True, download_name='stok_envanter.xlsx')
    
    except Exception as e:
        return {'success': False, 'error': str(e)}, 500

"""

# Find the /auto-sync route and insert before it
auto_sync_pos = content.find('@app.route("/auto-sync"')
if auto_sync_pos != -1:
    # Find the start of the line
    line_start = content.rfind('\n', 0, auto_sync_pos) + 1
    
    # Insert stock routes before the /auto-sync route
    new_content = content[:line_start] + stock_routes + '\n' + content[line_start:]
    
    # Write back
    with open('server/app.py', 'w', encoding='utf-8') as f:
        f.write(new_content)
    
    print('✓ Stock routes added successfully to server/app.py')
    print(f'  Lines added: {len(stock_routes.splitlines())}')
else:
    print('✗ Could not find /auto-sync route in server/app.py')
    exit(1)
