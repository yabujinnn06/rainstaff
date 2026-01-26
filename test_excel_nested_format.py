"""Test nested Excel format with numbered serials"""
import sys
sys.path.insert(0, 'puantaj_app')

from openpyxl import Workbook, load_workbook

# Create test Excel matching the screenshot format
wb = Workbook()
ws = wb.active

# Headers
ws.append(['Stok Kod', 'Stok Adı', 'Adet', 'Seri Sayım'])

# Main product row
ws.append([754, 'POS KARBON', 8, 5])

# Serial rows (with numbering like "1 ST87088")
ws.append([None, '1 ST87088', None, None])
ws.append([None, '2 ST87083', None, None])
ws.append([None, '3 ST87081', None, None])
ws.append([None, '4 ST87087', None, None])
ws.append([None, '5 ST87082', None, None])

# Save
test_file = 'test_nested_serials.xlsx'
wb.save(test_file)
print(f"✓ Created {test_file}")

# Now test parsing
wb = load_workbook(test_file, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
print(f"\n📊 Total rows: {len(rows)}")

for i, row in enumerate(rows):
    print(f"Row {i}: {row}")

# Test the parsing logic
print("\n🔍 PARSING TEST:")
print("-" * 50)

headers = rows[0]
print(f"Headers: {headers}")

i = 1
parsed_items = []

while i < len(rows):
    row = rows[i]
    stok_kod = row[0]
    
    if stok_kod:  # Main product row
        stok_adi = str(row[1]).strip() if row[1] else ''
        adet = row[2]
        print(f"\n✓ Found product: {stok_kod} - {stok_adi} (Adet: {adet})")
        
        i += 1
        serial_count = 0
        
        # Collect child rows (serials)
        while i < len(rows):
            child_row = rows[i]
            child_stok_kod = child_row[0]
            
            if child_stok_kod:  # Next product started
                break
            
            # This is a serial row
            serial_value = str(child_row[1]).strip() if child_row[1] else ''
            
            if serial_value:
                # Extract actual serial number (remove numbering like "1 ")
                parts = serial_value.split(maxsplit=1)
                if len(parts) == 2 and parts[0].isdigit():
                    actual_serial = parts[1]  # "ST87088"
                else:
                    actual_serial = serial_value  # Use as-is
                
                parsed_items.append({
                    'stok_kod': stok_kod,
                    'stok_adi': stok_adi,
                    'seri_no': actual_serial
                })
                serial_count += 1
                print(f"  └─ Serial {serial_count}: {actual_serial}")
            
            i += 1
    else:
        i += 1

print(f"\n✅ RESULT: Parsed {len(parsed_items)} serials")
for item in parsed_items:
    print(f"  {item['stok_kod']} | {item['stok_adi']} | {item['seri_no']}")

print("\n" + "="*50)
if len(parsed_items) == 5:
    print("✅ TEST PASSED: All 5 serials parsed correctly!")
else:
    print(f"❌ TEST FAILED: Expected 5 serials, got {len(parsed_items)}")
